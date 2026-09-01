import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.utils import get_column_letter, column_index_from_string
import datetime
import re
import os
import sys
import shutil

sys.stdout.reconfigure(encoding='utf-8')

# Target file
TARGET_FILE = "17-08-2026 13.58.57 - Yearly Event Tracker- DT-3 and DT-4 1.xlsx"
BACKUP_FILE = "17-08-2026 13.58.57 - Yearly Event Tracker- DT-3 and DT-4 1_Backup.xlsx"

# Ensure we start from the clean backup
if os.path.exists(BACKUP_FILE):
    print(f"Restoring clean baseline from: {BACKUP_FILE} ...")
    shutil.copyfile(BACKUP_FILE, TARGET_FILE)
else:
    print(f"Creating baseline backup: {BACKUP_FILE} ...")
    shutil.copyfile(TARGET_FILE, BACKUP_FILE)

print(f"Loading workbook: {TARGET_FILE} ...")
wb = openpyxl.load_workbook(TARGET_FILE)

# Styling Definitions
FONT_FAMILY = "Segoe UI"
REGULAR_FONT = Font(name=FONT_FAMILY, size=10, color='000000')
BOLD_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color='000000')
HEADER_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color='FFFFFF')
BANNER_TITLE_FONT = Font(name=FONT_FAMILY, size=11, bold=True, color='FFFFFF')
HOME_BTN_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color='FFFFFF', underline='single')

FONT_MAIN_TITLE = Font(name=FONT_FAMILY, size=15, bold=True, color='FFFFFF')
FONT_SUB_TITLE = Font(name=FONT_FAMILY, size=10, bold=False, color='DCE6F1')
FONT_SECTION_HEADER = Font(name=FONT_FAMILY, size=11, bold=True, color='FFFFFF')
FONT_TABLE_HEADER = Font(name=FONT_FAMILY, size=10, bold=True, color='FFFFFF')
FONT_CARD_TITLE = Font(name=FONT_FAMILY, size=9, bold=True, color='595959')
FONT_CARD_SUB = Font(name=FONT_FAMILY, size=8, bold=False, color='7F7F7F')
FONT_BOLD_NAVY = Font(name=FONT_FAMILY, size=10, bold=True, color='1B365D')
FONT_LINK = Font(name=FONT_FAMILY, size=10, bold=True, color='0055AA', underline='single')

FONT_KPI_NAVY = Font(name=FONT_FAMILY, size=18, bold=True, color='1B365D')
FONT_KPI_WHITE = Font(name=FONT_FAMILY, size=18, bold=True, color='FFFFFF')
FONT_KPI_BLACK = Font(name=FONT_FAMILY, size=18, bold=True, color='000000')
FONT_KPI_BLUE = Font(name=FONT_FAMILY, size=18, bold=True, color='1F4E79')

NAVY_DARK_FILL = PatternFill(start_color='1B365D', end_color='1B365D', fill_type='solid')
NAVY_LIGHT_FILL = PatternFill(start_color='2C4D75', end_color='2C4D75', fill_type='solid')
FILL_NAVY_SLATE = PatternFill(start_color='2C4D75', end_color='2C4D75', fill_type='solid')
FILL_ROYAL_BLUE = PatternFill(start_color='204A87', end_color='204A87', fill_type='solid')
FILL_BANNER_DARK = PatternFill(start_color='0F2537', end_color='0F2537', fill_type='solid')
FILL_CARD_GRAY = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
FILL_TOTAL_ROW = PatternFill(start_color='E9EEF4', end_color='E9EEF4', fill_type='solid')
FILL_TEAL_BTN = PatternFill(start_color='008080', end_color='008080', fill_type='solid')
FILL_PURPLE_CARD = PatternFill(start_color='4A69BD', end_color='4A69BD', fill_type='solid')
FILL_DROPDOWN = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
FILL_SOFT_GREEN = PatternFill(start_color='E2F0D9', end_color='E2F0D9', fill_type='solid')

# Status Colors
GREEN_FILL = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
GREEN_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color='FFFFFF')

YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
YELLOW_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color='000000')

RED_FILL = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
RED_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color='FFFFFF')

# Feedback Form Colors
FEEDBACK_SENT_FILL = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
FEEDBACK_SENT_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color='1F4E79')
NO_FILL = PatternFill(fill_type=None)

# Full Solid Black Grid Border
BLACK_THIN = Side(style='thin', color='000000')
BLACK_BORDER = Border(left=BLACK_THIN, right=BLACK_THIN, top=BLACK_THIN, bottom=BLACK_THIN)
ALIGN_CENTER_WRAP = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)


def style_and_merge(ws, start_col, end_col, val, font, fill, hyperlink_loc=None):
    for c in range(start_col, end_col + 1):
        cell = ws.cell(1, c)
        cell.font = font
        cell.fill = fill
        cell.border = BLACK_BORDER
        cell.alignment = ALIGN_CENTER_WRAP
        cell.value = None
    c_start = ws.cell(1, start_col)
    c_start.value = val
    if hyperlink_loc:
        c_start.hyperlink = Hyperlink(ref=f"{get_column_letter(start_col)}1", location=hyperlink_loc)
    if end_col > start_col:
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)


# =============================================================================
# 1. ADJUST HEADINGS (ROW 1 BANNER) & FEEDBACK DROPDOWNS ACROSS ALL SHEETS
# =============================================================================
print("Updating Row 1 banners, table-adjusted merging, and Feedback dropdowns...")

for sname in wb.sheetnames:
    if sname == 'Dashboard':
        continue
    ws = wb[sname]
    max_c = ws.max_column
    max_r = ws.max_row
    if max_r < 2 or max_c < 2:
        continue

    # Unmerge any existing Row 1 merged cells
    r1_merged = [m for m in list(ws.merged_cells.ranges) if m.min_row <= 1 <= m.max_row]
    for m in r1_merged:
        ws.unmerge_cells(str(m))

    ws.row_dimensions[1].height = 28

    # Determine sheet title and split columns
    if sname == "Master_Event_Tracker_2026":
        style_and_merge(ws, 1, 3, "🏠 ⮌ BACK TO DASHBOARD", HOME_BTN_FONT, NAVY_DARK_FILL, "'Dashboard'!A1")
        style_and_merge(ws, 4, 16, "MASTER EVENT OPERATIONS DATABASE • FULL YEAR 2026 (DOWNTOWN-3 & DOWNTOWN-4)", BANNER_TITLE_FONT, NAVY_LIGHT_FILL)
        style_and_merge(ws, 17, 22, "🟢 Green: Confirmed | 🟡 Yellow: Pending | 🔴 Red: Canceled", HEADER_FONT, NAVY_DARK_FILL)
    elif "DT-3" in sname:
        m_tag = sname.replace("DT-3", "").strip().upper()
        style_and_merge(ws, 1, 3, "🏠 ⮌ RETURN TO DASHBOARD", HOME_BTN_FONT, NAVY_DARK_FILL, "'Dashboard'!A1")
        style_and_merge(ws, 4, 12, f"DOWNTOWN-3 EVENT OPERATIONS TRACKER • {m_tag}", BANNER_TITLE_FONT, NAVY_LIGHT_FILL)
        style_and_merge(ws, 13, 17, "🟢 Green: Confirmed | 🟡 Yellow: Pending | 🔴 Red: Canceled", HEADER_FONT, NAVY_DARK_FILL)
    elif "DT-4" in sname:
        m_tag = sname.replace("DT-4", "").strip().upper()
        style_and_merge(ws, 1, 3, "🏠 ⮌ RETURN TO DASHBOARD", HOME_BTN_FONT, NAVY_DARK_FILL, "'Dashboard'!A1")
        style_and_merge(ws, 4, 13, f"DOWNTOWN-4 EVENT OPERATIONS TRACKER • {m_tag}", BANNER_TITLE_FONT, NAVY_LIGHT_FILL)
        style_and_merge(ws, 14, 18, "🟢 Green: Confirmed | 🟡 Yellow: Pending | 🔴 Red: Canceled", HEADER_FONT, NAVY_DARK_FILL)
    else:
        # Historical sheets
        if max_c <= 15:
            style_and_merge(ws, 1, 3, "🏠 ⮌ RETURN TO DASHBOARD", HOME_BTN_FONT, NAVY_DARK_FILL, "'Dashboard'!A1")
            style_and_merge(ws, 4, 10, f"EVENT OPERATIONS TRACKER • {sname.upper()}", BANNER_TITLE_FONT, NAVY_LIGHT_FILL)
            style_and_merge(ws, 11, max_c, "🟢 Green: Confirmed | 🟡 Yellow: Pending | 🔴 Red: Canceled", HEADER_FONT, NAVY_DARK_FILL)
        else:
            style_and_merge(ws, 1, 3, "🏠 ⮌ RETURN TO DASHBOARD", HOME_BTN_FONT, NAVY_DARK_FILL, "'Dashboard'!A1")
            style_and_merge(ws, 4, 11, f"EVENT OPERATIONS TRACKER • {sname.upper()}", BANNER_TITLE_FONT, NAVY_LIGHT_FILL)
            style_and_merge(ws, 12, max_c, "🟢 Green: Confirmed | 🟡 Yellow: Pending | 🔴 Red: Canceled", HEADER_FONT, NAVY_DARK_FILL)

    # Check Feedback Form column and Status column
    fb_col_idx = None
    status_col_idx = None
    for c in range(1, max_c + 1):
        h_val = str(ws.cell(2, c).value or '').strip().lower()
        if 'feedback' in h_val:
            fb_col_idx = c
        elif 'status' in h_val:
            status_col_idx = c

    # Ensure Row 2 headers are formatted
    ws.row_dimensions[2].height = 28
    for c in range(1, max_c + 1):
        cell = ws.cell(2, c)
        cell.font = HEADER_FONT
        cell.fill = NAVY_DARK_FILL
        cell.border = BLACK_BORDER
        cell.alignment = ALIGN_CENTER_WRAP

    # Update Data Validations and cell data for Status & Feedback
    if status_col_idx:
        s_col_letter = get_column_letter(status_col_idx)
        # Check if status validation already exists
        has_status_dv = any(s_col_letter in str(dv.sqref) for dv in ws.data_validations.dataValidation)
        if not has_status_dv:
            dv_status = DataValidation(type="list", formula1='"Confirmed (Done),Pending Calendar Booking,Canceled"', allow_blank=True)
            dv_status.add(f"{s_col_letter}3:{s_col_letter}{max_r + 50}")
            ws.add_data_validation(dv_status)

    if fb_col_idx:
        fb_col_letter = get_column_letter(fb_col_idx)
        
        # Remove old data validations covering feedback column
        new_dvs = []
        for dv in list(ws.data_validations.dataValidation):
            if fb_col_letter in str(dv.sqref):
                pass
            else:
                new_dvs.append(dv)
        ws.data_validations.dataValidation = new_dvs

        # Add new Feedback Data Validation: "Form sent,NA"
        dv_fb = DataValidation(type="list", formula1='"Form sent,NA"', allow_blank=True)
        dv_fb.add(f"{fb_col_letter}3:{fb_col_letter}{max_r + 50}")
        ws.add_data_validation(dv_fb)

        # Update cell values and styling in Feedback column
        for r in range(3, max_r + 1):
            cell = ws.cell(r, fb_col_idx)
            val = cell.value
            val_str = str(val or '').strip()
            
            # Check if this row has event data
            row_has_data = any(ws.cell(r, col_i).value is not None and str(ws.cell(r, col_i).value).strip() != '' for col_i in range(1, max_c + 1) if col_i != fb_col_idx)

            if val_str.lower() in ('form sent', 'form sent ', 'sent', 'forms sent'):
                cell.value = "Form sent"
                cell.fill = FEEDBACK_SENT_FILL
                cell.font = FEEDBACK_SENT_FONT
            elif val_str.lower() in ('na', 'n/a', 'n.a.', 'no'):
                cell.value = "NA"
                cell.fill = NO_FILL
                cell.font = REGULAR_FONT
            elif val_str.lower() in ('pending', 'pend', ''):
                if row_has_data:
                    cell.value = "NA"
                    cell.fill = NO_FILL
                    cell.font = REGULAR_FONT
                else:
                    cell.value = None
                    cell.fill = NO_FILL
                    cell.font = REGULAR_FONT
            else:
                if 'sent' in val_str.lower():
                    cell.value = "Form sent"
                    cell.fill = FEEDBACK_SENT_FILL
                    cell.font = FEEDBACK_SENT_FONT
                elif row_has_data:
                    cell.value = "NA"
                    cell.fill = NO_FILL
                    cell.font = REGULAR_FONT

            cell.border = BLACK_BORDER
            cell.alignment = ALIGN_CENTER_WRAP

    # Re-apply status formatting on rows 3..max_r
    if status_col_idx:
        for r in range(3, max_r + 1):
            cell = ws.cell(r, status_col_idx)
            s_val = str(cell.value or '').strip().lower()
            if 'confirm' in s_val or 'done' in s_val:
                cell.fill = GREEN_FILL
                cell.font = GREEN_FONT
            elif 'cancel' in s_val:
                cell.fill = RED_FILL
                cell.font = RED_FONT
            elif 'pending' in s_val:
                cell.fill = YELLOW_FILL
                cell.font = YELLOW_FONT

print("Successfully updated banners and feedback validations in all sheets.")


# =============================================================================
# 2. REBUILD & ENHANCE DASHBOARD WITH AUTOMATIC FEEDBACK FORM METRICS
# =============================================================================
print("Rebuilding Dashboard with automatic month-wise Feedback Form metrics...")

if "Dashboard" in wb.sheetnames:
    del wb["Dashboard"]

ws_dash = wb.create_sheet("Dashboard", 0)

def style_dash_cell(ws, r, c, val=None, font=None, fill=None, border=None, align=None, num_format=None, hyperlink_loc=None):
    cell = ws.cell(r, c)
    if val is not None: cell.value = val
    if font: cell.font = font
    if fill: cell.fill = fill
    if border: cell.border = border
    if align: cell.alignment = align
    if num_format: cell.number_format = num_format
    if hyperlink_loc:
        col_letter = get_column_letter(c)
        cell.hyperlink = Hyperlink(ref=f"{col_letter}{r}", location=hyperlink_loc)
    return cell

def style_dash_range(ws, min_r, min_c, max_r, max_c, font=None, fill=None, border=None, align=None, num_format=None):
    for r in range(min_r, max_r + 1):
        for c in range(min_c, max_c + 1):
            cell = ws.cell(r, c)
            if font: cell.font = font
            if fill: cell.fill = fill
            if border: cell.border = border
            if align: cell.alignment = align
            if num_format: cell.number_format = num_format

# Row 1 & 2: Header Banner
ws_dash.row_dimensions[1].height = 32
ws_dash.row_dimensions[2].height = 20
ws_dash.merge_cells('A1:S1')
ws_dash.merge_cells('A2:S2')

style_dash_cell(ws_dash, 1, 1, "🏢 CBRE | YEARLY EVENT OPERATIONS DASHBOARD", font=FONT_MAIN_TITLE, fill=FILL_BANNER_DARK, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 1, 1, 1, 19, fill=FILL_BANNER_DARK)

style_dash_cell(ws_dash, 2, 1, "Downtown-3 & Downtown-4 Facilities • 2026 Event Planning, Capacity Analytics & Operations Management", font=FONT_SUB_TITLE, fill=NAVY_DARK_FILL, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 2, 1, 2, 19, fill=NAVY_DARK_FILL)

# Rows 4 & 5: Operations Calendar Policy & Legend
ws_dash.row_dimensions[3].height = 8
ws_dash.row_dimensions[4].height = 20
ws_dash.row_dimensions[5].height = 26
ws_dash.merge_cells('A4:S4')
style_dash_cell(ws_dash, 4, 1, "OPERATIONS CALENDAR BOOKING POLICY & STATUS GUIDE:", font=FONT_BOLD_NAVY, align=ALIGN_LEFT_WRAP)

ws_dash.merge_cells('B5:E5')
style_dash_cell(ws_dash, 5, 2, "🟢 GREEN: All Done with Calendar Booking", font=FONT_TABLE_HEADER, fill=GREEN_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 5, 2, 5, 5, fill=GREEN_FILL, border=BLACK_BORDER)

ws_dash.merge_cells('G5:K5')
style_dash_cell(ws_dash, 5, 7, "🟡 YELLOW: Calendar Booking Pending (Action Required)", font=YELLOW_FONT, fill=YELLOW_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 5, 7, 5, 11, fill=YELLOW_FILL, border=BLACK_BORDER)

ws_dash.merge_cells('M5:P5')
style_dash_cell(ws_dash, 5, 13, "🔴 RED: Event Canceled by Organizer", font=RED_FONT, fill=RED_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 5, 13, 5, 16, fill=RED_FILL, border=BLACK_BORDER)

ws_dash.merge_cells('R5:S5')
style_dash_cell(ws_dash, 5, 18, "📑 Open Master Tracker ↗", font=FONT_TABLE_HEADER, fill=FILL_TEAL_BTN, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, hyperlink_loc="'Master_Event_Tracker_2026'!A1")
style_dash_range(ws_dash, 5, 18, 5, 19, fill=FILL_TEAL_BTN, border=BLACK_BORDER)

ws_dash.row_dimensions[6].height = 8

# Rows 7 to 9: Executive Summary KPI Cards (spanning B to S)
ws_dash.row_dimensions[7].height = 18
ws_dash.row_dimensions[8].height = 32
ws_dash.row_dimensions[9].height = 16

cards_cfg = [
    ('B', 'C', 'TOTAL EVENTS (2026)', '=B25+I28', 'DT-3 & DT-4 combined', FILL_CARD_GRAY, FONT_CARD_TITLE, FONT_KPI_NAVY, FONT_CARD_SUB),
    ('D', 'E', 'CONFIRMED (DONE)', '=C25+J28', '🟢 Calendar Booked', GREEN_FILL, FONT_TABLE_HEADER, FONT_KPI_WHITE, FONT_TABLE_HEADER),
    ('F', 'G', 'BOOKING PENDING', '=D25+K28', '🟡 Action Required', YELLOW_FILL, YELLOW_FONT, FONT_KPI_BLACK, YELLOW_FONT),
    ('H', 'I', 'CANCELED EVENTS', '=E25+L28', '🔴 Canceled Bookings', RED_FILL, FONT_TABLE_HEADER, FONT_KPI_WHITE, FONT_TABLE_HEADER),
    ('J', 'K', 'TOTAL ATTENDEES', '=F25+M28', '👥 Expected Pax', FILL_PURPLE_CARD, FONT_TABLE_HEADER, FONT_KPI_WHITE, FONT_TABLE_HEADER),
    ('L', 'M', 'DOWNTOWN-3 EVENTS', '=B25', '🏢 DT-3 Facility', FILL_CARD_GRAY, FONT_CARD_TITLE, FONT_KPI_NAVY, FONT_CARD_SUB),
    ('N', 'O', 'DOWNTOWN-4 EVENTS', '=I28', '🏢 DT-4 Facility', FILL_CARD_GRAY, FONT_CARD_TITLE, FONT_KPI_NAVY, FONT_CARD_SUB),
    ('P', 'Q', '📨 TOTAL FEEDBACK SENT', '=R25', 'Form Sent (2026)', FEEDBACK_SENT_FILL, Font(name=FONT_FAMILY, size=9, bold=True, color='1F4E79'), FONT_KPI_BLUE, Font(name=FONT_FAMILY, size=8, bold=True, color='1F4E79')),
    ('R', 'S', '⚪ TOTAL FEEDBACK NA', '=S25', 'Feedback NA (2026)', FILL_CARD_GRAY, FONT_CARD_TITLE, FONT_KPI_NAVY, FONT_CARD_SUB),
]

for c1, c2, title, formula, sub, fill, f_title, f_kpi, f_sub in cards_cfg:
    c1_idx = column_index_from_string(c1)
    c2_idx = column_index_from_string(c2)
    ws_dash.merge_cells(f'{c1}7:{c2}7')
    ws_dash.merge_cells(f'{c1}8:{c2}8')
    ws_dash.merge_cells(f'{c1}9:{c2}9')
    style_dash_range(ws_dash, 7, c1_idx, 9, c2_idx, fill=fill, border=BLACK_BORDER)
    style_dash_cell(ws_dash, 7, c1_idx, title, font=f_title, align=ALIGN_CENTER_WRAP)
    style_dash_cell(ws_dash, 8, c1_idx, formula, font=f_kpi, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, 9, c1_idx, sub, font=f_sub, align=ALIGN_CENTER_WRAP)

ws_dash.row_dimensions[10].height = 10

# Rows 11 to 28: Three Side-by-Side Operational Navigation & Feedback Tracking Tables
ws_dash.row_dimensions[11].height = 26
ws_dash.row_dimensions[12].height = 24

# Table 1: DT-3 Monthly Tracker (Cols A to F)
ws_dash.merge_cells('A11:F11')
style_dash_cell(ws_dash, 11, 1, "🏢 DOWNTOWN-3 (DT-3) MONTHLY EVENT TRACKER & NAVIGATION", font=FONT_SECTION_HEADER, fill=FILL_ROYAL_BLUE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 11, 1, 11, 6, fill=FILL_ROYAL_BLUE, border=BLACK_BORDER)

dt3_headers = ['Month (Click to Open)', 'Total Events', '🟢 Confirmed', '🟡 Pending', '🔴 Canceled', 'Total Pax']
for idx, h in enumerate(dt3_headers, 1):
    style_dash_cell(ws_dash, 12, idx, h, font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)

# Table 2: DT-4 Monthly Tracker (Cols H to M)
ws_dash.merge_cells('H11:M11')
style_dash_cell(ws_dash, 11, 8, "🏢 DOWNTOWN-4 (DT-4) MONTHLY EVENT TRACKER & NAVIGATION", font=FONT_SECTION_HEADER, fill=NAVY_DARK_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 11, 8, 11, 13, fill=NAVY_DARK_FILL, border=BLACK_BORDER)

dt4_headers = ['Month (Click to Open)', 'Total Events', '🟢 Confirmed', '🟡 Pending', '🔴 Canceled', 'Total Pax']
for idx, h in enumerate(dt4_headers, 8):
    style_dash_cell(ws_dash, 12, idx, h, font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)

# Table 3: Dedicated Feedback Form Tracker (Cols O to S)
ws_dash.merge_cells('O11:S11')
style_dash_cell(ws_dash, 11, 15, "📋 FEEDBACK FORM TRACKER (2026 MONTH-WISE)", font=FONT_SECTION_HEADER, fill=FILL_ROYAL_BLUE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 11, 15, 11, 19, fill=FILL_ROYAL_BLUE, border=BLACK_BORDER)

fb_headers = ['Month', 'DT-3 Sent', 'DT-4 Sent', '📨 Total Sent', '⚪ Total NA']
for idx, h in enumerate(fb_headers, 15):
    style_dash_cell(ws_dash, 12, idx, h, font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)

months_26 = ['Jan26', 'Feb26', 'Mar26', 'Apr26', 'May26', 'Jun26', 'Jul26', 'Aug26', 'Sep26', 'Oct26', 'Nov26', 'Dec26']
month_labels = ['Jan 2026 ↗', 'Feb 2026 ↗', 'Mar 2026 ↗', 'Apr 2026 ↗', 'May 2026 ↗', 'Jun 2026 ↗', 'Jul 2026 ↗', 'Aug 2026 ↗', 'Sep 2026 ↗', 'Oct 2026 ↗', 'Nov 2026 ↗', 'Dec 2026 ↗']

for idx, (m_tag, m_lbl) in enumerate(zip(months_26, month_labels)):
    r = 13 + idx
    ws_dash.row_dimensions[r].height = 20
    s_dt3 = f"DT-3 {m_tag}"
    s_dt4 = f"DT-4 {m_tag}"

    # DT-3 Table (Cols A to F)
    style_dash_cell(ws_dash, r, 1, m_lbl, font=FONT_LINK, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, hyperlink_loc=f"'{s_dt3}'!A1")
    style_dash_cell(ws_dash, r, 2, f'=COUNTIF(\'{s_dt3}\'!$N$3:$N$500, "<>")', font=BOLD_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 3, f'=COUNTIF(\'{s_dt3}\'!$N$3:$N$500, "Confirmed (Done)")', font=GREEN_FONT, fill=GREEN_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 4, f'=COUNTIF(\'{s_dt3}\'!$N$3:$N$500, "Pending Calendar Booking")', font=YELLOW_FONT, fill=YELLOW_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 5, f'=COUNTIF(\'{s_dt3}\'!$N$3:$N$500, "Canceled")', font=RED_FONT, fill=RED_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 6, f'=SUM(\'{s_dt3}\'!$F$3:$F$500)', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    
    # DT-4 Table (Cols H to M)
    style_dash_cell(ws_dash, r, 8, m_lbl, font=FONT_LINK, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, hyperlink_loc=f"'{s_dt4}'!A1")
    style_dash_cell(ws_dash, r, 9, f'=COUNTIF(\'{s_dt4}\'!$O$3:$O$500, "<>")', font=BOLD_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 10, f'=COUNTIF(\'{s_dt4}\'!$O$3:$O$500, "Confirmed (Done)")', font=GREEN_FONT, fill=GREEN_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 11, f'=COUNTIF(\'{s_dt4}\'!$O$3:$O$500, "Pending Calendar Booking")', font=YELLOW_FONT, fill=YELLOW_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 12, f'=COUNTIF(\'{s_dt4}\'!$O$3:$O$500, "Canceled")', font=RED_FONT, fill=RED_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 13, f'=SUM(\'{s_dt4}\'!$G$3:$G$500)', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

    # Feedback Form Month-Wise Table (Cols O to S)
    # Col O: Month Name
    style_dash_cell(ws_dash, r, 15, m_lbl.replace(" ↗", ""), font=BOLD_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
    # Col P: DT-3 Form Sent (Col O in DT-3)
    style_dash_cell(ws_dash, r, 16, f'=COUNTIF(\'{s_dt3}\'!$O$3:$O$500, "Form sent")', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    # Col Q: DT-4 Form Sent (Col P in DT-4)
    style_dash_cell(ws_dash, r, 17, f'=COUNTIF(\'{s_dt4}\'!$P$3:$P$500, "Form sent")', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    # Col R: Total Sent
    style_dash_cell(ws_dash, r, 18, f'=$P{r}+$Q{r}', font=FEEDBACK_SENT_FONT, fill=FEEDBACK_SENT_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    # Col S: Total NA
    style_dash_cell(ws_dash, r, 19, f'=COUNTIF(\'{s_dt3}\'!$O$3:$O$500, "NA")+COUNTIF(\'{s_dt4}\'!$P$3:$P$500, "NA")', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

# DT-3 Total (Row 25)
ws_dash.row_dimensions[25].height = 22
style_dash_cell(ws_dash, 25, 1, "DT-3 Full Year Total", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_cell(ws_dash, 25, 2, "=SUM(B13:B24)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 25, 3, "=SUM(C13:C24)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 25, 4, "=SUM(D13:D24)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 25, 5, "=SUM(E13:E24)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 25, 6, "=SUM(F13:F24)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

# Feedback Full Year Total (Row 25 in Cols O to S)
style_dash_cell(ws_dash, 25, 15, "Full Year Total", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_cell(ws_dash, 25, 16, "=SUM(P13:P24)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 25, 17, "=SUM(Q13:Q24)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 25, 18, "=SUM(R13:R24)", font=FEEDBACK_SENT_FONT, fill=FEEDBACK_SENT_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 25, 19, "=SUM(S13:S24)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

# DT-4 2027 Rows (Rows 25-27 in Cols H to M)
months_27 = [('Jan27', "Jan '27 ↗"), ('Feb27', "Feb '27 ↗"), ('Mar27', "Mar '27 ↗")]
for idx, (m_tag, m_lbl) in enumerate(months_27):
    r = 25 + idx
    s_dt4_27 = f"DT-4 {m_tag}"
    style_dash_cell(ws_dash, r, 8, m_lbl, font=FONT_LINK, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, hyperlink_loc=f"'{s_dt4_27}'!A1")
    style_dash_cell(ws_dash, r, 9, f'=COUNTIF(\'{s_dt4_27}\'!$O$3:$O$500, "<>")', font=BOLD_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 10, f'=COUNTIF(\'{s_dt4_27}\'!$O$3:$O$500, "Confirmed (Done)")', font=GREEN_FONT, fill=GREEN_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 11, f'=COUNTIF(\'{s_dt4_27}\'!$O$3:$O$500, "Pending Calendar Booking")', font=YELLOW_FONT, fill=YELLOW_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 12, f'=COUNTIF(\'{s_dt4_27}\'!$O$3:$O$500, "Canceled")', font=RED_FONT, fill=RED_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 13, f'=SUM(\'{s_dt4_27}\'!$G$3:$G$500)', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

# DT-4 Subtotal (Row 28)
ws_dash.row_dimensions[28].height = 22
style_dash_cell(ws_dash, 28, 8, "DT-4 2026 Subtotal", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_cell(ws_dash, 28, 9, "=SUM(I13:I24)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 28, 10, "=SUM(J13:J24)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 28, 11, "=SUM(K13:K24)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 28, 12, "=SUM(L13:L24)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 28, 13, "=SUM(M13:M27)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

# Rows 27 to 30: Executive Summary Mini Boxes (Cols O-P and Cols R-S)
# Left Mini-Box: Feedback Form Summary (Cols O-P)
style_dash_cell(ws_dash, 27, 15, "Feedback Metric", font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_cell(ws_dash, 27, 16, "Summary", font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)

style_dash_cell(ws_dash, 28, 15, "📨 Total Form Sent", font=FEEDBACK_SENT_FONT, fill=FEEDBACK_SENT_FILL, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
style_dash_cell(ws_dash, 28, 16, "=R25", font=FEEDBACK_SENT_FONT, fill=FEEDBACK_SENT_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

style_dash_cell(ws_dash, 29, 15, "⚪ Total Form NA", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
style_dash_cell(ws_dash, 29, 16, "=S25", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

style_dash_cell(ws_dash, 30, 15, "Feedback Sent Rate", font=Font(name=FONT_FAMILY, size=10, bold=True, color='276A3C'), fill=FILL_SOFT_GREEN, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
style_dash_cell(ws_dash, 30, 16, "=IFERROR(R25/(R25+S25), 0)", font=Font(name=FONT_FAMILY, size=10, bold=True, color='276A3C'), fill=FILL_SOFT_GREEN, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='0.0%')

# Right Mini-Box: Status Distribution Mini Box (Cols R-S)
style_dash_cell(ws_dash, 27, 18, "Status", font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_cell(ws_dash, 27, 19, "Count", font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_cell(ws_dash, 28, 18, "Confirmed (Done)", font=GREEN_FONT, fill=GREEN_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_cell(ws_dash, 28, 19, "=E45", font=GREEN_FONT, fill=GREEN_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 29, 18, "Booking Pending", font=YELLOW_FONT, fill=YELLOW_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_cell(ws_dash, 29, 19, "=F45", font=YELLOW_FONT, fill=YELLOW_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 30, 18, "Canceled", font=RED_FONT, fill=RED_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_cell(ws_dash, 30, 19, "=G45", font=RED_FONT, fill=RED_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

# Combined Monthly Events Summary (Left: Cols A to I)
ws_dash.row_dimensions[31].height = 26
ws_dash.row_dimensions[32].height = 24

ws_dash.merge_cells('A31:I31')
style_dash_cell(ws_dash, 31, 1, "📊 2026 COMBINED MONTHLY EVENTS SUMMARY", font=FONT_SECTION_HEADER, fill=NAVY_DARK_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 31, 1, 31, 9, fill=NAVY_DARK_FILL, border=BLACK_BORDER)

comb_headers = ['Month', 'DT-3 Events', 'DT-4 Events', 'Total Events', '🟢 Confirmed', '🟡 Pending', '🔴 Canceled', 'Confirmed %', 'Total Pax']
for idx, h in enumerate(comb_headers, 1):
    style_dash_cell(ws_dash, 32, idx, h, font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)

months_short = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
for idx, m_short in enumerate(months_short):
    r = 33 + idx
    ref_r = 13 + idx
    ws_dash.row_dimensions[r].height = 20
    style_dash_cell(ws_dash, r, 1, m_short, font=BOLD_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
    style_dash_cell(ws_dash, r, 2, f'=B{ref_r}', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 3, f'=I{ref_r}', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 4, f'=$B{r}+$C{r}', font=BOLD_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 5, f'=C{ref_r}+J{ref_r}', font=GREEN_FONT, fill=GREEN_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 6, f'=D{ref_r}+K{ref_r}', font=YELLOW_FONT, fill=YELLOW_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 7, f'=E{ref_r}+L{ref_r}', font=RED_FONT, fill=RED_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 8, f'=IFERROR($E{r}/($D{r}-$G{r}), 0)', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='0.0%')
    style_dash_cell(ws_dash, r, 9, f'=F{ref_r}+M{ref_r}', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

# Full Year Total (Row 45)
ws_dash.row_dimensions[45].height = 22
style_dash_cell(ws_dash, 45, 1, "Full Year 2026 Total", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_cell(ws_dash, 45, 2, "=SUM(B33:B44)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 45, 3, "=SUM(C33:C44)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 45, 4, "=SUM(D33:D44)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 45, 5, "=SUM(E33:E44)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 45, 6, "=SUM(F33:F44)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 45, 7, "=SUM(G33:G44)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 45, 8, "=IFERROR(E45/(D45-G45), 0)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='0.0%')
style_dash_cell(ws_dash, 45, 9, "=SUM(I33:I44)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

# Interactive Month Selection & Deep Dive (Right: Cols K to S)
ws_dash.merge_cells('K31:S31')
style_dash_cell(ws_dash, 31, 11, "🔍 INTERACTIVE MONTH SELECTION & DEEP DIVE", font=FONT_SECTION_HEADER, fill=FILL_ROYAL_BLUE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 31, 11, 31, 19, fill=FILL_ROYAL_BLUE, border=BLACK_BORDER)

ws_dash.merge_cells('K32:L32')
style_dash_cell(ws_dash, 32, 11, "📅 SELECT MONTH:", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 32, 11, 32, 12, fill=FILL_TOTAL_ROW, border=BLACK_BORDER)

ws_dash.merge_cells('M32:O32')
style_dash_cell(ws_dash, 32, 13, "Jan", font=Font(name=FONT_FAMILY, size=12, bold=True, color='1B365D'), fill=FILL_DROPDOWN, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 32, 13, 32, 15, fill=FILL_DROPDOWN, border=BLACK_BORDER)

ws_dash.merge_cells('P32:S32')
style_dash_cell(ws_dash, 32, 16, "🟢 Live Auto-Calculated Drilldown", font=Font(name=FONT_FAMILY, size=9, bold=True, color='276A3C'), fill=FILL_SOFT_GREEN, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 32, 16, 32, 19, fill=FILL_SOFT_GREEN, border=BLACK_BORDER)

dv_month = DataValidation(type="list", formula1='"Full Year 2026,Jan,Feb,Mar,Apr,May,Jun,Jul,Aug,Sep,Oct,Nov,Dec"', allow_blank=False)
dv_month.add("M32:O32")
ws_dash.add_data_validation(dv_month)

ws_dash.row_dimensions[33].height = 14
ws_dash.row_dimensions[34].height = 24
ws_dash.row_dimensions[35].height = 14
ws_dash.row_dimensions[36].height = 24
ws_dash.row_dimensions[37].height = 22

# Metric Cards Row 1
ws_dash.merge_cells('K33:L33')
style_dash_cell(ws_dash, 33, 11, "TOTAL EVENTS", font=FONT_CARD_TITLE, fill=FILL_CARD_GRAY, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 33, 11, 33, 12, fill=FILL_CARD_GRAY, border=BLACK_BORDER)

ws_dash.merge_cells('K34:L34')
style_dash_cell(ws_dash, 34, 11, '=IF($M$32="Full Year 2026", D45, IFERROR(INDEX($D$33:$D$44, MATCH($M$32, $A$33:$A$44, 0)), 0))', font=Font(name=FONT_FAMILY, size=14, bold=True, color='1B365D'), fill=FILL_CARD_GRAY, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_range(ws_dash, 34, 11, 34, 12, fill=FILL_CARD_GRAY, border=BLACK_BORDER)

ws_dash.merge_cells('M33:O33')
style_dash_cell(ws_dash, 33, 13, "CONFIRMED (DONE)", font=FONT_TABLE_HEADER, fill=GREEN_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 33, 13, 33, 15, fill=GREEN_FILL, border=BLACK_BORDER)

ws_dash.merge_cells('M34:O34')
style_dash_cell(ws_dash, 34, 13, '=IF($M$32="Full Year 2026", E45, IFERROR(INDEX($E$33:$E$44, MATCH($M$32, $A$33:$A$44, 0)), 0))', font=Font(name=FONT_FAMILY, size=14, bold=True, color='FFFFFF'), fill=GREEN_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_range(ws_dash, 34, 13, 34, 15, fill=GREEN_FILL, border=BLACK_BORDER)

ws_dash.merge_cells('P33:R33')
style_dash_cell(ws_dash, 33, 16, "PENDING BOOKING", font=YELLOW_FONT, fill=YELLOW_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 33, 16, 33, 18, fill=YELLOW_FILL, border=BLACK_BORDER)

ws_dash.merge_cells('P34:R34')
style_dash_cell(ws_dash, 34, 16, '=IF($M$32="Full Year 2026", F45, IFERROR(INDEX($F$33:$F$44, MATCH($M$32, $A$33:$A$44, 0)), 0))', font=Font(name=FONT_FAMILY, size=14, bold=True, color='000000'), fill=YELLOW_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_range(ws_dash, 34, 16, 34, 18, fill=YELLOW_FILL, border=BLACK_BORDER)

style_dash_cell(ws_dash, 33, 19, "CANCELED", font=FONT_TABLE_HEADER, fill=RED_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_cell(ws_dash, 34, 19, '=IF($M$32="Full Year 2026", G45, IFERROR(INDEX($G$33:$G$44, MATCH($M$32, $A$33:$A$44, 0)), 0))', font=Font(name=FONT_FAMILY, size=14, bold=True, color='FFFFFF'), fill=RED_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

# Metric Cards Row 2
ws_dash.merge_cells('K35:L35')
style_dash_cell(ws_dash, 35, 11, "CONFIRMED % RATE", font=FONT_CARD_TITLE, fill=FILL_CARD_GRAY, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 35, 11, 35, 12, fill=FILL_CARD_GRAY, border=BLACK_BORDER)

ws_dash.merge_cells('K36:L36')
style_dash_cell(ws_dash, 36, 11, '=IF($M$32="Full Year 2026", H45, IFERROR(INDEX($H$33:$H$44, MATCH($M$32, $A$33:$A$44, 0)), 0))', font=Font(name=FONT_FAMILY, size=13, bold=True, color='1B365D'), fill=FILL_CARD_GRAY, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='0.0%')
style_dash_range(ws_dash, 36, 11, 36, 12, fill=FILL_CARD_GRAY, border=BLACK_BORDER)

ws_dash.merge_cells('M35:O35')
style_dash_cell(ws_dash, 35, 13, "TOTAL ATTENDEES (PAX)", font=FONT_TABLE_HEADER, fill=FILL_PURPLE_CARD, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 35, 13, 35, 15, fill=FILL_PURPLE_CARD, border=BLACK_BORDER)

ws_dash.merge_cells('M36:O36')
style_dash_cell(ws_dash, 36, 13, '=IF($M$32="Full Year 2026", I45, IFERROR(INDEX($I$33:$I$44, MATCH($M$32, $A$33:$A$44, 0)), 0))', font=Font(name=FONT_FAMILY, size=14, bold=True, color='FFFFFF'), fill=FILL_PURPLE_CARD, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_range(ws_dash, 36, 13, 36, 15, fill=FILL_PURPLE_CARD, border=BLACK_BORDER)

ws_dash.merge_cells('P35:Q35')
style_dash_cell(ws_dash, 35, 16, "DT-3 EVENTS", font=FONT_CARD_TITLE, fill=FILL_CARD_GRAY, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 35, 16, 35, 17, fill=FILL_CARD_GRAY, border=BLACK_BORDER)

ws_dash.merge_cells('P36:Q36')
style_dash_cell(ws_dash, 36, 16, '=IF($M$32="Full Year 2026", B45, IFERROR(INDEX($B$33:$B$44, MATCH($M$32, $A$33:$A$44, 0)), 0))', font=Font(name=FONT_FAMILY, size=13, bold=True, color='1B365D'), fill=FILL_CARD_GRAY, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_range(ws_dash, 36, 16, 36, 17, fill=FILL_CARD_GRAY, border=BLACK_BORDER)

ws_dash.merge_cells('R35:S35')
style_dash_cell(ws_dash, 35, 18, "DT-4 EVENTS", font=FONT_CARD_TITLE, fill=FILL_CARD_GRAY, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 35, 18, 35, 19, fill=FILL_CARD_GRAY, border=BLACK_BORDER)

ws_dash.merge_cells('R36:S36')
style_dash_cell(ws_dash, 36, 18, '=IF($M$32="Full Year 2026", C45, IFERROR(INDEX($C$33:$C$44, MATCH($M$32, $A$33:$A$44, 0)), 0))', font=Font(name=FONT_FAMILY, size=13, bold=True, color='1B365D'), fill=FILL_CARD_GRAY, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_range(ws_dash, 36, 18, 36, 19, fill=FILL_CARD_GRAY, border=BLACK_BORDER)

# Metric Row 3: Feedback Form Monthly Live Drilldown (Row 37)
ws_dash.merge_cells('K37:L37')
style_dash_cell(ws_dash, 37, 11, "📨 FEEDBACK SENT:", font=Font(name=FONT_FAMILY, size=9, bold=True, color='1F4E79'), fill=FEEDBACK_SENT_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 37, 11, 37, 12, fill=FEEDBACK_SENT_FILL, border=BLACK_BORDER)

ws_dash.merge_cells('M37:O37')
style_dash_cell(ws_dash, 37, 13, '=IF($M$32="Full Year 2026", R25, IFERROR(INDEX($R$13:$R$24, MATCH($M$32, $A$33:$A$44, 0)), 0))', font=Font(name=FONT_FAMILY, size=13, bold=True, color='1F4E79'), fill=FEEDBACK_SENT_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_range(ws_dash, 37, 13, 37, 15, fill=FEEDBACK_SENT_FILL, border=BLACK_BORDER)

ws_dash.merge_cells('P37:Q37')
style_dash_cell(ws_dash, 37, 16, "⚪ FEEDBACK NA:", font=Font(name=FONT_FAMILY, size=9, bold=True, color='1B365D'), fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 37, 16, 37, 17, fill=FILL_TOTAL_ROW, border=BLACK_BORDER)

ws_dash.merge_cells('R37:S37')
style_dash_cell(ws_dash, 37, 18, '=IF($M$32="Full Year 2026", S25, IFERROR(INDEX($S$13:$S$24, MATCH($M$32, $A$33:$A$44, 0)), 0))', font=Font(name=FONT_FAMILY, size=13, bold=True, color='1B365D'), fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_range(ws_dash, 37, 18, 37, 19, fill=FILL_TOTAL_ROW, border=BLACK_BORDER)

# Navigation Box
ws_dash.merge_cells('K38:S38')
style_dash_cell(ws_dash, 38, 11, "🚀 QUICK NAVIGATION TO MONTHLY TRACKERS", font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 38, 11, 38, 19, fill=FILL_NAVY_SLATE, border=BLACK_BORDER)

ws_dash.merge_cells('K39:S44')
help_text = "💡 HOW TO USE THIS DASHBOARD:\n1. Change the Month dropdown above (M32) to instantly update the Monthly Deep Dive KPIs.\n2. Click any Month in the DT-3 / DT-4 tables (Rows 13-27) to jump directly to that month's sheet.\n3. In any monthly sheet, click '🏠 ⮌ RETURN TO DASHBOARD' (A1) to return back here."
style_dash_cell(ws_dash, 39, 11, help_text, font=Font(name=FONT_FAMILY, size=10, bold=False, color='1B365D'), fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
style_dash_range(ws_dash, 39, 11, 44, 19, fill=FILL_TOTAL_ROW, border=BLACK_BORDER)

ws_dash.row_dimensions[46].height = 10

# Facility Space & Venue Distribution (Rows 47 to 58, Cols A-D)
ws_dash.row_dimensions[47].height = 26
ws_dash.row_dimensions[48].height = 24

ws_dash.merge_cells('A47:D47')
style_dash_cell(ws_dash, 47, 1, "📍 FACILITY SPACE & VENUE UTILIZATION", font=FONT_SECTION_HEADER, fill=NAVY_DARK_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 47, 1, 47, 4, fill=NAVY_DARK_FILL, border=BLACK_BORDER)

venue_headers = ['Key Venue / Space', 'Building', 'Events', 'Total Pax']
for idx, h in enumerate(venue_headers, 1):
    style_dash_cell(ws_dash, 48, idx, h, font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)

venues = [
    ('Innovation Hub', 'Downtown-4', '=COUNTIFS(Master_Event_Tracker_2026!$H$3:$H$5000, "*Innovation*", Master_Event_Tracker_2026!$F$3:$F$5000, "Downtown-4")', '=SUMIFS(Master_Event_Tracker_2026!$K$3:$K$5000, Master_Event_Tracker_2026!$H$3:$H$5000, "*Innovation*", Master_Event_Tracker_2026!$F$3:$F$5000, "Downtown-4")'),
    ('Cafeteria / Townhall Area', 'Downtown-4', '=COUNTIFS(Master_Event_Tracker_2026!$H$3:$H$5000, "*Cafeteria*", Master_Event_Tracker_2026!$F$3:$F$5000, "Downtown-4")', '=SUMIFS(Master_Event_Tracker_2026!$K$3:$K$5000, Master_Event_Tracker_2026!$H$3:$H$5000, "*Cafeteria*", Master_Event_Tracker_2026!$F$3:$F$5000, "Downtown-4")'),
    ('Zone E Breakout', 'Downtown-4', '=COUNTIFS(Master_Event_Tracker_2026!$H$3:$H$5000, "*Zone E*", Master_Event_Tracker_2026!$F$3:$F$5000, "Downtown-4")', '=SUMIFS(Master_Event_Tracker_2026!$K$3:$K$5000, Master_Event_Tracker_2026!$H$3:$H$5000, "*Zone E*", Master_Event_Tracker_2026!$F$3:$F$5000, "Downtown-4")'),
    ('Social Hub Areas', 'Downtown-4', '=COUNTIFS(Master_Event_Tracker_2026!$H$3:$H$5000, "*Social Hub*", Master_Event_Tracker_2026!$F$3:$F$5000, "Downtown-4")', '=SUMIFS(Master_Event_Tracker_2026!$K$3:$K$5000, Master_Event_Tracker_2026!$H$3:$H$5000, "*Social Hub*", Master_Event_Tracker_2026!$F$3:$F$5000, "Downtown-4")'),
    ('Cafeteria (L-5)', 'Downtown-4', '=COUNTIFS(Master_Event_Tracker_2026!$H$3:$H$5000, "*Cafeteria*", Master_Event_Tracker_2026!$F$3:$F$5000, "Downtown-4")', '=SUMIFS(Master_Event_Tracker_2026!$K$3:$K$5000, Master_Event_Tracker_2026!$H$3:$H$5000, "*Cafeteria*", Master_Event_Tracker_2026!$F$3:$F$5000, "Downtown-4")'),
    ('Meeting Rooms (405, 506, VC)', 'Downtown-4', '=COUNTIFS(Master_Event_Tracker_2026!$H$3:$H$5000, "*MR*", Master_Event_Tracker_2026!$F$3:$F$5000, "Downtown-4")', '=SUMIFS(Master_Event_Tracker_2026!$K$3:$K$5000, Master_Event_Tracker_2026!$H$3:$H$5000, "*MR*", Master_Event_Tracker_2026!$F$3:$F$5000, "Downtown-4")'),
    ('Cafeteria Breakout', 'Downtown-3', '=COUNTIFS(Master_Event_Tracker_2026!$H$3:$H$5000, "*Cafeteria*", Master_Event_Tracker_2026!$F$3:$F$5000, "Downtown-3")', '=SUMIFS(Master_Event_Tracker_2026!$K$3:$K$5000, Master_Event_Tracker_2026!$H$3:$H$5000, "*Cafeteria*", Master_Event_Tracker_2026!$F$3:$F$5000, "Downtown-3")'),
    ('Breakout Area', 'Downtown-3', '=COUNTIFS(Master_Event_Tracker_2026!$H$3:$H$5000, "*Breakout*", Master_Event_Tracker_2026!$F$3:$F$5000, "Downtown-3")', '=SUMIFS(Master_Event_Tracker_2026!$K$3:$K$5000, Master_Event_Tracker_2026!$H$3:$H$5000, "*Breakout*", Master_Event_Tracker_2026!$F$3:$F$5000, "Downtown-3")'),
    ('Meeting Rooms (107, 305/306)', 'Downtown-3', '=COUNTIFS(Master_Event_Tracker_2026!$H$3:$H$5000, "*MR*", Master_Event_Tracker_2026!$F$3:$F$5000, "Downtown-3")', '=SUMIFS(Master_Event_Tracker_2026!$K$3:$K$5000, Master_Event_Tracker_2026!$H$3:$H$5000, "*MR*", Master_Event_Tracker_2026!$F$3:$F$5000, "Downtown-3")'),
]

for idx, (v_name, bldg, f_cnt, f_pax) in enumerate(venues):
    r = 49 + idx
    ws_dash.row_dimensions[r].height = 20
    style_dash_cell(ws_dash, r, 1, v_name, font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
    style_dash_cell(ws_dash, r, 2, bldg, font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
    style_dash_cell(ws_dash, r, 3, f_cnt, font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 4, f_pax, font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

ws_dash.merge_cells('A58:B58')
ws_dash.row_dimensions[58].height = 22
style_dash_cell(ws_dash, 58, 1, "Top Venues Subtotal", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 58, 1, 58, 2, fill=FILL_TOTAL_ROW, border=BLACK_BORDER)
style_dash_cell(ws_dash, 58, 3, "=SUM(C49:C57)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 58, 4, "=SUM(D49:D57)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

# Event Category & Purpose Distribution (Rows 47 to 58, Cols F-I)
ws_dash.merge_cells('F47:I47')
style_dash_cell(ws_dash, 47, 6, "🎯 EVENT CATEGORY & PURPOSE DISTRIBUTION", font=FONT_SECTION_HEADER, fill=NAVY_DARK_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 47, 6, 47, 9, fill=NAVY_DARK_FILL, border=BLACK_BORDER)

cat_headers = ['Event Category', 'Total Events', 'Confirmed', 'Total Pax']
for idx, h in enumerate(cat_headers, 6):
    style_dash_cell(ws_dash, 48, idx, h, font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)

categories = [
    'Townhall & All-Hands',
    'Training & Enablement',
    'Induction & Onboarding',
    'Assessment & Hiring',
    'Leadership & VIP Visit',
    'Social & Team Connect',
    'Tech & Innovation',
    'Dry Run & Rehearsal',
    'Team Meeting & Workshop'
]

for idx, cat_name in enumerate(categories):
    r = 49 + idx
    ws_dash.row_dimensions[r].height = 20
    style_dash_cell(ws_dash, r, 6, cat_name, font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
    style_dash_cell(ws_dash, r, 7, f'=COUNTIF(Master_Event_Tracker_2026!$J$3:$J$5000, "{cat_name}")', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 8, f'=COUNTIFS(Master_Event_Tracker_2026!$J$3:$J$5000, "{cat_name}", Master_Event_Tracker_2026!$R$3:$R$5000, "Confirmed (Done)")', font=GREEN_FONT, fill=GREEN_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 9, f'=SUMIF(Master_Event_Tracker_2026!$J$3:$J$5000, "{cat_name}", Master_Event_Tracker_2026!$K$3:$K$5000)', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

ws_dash.row_dimensions[58].height = 22
style_dash_cell(ws_dash, 58, 6, "Total Events Tracked", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_cell(ws_dash, 58, 7, "=SUM(G49:G57)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 58, 8, "=SUM(H49:H57)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 58, 9, "=SUM(I49:I57)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

ws_dash.row_dimensions[59].height = 10
ws_dash.row_dimensions[60].height = 6

# Priority Action Items (Rows 61 to 77)
ws_dash.row_dimensions[61].height = 26
ws_dash.row_dimensions[62].height = 24

ws_dash.merge_cells('A61:I61')
style_dash_cell(ws_dash, 61, 1, "⚡ PRIORITY ACTION ITEMS: UPCOMING PENDING CALENDAR BOOKINGS (REQUIRE CALENDAR INVITE)", font=FONT_SECTION_HEADER, fill=NAVY_DARK_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 61, 1, 61, 9, fill=NAVY_DARK_FILL, border=BLACK_BORDER)

action_headers = ['Event ID', 'Date', 'Day', 'Building', 'Location', 'Event Title', 'Pax', 'Booked By', 'Status']
for idx, h in enumerate(action_headers, 1):
    style_dash_cell(ws_dash, 62, idx, h, font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)

pending_sample = [
    ('EVT-2026-0001', datetime.date(2026, 1, 5), 'Mon', 'Downtown-3', 'MR-305 & 306', 'Induction Session', 35, 'Deshraj'),
    ('EVT-2026-0002', datetime.date(2026, 1, 5), 'Mon', 'Downtown-3', 'Breakout Area', 'Induction Session', 35, 'Deshraj'),
    ('EVT-2026-0003', datetime.date(2026, 1, 10), 'Sat', 'Downtown-3', 'MR-305 & 306', "Driver's Training", 20, 'Prem Chaudhary'),
    ('EVT-2026-0004', datetime.date(2026, 1, 10), 'Sat', 'Downtown-3', 'Hiring drive', 'ZONE C, D and E', 0, 'Manoj Thakral'),
    ('EVT-2026-0005', datetime.date(2026, 1, 12), 'Mon', 'Downtown-3', 'MR-305 & 306', 'Inductin Session', 40, 'Deshraj'),
    ('EVT-2026-0006', datetime.date(2026, 1, 12), 'Mon', 'Downtown-3', 'Breakout Area', 'Inductin Session', 20, 'Deshraj'),
    ('EVT-2026-0007', datetime.date(2026, 1, 14), 'Wed', 'Downtown-3', 'MR-305 & 306', 'Inductin Session', 25, 'Deshraj'),
    ('EVT-2026-0008', datetime.date(2026, 1, 14), 'Wed', 'Downtown-3', 'Breakout Area', 'Scheduled Event', 0, 'Internal Team'),
    ('EVT-2026-0009', datetime.date(2026, 1, 15), 'Thu', 'Downtown-3', 'MR-305 & 306', 'Communicating with Impact (FCO + Nom)', 25, 'Deshraj'),
    ('EVT-2026-0010', datetime.date(2026, 1, 15), 'Thu', 'Downtown-3', 'Cafeteria + Breakout Area', 'Risk Awareness session for Enterprise Data team', 200, 'Maanshu Chugh'),
    ('EVT-2026-0011', datetime.date(2026, 1, 16), 'Fri', 'Downtown-3', 'Cafeteria', 'Culture Immersion NH + Nominations', 50, 'Deshraj'),
    ('EVT-2026-0012', datetime.date(2026, 1, 17), 'Sat', 'Downtown-3', 'Main Space', 'Hiring Drive', 0, 'Mohit Mishra/ Monika Nigam'),
    ('EVT-2026-0013', datetime.date(2026, 1, 19), 'Mon', 'Downtown-3', 'MR-305 & 306', 'Induction Session (perm -305 and 306) contractors- (breakout area)', 70, 'Deshraj'),
    ('EVT-2026-0014', datetime.date(2026, 1, 19), 'Mon', 'Downtown-3', 'Breakout Area', 'Scheduled Event', 0, 'Deshraj'),
    ('EVT-2026-0015', datetime.date(2026, 1, 20), 'Tue', 'Downtown-3', 'MR-305 & 306', 'ALT Workshop 2', 0, 'Deshraj'),
]

for idx, (eid, dt, day, bldg, loc, title, pax, booked) in enumerate(pending_sample):
    r = 63 + idx
    ws_dash.row_dimensions[r].height = 20
    style_dash_cell(ws_dash, r, 1, eid, font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
    style_dash_cell(ws_dash, r, 2, dt, font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='dd-mmm-yyyy')
    style_dash_cell(ws_dash, r, 3, day, font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
    style_dash_cell(ws_dash, r, 4, bldg, font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
    style_dash_cell(ws_dash, r, 5, loc, font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
    style_dash_cell(ws_dash, r, 6, title, font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
    style_dash_cell(ws_dash, r, 7, pax, font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 8, booked, font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
    style_dash_cell(ws_dash, r, 9, "🟡 Pending Booking", font=YELLOW_FONT, fill=YELLOW_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)

# Column Widths for Dashboard
dash_col_widths = {
    'A': 20.0, 'B': 14.0, 'C': 14.0, 'D': 14.0, 'E': 14.0, 'F': 22.0, 'G': 14.0, 'H': 20.0, 'I': 14.0,
    'J': 14.0, 'K': 14.0, 'L': 14.0, 'M': 14.0, 'N': 14.0, 'O': 18.0, 'P': 15.0, 'Q': 15.0, 'R': 15.0, 'S': 15.0
}
for col_l, w in dash_col_widths.items():
    ws_dash.column_dimensions[col_l].width = w

if hasattr(ws_dash, 'views') and ws_dash.views and ws_dash.views.sheetView:
    ws_dash.views.sheetView[0].showGridLines = True
elif hasattr(ws_dash, 'sheet_view'):
    ws_dash.sheet_view.showGridLines = True

# Save Workbook
print(f"Saving changes to {TARGET_FILE} ...")
wb.save(TARGET_FILE)
print(f"Successfully modified and saved {TARGET_FILE}!")
