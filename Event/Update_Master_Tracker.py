import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.utils import get_column_letter, column_index_from_string
import datetime
import re
import os
import sys
import shutil
import time

sys.stdout.reconfigure(encoding='utf-8')

# Target Excel File detection
CANDIDATE_FILES = [
    "Yearly Event Tracker- DT-3 and DT-4 Updated.xlsx",
    "Yearly Event Tracker- DT-3 and DT-4 Latest(19th august).xlsx",
    "Yearly Event Tracker- DT-3 and DT-4.xlsx"
]

EXCEL_FILE = None
for f in CANDIDATE_FILES:
    if os.path.exists(f):
        EXCEL_FILE = f
        break

if not EXCEL_FILE:
    EXCEL_FILE = "Yearly Event Tracker- DT-3 and DT-4 Updated.xlsx"

print(f"Loading workbook: {EXCEL_FILE} ...")
try:
    wb = openpyxl.load_workbook(EXCEL_FILE)
except PermissionError:
    print(f" [INFO] '{EXCEL_FILE}' is locked by Excel. Reading from temporary snapshot...")
    temp_read_file = "Yearly_Tracker_read_temp.xlsx"
    shutil.copyfile(EXCEL_FILE, temp_read_file)
    wb = openpyxl.load_workbook(temp_read_file)
    try:
        os.remove(temp_read_file)
    except: pass

if "Master_Event_Tracker_2026" not in wb.sheetnames:
    ws_master = wb.create_sheet("Master_Event_Tracker_2026")
else:
    ws_master = wb["Master_Event_Tracker_2026"]

FONT_FAMILY = "Segoe UI"
REGULAR_FONT = Font(name=FONT_FAMILY, size=10, color='000000')
BOLD_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color='000000')
HEADER_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color='FFFFFF')
BANNER_TITLE_FONT = Font(name=FONT_FAMILY, size=11, bold=True, color='FFFFFF')
HOME_BTN_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color='FFFFFF', underline='single')

FONT_MAIN_TITLE = Font(name=FONT_FAMILY, size=15, bold=True, color='FFFFFF')
FONT_SUB_TITLE = Font(name=FONT_FAMILY, size=10, bold=False, color='DCE6F1')
FONT_SECTION_HEADER = Font(name=FONT_FAMILY, size=11, bold=True, color='FFFFFF')
FONT_TABLE_HEADER = Font(name=FONT_FAMILY, size=10, bold=True, color='FFFFFF')
FONT_CARD_TITLE = Font(name=FONT_FAMILY, size=9, bold=True, color='595959')
FONT_CARD_SUB = Font(name=FONT_FAMILY, size=8, bold=False, color='7F7F7F')
FONT_BOLD_NAVY = Font(name=FONT_FAMILY, size=10, bold=True, color='1B365D')
FONT_LINK = Font(name=FONT_FAMILY, size=10, bold=True, color='0055AA', underline='single')

FONT_KPI_NAVY = Font(name=FONT_FAMILY, size=18, bold=True, color='1B365D')
FONT_KPI_WHITE = Font(name=FONT_FAMILY, size=18, bold=True, color='FFFFFF')
FONT_KPI_BLACK = Font(name=FONT_FAMILY, size=18, bold=True, color='000000')
FONT_KPI_BLUE = Font(name=FONT_FAMILY, size=18, bold=True, color='1F4E79')

NAVY_DARK_FILL = PatternFill(start_color='1B365D', end_color='1B365D', fill_type='solid')
NAVY_LIGHT_FILL = PatternFill(start_color='2C4D75', end_color='2C4D75', fill_type='solid')
FILL_NAVY_SLATE = PatternFill(start_color='2C4D75', end_color='2C4D75', fill_type='solid')
FILL_ROYAL_BLUE = PatternFill(start_color='204A87', end_color='204A87', fill_type='solid')
FILL_BANNER_DARK = PatternFill(start_color='0F2537', end_color='0F2537', fill_type='solid')
FILL_CARD_GRAY = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
FILL_TOTAL_ROW = PatternFill(start_color='E9EEF4', end_color='E9EEF4', fill_type='solid')
FILL_TEAL_BTN = PatternFill(start_color='008080', end_color='008080', fill_type='solid')
FILL_PURPLE_CARD = PatternFill(start_color='4A69BD', end_color='4A69BD', fill_type='solid')
FILL_DROPDOWN = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
FILL_SOFT_GREEN = PatternFill(start_color='E2F0D9', end_color='E2F0D9', fill_type='solid')

# Status Colors
GREEN_FILL = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
GREEN_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color='FFFFFF')

YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
YELLOW_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color='000000')

RED_FILL = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
RED_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color='FFFFFF')

# Feedback Form Colors
FEEDBACK_SENT_FILL = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
FEEDBACK_SENT_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color='1F4E79')
FEEDBACK_NOT_REQ_FILL = PatternFill(start_color='F2F4F7', end_color='F2F4F7', fill_type='solid')
FEEDBACK_NOT_REQ_FONT = Font(name=FONT_FAMILY, size=10, bold=False, color='595959')

BIG_EVENT_FILL = PatternFill(start_color='EBF5FB', end_color='EBF5FB', fill_type='solid')
BIG_EVENT_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color='1B365D')
SMALL_EVENT_FILL = PatternFill(fill_type=None)
SMALL_EVENT_FONT = Font(name=FONT_FAMILY, size=10, bold=False, color='333333')

NO_FILL = PatternFill(fill_type=None)

# Full Solid Black Grid Border
BLACK_THIN = Side(style='thin', color='000000')
BLACK_BORDER = Border(left=BLACK_THIN, right=BLACK_THIN, top=BLACK_THIN, bottom=BLACK_THIN)
ALIGN_CENTER_WRAP = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)

MONTH_ORDER = {'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6, 'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12}
MONTH_MAP = {'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6, 'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12, 'ap': 4}

HISTORICAL_SENT_ROWS = {
    ('DT-3 Aug26', 8), ('DT-3 Aug26', 9), ('DT-3 Aug26', 10), ('DT-3 Aug26', 11),
    ('DT-3 Aug26', 12), ('DT-3 Aug26', 24), ('DT-3 Aug26', 25), ('DT-3 Aug26', 26),
    ('DT-3 Aug26', 27), ('DT-3 Aug26', 28), ('DT-3 Aug26', 35), ('DT-3 Aug26', 36),
    ('DT-4 Aug26', 8), ('DT-4 Aug26', 9), ('DT-4 Aug26', 10), ('DT-4 Aug26', 11),
    ('DT-4 Aug26', 14), ('DT-4 Aug26', 15), ('DT-4 Aug26', 20), ('DT-4 Aug26', 23),
    ('DT-4 Aug26', 24), ('DT-4 Aug26', 27), ('DT-4 Aug26', 28), ('DT-4 Aug26', 31),
    ('DT-4 Aug26', 32), ('DT-4 Aug26', 35), ('DT-4 Aug26', 36), ('DT-4 Aug26', 37),
    ('DT-4 Aug26', 38), ('DT-4 Aug26', 42), ('DT-4 Aug26', 45), ('DT-4 Aug26', 47),
    ('DT-4 Aug26', 50), ('DT-4 Aug26', 55), ('DT-4 Aug26', 56), ('DT-4 Aug26', 58),
    ('DT-4 Aug26', 62), ('DT-4 Aug26', 63), ('DT-4 Aug26', 64), ('DT-4 Aug26', 67),
    ('DT-4 Aug26', 68), ('DT-4 Aug26', 70), ('DT-4 Aug26', 71), ('DT-4 Aug26', 72),
    ('DT-4 Aug26', 73), ('DT-4 Aug26', 74), ('DT-4 Aug26', 75), ('DT-4 Aug26', 76),
    ('DT-4 Aug26', 77), ('DT-4 Aug26', 80)
}

CUTOFF_DATE = datetime.date(2026, 8, 21)

def parse_date(val, default_year=2026):
    if val is None:
        return None
    if isinstance(val, (datetime.date, datetime.datetime)):
        return val.date() if isinstance(val, datetime.datetime) else val
    s = str(val).strip()
    if not s or s.lower() in ('na', 'none', '-', 'holiday', 'nl', 'no'):
        return None
    m = re.search(r'(\d{4})-(\d{1,2})-(\d{1,2})', s)
    if m:
        try:
            return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except: pass
    m = re.search(r'(\d{1,2})[/.-](\d{1,2})[/.-](\d{4})', s)
    if m:
        try:
            return datetime.date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except: pass
    m = re.search(r'(\d{1,2})(?:st|nd|rd|th)?[\s\'\-_]*([a-zA-Z]+)[\s\'\-_]*(\d{2,4})?', s)
    if m:
        day = int(m.group(1))
        mon_str = m.group(2).lower()[:3]
        mon = MONTH_MAP.get(mon_str)
        yr_str = m.group(3)
        if yr_str:
            yr = int(yr_str)
            if yr < 100: yr += 2000
        else:
            yr = default_year
        if mon:
            try:
                return datetime.date(yr, mon, day)
            except: pass
    return None

def format_date_val(val, default_year=2026):
    parsed = parse_date(val, default_year)
    return parsed if parsed else str(val or '').strip()

def parse_time_str(val):
    if val is None:
        return None
    if isinstance(val, datetime.time):
        return val.strftime('%I:%M %p')
    if isinstance(val, datetime.datetime):
        return val.strftime('%I:%M %p')
    if isinstance(val, (int, float)):
        if 0 <= val < 1:
            total_seconds = int(round(val * 86400))
            hours = (total_seconds // 3600) % 24
            minutes = (total_seconds % 3600) // 60
            ampm = 'PM' if hours >= 12 else 'AM'
            h12 = hours % 12
            if h12 == 0: h12 = 12
            return f"{h12:02d}:{minutes:02d} {ampm}"
        val = str(val)
    
    s = str(val).strip()
    if not s or s.lower() in ('na', 'none', '-', 'no', 'nl', 'yes'):
        return None
    if s.lower() in ('holiday', 'tbc', 'full day', 'all day'):
        return s.title()
        
    s_clean = s.replace(';', ':').replace('o', '0').replace('O', '0').replace('.', ':')
    s_clean = re.sub(r':([apAP][mM])', r' \1', s_clean)
    s_clean = re.sub(r'\s*:\s*', ':', s_clean)
    
    m = re.search(r'(\d{1,2}):(\d{2})(?::\d{2})?\s*([apAP][mM])?', s_clean)
    if m:
        hr = int(m.group(1))
        mn = int(m.group(2))
        ampm = m.group(3)
        if ampm:
            ampm = ampm.upper()
            if hr > 12: hr -= 12
            elif hr == 0: hr = 12
            return f"{hr:02d}:{mn:02d} {ampm}"
        else:
            if hr >= 12:
                ampm = 'PM'
                if hr > 12: hr -= 12
            else:
                ampm = 'AM'
                if hr == 0: hr = 12
            return f"{hr:02d}:{mn:02d} {ampm}"
            
    m2 = re.search(r'(\d{1,2})\s*([apAP][a-zA-Z]*)', s_clean)
    if m2:
        hr = int(m2.group(1))
        ampm_str = m2.group(2).upper()
        ampm = 'PM' if 'P' in ampm_str else 'AM'
        if hr > 12: hr -= 12
        elif hr == 0: hr = 12
        return f"{hr:02d}:00 {ampm}"
        
    m3 = re.search(r'(\d{2})(\d{2})\s*([apAP][mM])', s_clean)
    if m3:
        hr = int(m3.group(1))
        mn = int(m3.group(2))
        ampm = m3.group(3).upper()
        if hr > 12: hr -= 12
        elif hr == 0: hr = 12
        return f"{hr:02d}:{mn:02d} {ampm}"
        
    return None

def parse_day_str(val, date_obj=None):
    if date_obj and isinstance(date_obj, (datetime.date, datetime.datetime)):
        return date_obj.strftime('%a')
    if not val:
        return ''
    day_map = {
        'mon': 'Mon', 'monday': 'Mon',
        'tue': 'Tue', 'tues': 'Tue', 'tuesday': 'Tue',
        'wed': 'Wed', 'wednesday': 'Wed',
        'thu': 'Thu', 'thur': 'Thu', 'thurs': 'Thu', 'thursday': 'Thu',
        'fri': 'Fri', 'friday': 'Fri',
        'sat': 'Sat', 'saturday': 'Sat',
        'sun': 'Sun', 'sunday': 'Sun'
    }
    return day_map.get(str(val).strip().lower(), str(val).strip())

def parse_capacity(val):
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip().lower()
    if s in ('na', 'none', '-', '', 'nl'):
        return 0
    m = re.search(r'(\d+)', s)
    if m:
        return int(m.group(1))
    return 0

def standardize_status_with_cutoff(status_val, event_date_obj=None, sname=""):
    if not status_val:
        cur = 'Pending Calendar Booking'
    else:
        s = str(status_val).strip().lower()
        if 'cancel' in s:
            return 'Canceled'
        elif 'confirm' in s or 'done' in s:
            return 'Confirmed (Done)'
        else:
            cur = 'Pending Calendar Booking'

    if event_date_obj and isinstance(event_date_obj, (datetime.date, datetime.datetime)):
        d = event_date_obj.date() if isinstance(event_date_obj, datetime.datetime) else event_date_obj
        if d <= CUTOFF_DATE:
            return 'Confirmed (Done)'
        else:
            return cur
    else:
        past_months = ['Jan26', 'Feb26', 'Mar26', 'Apr26', 'May26', 'Jun26', 'Jul26', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', '25', '24']
        if any(m in sname for m in past_months):
            return 'Confirmed (Done)'
        return cur

def get_event_scale(pax_val):
    pax = parse_capacity(pax_val)
    if pax >= 100:
        return "Big Event"
    else:
        return "Small Event"

def categorize_event(title, loc, remarks):
    t = (str(title or '') + ' ' + str(loc or '') + ' ' + str(remarks or '')).lower()
    if 'townhall' in t or 'town hall' in t or 'all hands' in t or 'all-hands' in t:
        return 'Townhall & All-Hands'
    elif 'grad' in t or 'training' in t or 'enablement' in t or 'learning' in t:
        return 'Training & Enablement'
    elif 'induction' in t or 'onboarding' in t or 'on boarding' in t or 'new joiner' in t:
        return 'Induction & Onboarding'
    elif 'codility' in t or 'assesment' in t or 'assessment' in t or 'hiring' in t or 'interview' in t:
        return 'Assessment & Hiring'
    elif 'leadership' in t or 'slt' in t or 'xlt' in t or 'visitor' in t or 'visit' in t or 'cio' in t:
        return 'Leadership & VIP Visit'
    elif 'celebration' in t or 'birthday' in t or 'hi tea' in t or 'hi-tea' in t or 'independence' in t or 'social' in t or 'lunch' in t or 'connect' in t:
        return 'Social & Team Connect'
    elif 'hackable' in t or 'hackathon' in t or 'copilot' in t or 'tech' in t:
        return 'Tech & Innovation'
    elif 'dry run' in t or 'rehearsal' in t:
        return 'Dry Run & Rehearsal'
    else:
        return 'Team Meeting & Workshop'

def style_and_merge(ws, start_col, end_col, val, font, fill, hyperlink_loc=None):
    for c in range(start_col, end_col + 1):
        cell = ws.cell(1, c)
        cell.font = font
        cell.fill = fill
        cell.border = BLACK_BORDER
        cell.alignment = ALIGN_CENTER_WRAP
        cell.value = None
    c_start = ws.cell(1, start_col)
    c_start.value = val
    if hyperlink_loc:
        c_start.hyperlink = Hyperlink(ref=f"{get_column_letter(start_col)}1", location=hyperlink_loc)
    if end_col > start_col:
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

# =============================================================================
# 1. FORMAT ALL MONTHLY SHEETS
# =============================================================================
def format_monthly_sheet(ws, sname):
    max_r = ws.max_row
    max_c = ws.max_column
    if max_r < 2 or max_c < 2:
        return
        
    is_dt3 = "DT-3" in sname
    is_dt4 = "DT-4" in sname

    r1_merged = [m for m in list(ws.merged_cells.ranges) if m.min_row <= 1 <= m.max_row]
    for m in r1_merged:
        try: ws.unmerge_cells(str(m))
        except: pass

    row2_headers = [str(ws.cell(2, c).value or '').strip().lower() for c in range(1, max_c + 1)]
    has_scale_col = any('scale' in h or 'event size' in h for h in row2_headers)

    if not has_scale_col:
        cap_col_idx = None
        for c in range(1, max_c + 1):
            if 'capacity' in row2_headers[c-1] or 'pax' in row2_headers[c-1]:
                cap_col_idx = c
                break
        if not cap_col_idx:
            cap_col_idx = 6 if is_dt3 else (7 if is_dt4 else None)

        if cap_col_idx:
            insert_pos = cap_col_idx + 1
            ws.insert_cols(insert_pos)
            ws.cell(2, insert_pos, "Event Scale")
            max_c = ws.max_column

    header_cols = {}
    for c in range(1, max_c + 1):
        val = str(ws.cell(2, c).value or '').strip().lower()
        if 'event id' in val or val == 'id': header_cols['event_id'] = c
        elif 'date' in val and 'booking' not in val: header_cols['event_date'] = c
        elif val == 'day': header_cols['day'] = c
        elif 'floor' in val: header_cols['floor'] = c
        elif 'location' in val: header_cols['location'] = c
        elif 'title' in val or 'event type' in val: header_cols['title'] = c
        elif 'capacity' in val or 'pax' in val: header_cols['capacity'] = c
        elif 'scale' in val or 'event size' in val: header_cols['scale'] = c
        elif 'start' in val and 'time' in val: header_cols['start_time'] = c
        elif 'end' in val and 'time' in val: header_cols['end_time'] = c
        elif 'f&' in val or 'f & b' in val or 'f&b' in val: header_cols['fnb'] = c
        elif 'logistic' in val: header_cols['logistics'] = c
        elif 'booked by' in val: header_cols['booked_by'] = c
        elif 'email' in val: header_cols['email'] = c
        elif 'booking date' in val: header_cols['booking_date'] = c
        elif 'status' in val: header_cols['status'] = c
        elif 'feedback' in val: header_cols['feedback'] = c
        elif 'remarks' in val: header_cols['remarks'] = c
        elif 'taken by' in val: header_cols['taken_by'] = c

    ws.row_dimensions[1].height = 28

    if "DT-3" in sname:
        m_tag = sname.replace("DT-3", "").strip().upper()
        style_and_merge(ws, 1, 3, "🏠 ⮌ RETURN TO DASHBOARD", HOME_BTN_FONT, NAVY_DARK_FILL, "'Dashboard'!A1")
        style_and_merge(ws, 4, 13, f"DOWNTOWN-3 EVENT OPERATIONS TRACKER • {m_tag}", BANNER_TITLE_FONT, NAVY_LIGHT_FILL)
        style_and_merge(ws, 14, max_c, "🟢 Green: Confirmed | 🟡 Yellow: Pending | 🔴 Red: Canceled", HEADER_FONT, NAVY_DARK_FILL)
    elif "DT-4" in sname:
        m_tag = sname.replace("DT-4", "").strip().upper()
        style_and_merge(ws, 1, 3, "🏠 ⮌ RETURN TO DASHBOARD", HOME_BTN_FONT, NAVY_DARK_FILL, "'Dashboard'!A1")
        style_and_merge(ws, 4, 14, f"DOWNTOWN-4 EVENT OPERATIONS TRACKER • {m_tag}", BANNER_TITLE_FONT, NAVY_LIGHT_FILL)
        style_and_merge(ws, 15, max_c, "🟢 Green: Confirmed | 🟡 Yellow: Pending | 🔴 Red: Canceled", HEADER_FONT, NAVY_DARK_FILL)
    else:
        style_and_merge(ws, 1, 3, "🏠 ⮌ RETURN TO DASHBOARD", HOME_BTN_FONT, NAVY_DARK_FILL, "'Dashboard'!A1")
        style_and_merge(ws, 4, 11, f"EVENT OPERATIONS TRACKER • {sname.upper()}", BANNER_TITLE_FONT, NAVY_LIGHT_FILL)
        style_and_merge(ws, 12, max_c, "🟢 Green: Confirmed | 🟡 Yellow: Pending | 🔴 Red: Canceled", HEADER_FONT, NAVY_DARK_FILL)

    ws.row_dimensions[2].height = 28
    for c in range(1, max_c + 1):
        cell = ws.cell(2, c)
        cell.font = HEADER_FONT
        cell.fill = NAVY_DARK_FILL
        cell.alignment = ALIGN_CENTER_WRAP
        cell.border = BLACK_BORDER

    for r in range(3, max_r + 1):
        ws.row_dimensions[r].height = 24
        row_has_data = any(ws.cell(r, col_i).value is not None and str(ws.cell(r, col_i).value).strip() != '' for col_i in range(1, max_c + 1) if col_i not in (header_cols.get('feedback'), header_cols.get('scale')))

        date_cell_val = ws.cell(r, header_cols.get('event_date')).value if header_cols.get('event_date') else None
        date_obj = parse_date(date_cell_val)
        cap_val = ws.cell(r, header_cols.get('capacity')).value if header_cols.get('capacity') else None
        status_raw = ws.cell(r, header_cols.get('status')).value if header_cols.get('status') else None
        pax_num = parse_capacity(cap_val)
        scale_val = "Big Event" if pax_num >= 100 else ("Small Event" if row_has_data else "")

        final_status = standardize_status_with_cutoff(status_raw, date_obj, sname) if row_has_data else None

        is_orig_sent = (sname, r) in HISTORICAL_SENT_ROWS
        if is_orig_sent:
            final_feedback = "Form sent"
        else:
            final_feedback = "Not Required" if row_has_data else None

        for c in range(1, max_c + 1):
            cell = ws.cell(r, c)
            val = cell.value
            val_str = str(val or '').strip()
            
            cell.border = BLACK_BORDER
            cell.alignment = ALIGN_CENTER_WRAP
            cell.font = REGULAR_FONT
            cell.fill = NO_FILL

            if c == header_cols.get('scale'):
                if row_has_data:
                    cell.value = scale_val
                    if scale_val == "Big Event":
                        cell.fill = BIG_EVENT_FILL
                        cell.font = BIG_EVENT_FONT
                    else:
                        cell.fill = SMALL_EVENT_FILL
                        cell.font = SMALL_EVENT_FONT
                else:
                    cell.value = None
                    cell.fill = NO_FILL
                    cell.font = REGULAR_FONT

            elif c == header_cols.get('status'):
                if row_has_data:
                    cell.value = final_status
                    if final_status == 'Confirmed (Done)':
                        cell.fill = GREEN_FILL
                        cell.font = GREEN_FONT
                    elif final_status == 'Canceled':
                        cell.fill = RED_FILL
                        cell.font = RED_FONT
                    else:
                        cell.fill = YELLOW_FILL
                        cell.font = YELLOW_FONT
                else:
                    cell.value = None
                    cell.fill = NO_FILL
                    cell.font = REGULAR_FONT

            elif c == header_cols.get('feedback'):
                if row_has_data:
                    cell.value = final_feedback
                    if final_feedback == "Form sent":
                        cell.fill = FEEDBACK_SENT_FILL
                        cell.font = FEEDBACK_SENT_FONT
                    else:
                        cell.fill = FEEDBACK_NOT_REQ_FILL
                        cell.font = FEEDBACK_NOT_REQ_FONT
                else:
                    cell.value = None
                    cell.fill = NO_FILL
                    cell.font = REGULAR_FONT

            elif c == header_cols.get('capacity'):
                cell.number_format = '#,##0'
            elif c in (header_cols.get('start_time'), header_cols.get('end_time')):
                parsed_t = parse_time_str(val)
                if parsed_t: cell.value = parsed_t
                if isinstance(val, (datetime.time, datetime.datetime)): cell.number_format = 'hh:mm AM/PM'
            elif c in (header_cols.get('event_date'), header_cols.get('booking_date')):
                if isinstance(val, (datetime.date, datetime.datetime)): cell.number_format = 'dd-mmm-yyyy'

    if header_cols.get('status'):
        col_letter = get_column_letter(header_cols['status'])
        dv_status = DataValidation(type="list", formula1='"Confirmed (Done),Pending Calendar Booking,Canceled"', allow_blank=True)
        dv_status.add(f"{col_letter}3:{col_letter}{max_r}")
        ws.add_data_validation(dv_status)

    if header_cols.get('feedback'):
        col_letter = get_column_letter(header_cols['feedback'])
        dv_fb = DataValidation(type="list", formula1='"Form sent,Not Required,NA"', allow_blank=True)
        dv_fb.add(f"{col_letter}3:{col_letter}{max_r}")
        ws.add_data_validation(dv_fb)

    if header_cols.get('scale'):
        col_letter = get_column_letter(header_cols['scale'])
        dv_scale = DataValidation(type="list", formula1='"Big Event,Small Event"', allow_blank=True)
        dv_scale.add(f"{col_letter}3:{col_letter}{max_r}")
        ws.add_data_validation(dv_scale)

    ws.freeze_panes = 'A3'
    if hasattr(ws, 'views') and ws.views and ws.views.sheetView:
        ws.views.sheetView[0].showGridLines = True
    elif hasattr(ws, 'sheet_view'):
        ws.sheet_view.showGridLines = True

    for c in range(1, max_c + 1):
        col_letter = get_column_letter(c)
        max_len = 0
        for r in range(2, max_r + 1):
            cell = ws.cell(r, c)
            val = cell.value
            if val is None: continue
            v_str = str(val).strip()
            if v_str.startswith('='): v_len = 15
            elif isinstance(val, (datetime.date, datetime.datetime)): v_len = 12
            elif isinstance(val, datetime.time): v_len = 10
            else: v_len = len(v_str)
            if v_len > max_len: max_len = v_len
                
        min_w = 12
        if c == header_cols.get('event_id'): min_w = 16
        elif c == header_cols.get('event_date'): min_w = 15
        elif c == header_cols.get('title'): min_w = 28
        elif c == header_cols.get('location'): min_w = 22
        elif c == header_cols.get('remarks'): min_w = 32
        elif c == header_cols.get('status'): min_w = 25
        elif c == header_cols.get('feedback'): min_w = 18
        elif c == header_cols.get('scale'): min_w = 15
        elif c == header_cols.get('email'): min_w = 25
        elif c == header_cols.get('booked_by'): min_w = 20
        elif c == header_cols.get('logistics'): min_w = 22
        elif c == header_cols.get('fnb'): min_w = 18

        ws.column_dimensions[col_letter].width = min(max(max_len + 3, min_w), 48)

print("Formatting all monthly event sheets, restoring Form sent, and marking Confirmed till date...")
monthly_sheets = [s for s in wb.sheetnames if s not in ('Dashboard', 'Master_Event_Tracker_2026')]
for idx, sname in enumerate(monthly_sheets, 1):
    format_monthly_sheet(wb[sname], sname)
print(f" Successfully formatted {len(monthly_sheets)} monthly sheets.")

# =============================================================================
# 2. SYNCHRONIZE MASTER EVENT TRACKER 2026
# =============================================================================
print("Synchronizing Master Event Tracker (2026)...")
r1_m_merged = [m for m in list(ws_master.merged_cells.ranges) if m.min_row <= 1 <= m.max_row]
for m in r1_m_merged:
    try: ws_master.unmerge_cells(str(m))
    except: pass

ws_master.delete_rows(1, ws_master.max_row + 10)

ws_master.row_dimensions[1].height = 28
style_and_merge(ws_master, 1, 3, "🏠 ⮌ BACK TO DASHBOARD", HOME_BTN_FONT, NAVY_DARK_FILL, "'Dashboard'!A1")
style_and_merge(ws_master, 4, 16, "MASTER EVENT OPERATIONS DATABASE • FULL YEAR 2026 (DOWNTOWN-3 & DOWNTOWN-4)", BANNER_TITLE_FONT, NAVY_LIGHT_FILL)
style_and_merge(ws_master, 17, 23, "🟢 Green: Confirmed | 🟡 Yellow: Pending | 🔴 Red: Canceled", HEADER_FONT, NAVY_DARK_FILL)

master_headers = [
    "Event ID", "Year", "Month", "Event Date", "Day", "Building", "Floor", 
    "Location", "Event Title / Purpose", "Category", "Capacity / Pax", "Event Scale",
    "Start Time", "End Time", "F&B Support", "Logistic Support", 
    "Booked By", "Requester Email", "Calendar Booking Status", 
    "Feedback Form", "Booking Date", "Remarks", "Booking Taken By"
]

ws_master.row_dimensions[2].height = 28
for col_idx, h_text in enumerate(master_headers, 1):
    c = ws_master.cell(2, col_idx, h_text)
    c.fill = NAVY_DARK_FILL
    c.font = HEADER_FONT
    c.alignment = ALIGN_CENTER_WRAP
    c.border = BLACK_BORDER

sheets_active = [s for s in wb.sheetnames if ('DT-3' in s or 'DT-4' in s) and ('26' in s or '27' in s)]

def sort_key(s):
    yr = 1 if '27' in s else 0
    b = 0 if 'DT-3' in s else 1
    m_num = 1
    for k, v in MONTH_ORDER.items():
        if k in s:
            m_num = v
            break
    return (yr, m_num, b)

sheets_active.sort(key=sort_key)

current_out_row = 3
total_synced = 0

for sname in sheets_active:
    ws = wb[sname]
    bldg = "Downtown-3" if "DT-3" in sname else "Downtown-4"
    is_dt3 = (bldg == "Downtown-3")
    yr_val = 2027 if "27" in sname else 2026
    m_name = "Jan"
    for k in MONTH_ORDER.keys():
        if k in sname:
            m_name = k
            break
            
    for r in range(3, ws.max_row + 1):
        d_val = ws.cell(r, 2).value
        day_val = ws.cell(r, 3).value
        loc_val = str(ws.cell(r, 4).value or '').strip()
        
        if is_dt3:
            floor_val = 'L-1' if '1' in loc_val else ('L-5' if '5' in loc_val else ('L-4' if '4' in loc_val else ('L-6' if '6' in loc_val else 'General')))
            title_val = str(ws.cell(r, 5).value or '').strip()
            cap_val = ws.cell(r, 6).value
            scale_val = str(ws.cell(r, 7).value or '').strip()
            start_time = ws.cell(r, 8).value
            end_time = ws.cell(r, 9).value
            fnb_val = str(ws.cell(r, 10).value or '').strip()
            logistics_val = str(ws.cell(r, 11).value or '').strip()
            booked_by_val = str(ws.cell(r, 12).value or '').strip()
            email_val = str(ws.cell(r, 13).value or '').strip()
            booking_date_val = ws.cell(r, 14).value
            status_val = ws.cell(r, 15).value
            feedback_val = ws.cell(r, 16).value
            remarks_val = str(ws.cell(r, 17).value or '').strip()
            taken_by_val = str(ws.cell(r, 18).value or '').strip()
        else:
            floor_val = str(ws.cell(r, 5).value or '').strip()
            title_val = str(ws.cell(r, 6).value or '').strip()
            cap_val = ws.cell(r, 7).value
            scale_val = str(ws.cell(r, 8).value or '').strip()
            start_time = ws.cell(r, 9).value
            end_time = ws.cell(r, 10).value
            fnb_val = str(ws.cell(r, 11).value or '').strip()
            logistics_val = str(ws.cell(r, 12).value or '').strip()
            booked_by_val = str(ws.cell(r, 13).value or '').strip()
            email_val = str(ws.cell(r, 14).value or '').strip()
            booking_date_val = ws.cell(r, 15).value
            status_val = ws.cell(r, 16).value
            feedback_val = ws.cell(r, 17).value
            remarks_val = str(ws.cell(r, 18).value or '').strip()
            taken_by_val = str(ws.cell(r, 19).value or '').strip()
            
        pax = parse_capacity(cap_val)
        has_event = any(v is not None and str(v).strip() != '' for v in (d_val, day_val, loc_val, title_val, booked_by_val, remarks_val, status_val)) or (pax > 0)
        
        if not has_event:
            continue
            
        total_synced += 1
        evt_formula = f'=IF(OR(D{current_out_row}<>"",H{current_out_row}<>"",I{current_out_row}<>"",Q{current_out_row}<>"",S{current_out_row}<>""), "EVT-" & TEXT(B{current_out_row}, "0000") & "-" & TEXT(COUNTIF(OFFSET($A$1, 1, 0, ROW()-2, 1), "EVT*") + 1, "0000"), "")'
        
        date_obj = parse_date(d_val, yr_val)
        formatted_date = date_obj if date_obj else format_date_val(d_val, yr_val)
        formatted_day = parse_day_str(day_val, date_obj)
        formatted_start = parse_time_str(start_time)
        formatted_end = parse_time_str(end_time)
        booking_date_obj = parse_date(booking_date_val, yr_val)
        formatted_booking_date = booking_date_obj if booking_date_obj else format_date_val(booking_date_val, yr_val)
        
        final_status = standardize_status_with_cutoff(status_val, date_obj, sname)
        final_scale = get_event_scale(pax)
        
        if (sname, r) in HISTORICAL_SENT_ROWS:
            final_feedback = "Form sent"
        else:
            final_feedback = "Not Required"
            
        cat = categorize_event(title_val, loc_val, remarks_val)
        
        row_vals = [
            evt_formula, yr_val, m_name, formatted_date, formatted_day, bldg, floor_val,
            loc_val, title_val, cat, pax, final_scale, formatted_start, formatted_end,
            fnb_val, logistics_val, booked_by_val, email_val, final_status,
            final_feedback, formatted_booking_date, remarks_val, taken_by_val
        ]
        
        ws_master.row_dimensions[current_out_row].height = 24
        for mc_idx, m_val in enumerate(row_vals, 1):
            mc = ws_master.cell(current_out_row, mc_idx, m_val)
            mc.font = REGULAR_FONT
            mc.border = BLACK_BORDER
            mc.alignment = ALIGN_CENTER_WRAP
            
            if mc_idx == 11: mc.number_format = '#,##0'
            elif mc_idx in (13, 14): mc.number_format = 'hh:mm AM/PM'
            elif mc_idx in (4, 21):
                if isinstance(m_val, (datetime.date, datetime.datetime)): mc.number_format = 'dd-mmm-yyyy'
                    
            if mc_idx == 12:
                if m_val == "Big Event":
                    mc.fill = BIG_EVENT_FILL
                    mc.font = BIG_EVENT_FONT
                else:
                    mc.fill = SMALL_EVENT_FILL
                    mc.font = SMALL_EVENT_FONT
            elif mc_idx == 19:
                if "Confirmed" in final_status:
                    mc.fill = GREEN_FILL
                    mc.font = GREEN_FONT
                elif "Canceled" in final_status:
                    mc.fill = RED_FILL
                    mc.font = RED_FONT
                else:
                    mc.fill = YELLOW_FILL
                    mc.font = YELLOW_FONT
            elif mc_idx == 20:
                if final_feedback == "Form sent":
                    mc.fill = FEEDBACK_SENT_FILL
                    mc.font = FEEDBACK_SENT_FONT
                else:
                    mc.fill = FEEDBACK_NOT_REQ_FILL
                    mc.font = FEEDBACK_NOT_REQ_FONT

        current_out_row += 1

print(f"Total synchronized events in Master Tracker: {total_synced}")

dv_m_scale = DataValidation(type="list", formula1='"Big Event,Small Event"', allow_blank=True)
dv_m_scale.add(f"L3:L{total_synced+50}")
ws_master.add_data_validation(dv_m_scale)

dv_m_status = DataValidation(type="list", formula1='"Confirmed (Done),Pending Calendar Booking,Canceled"', allow_blank=True)
dv_m_status.add(f"S3:S{total_synced+50}")
ws_master.add_data_validation(dv_m_status)

dv_m_fb = DataValidation(type="list", formula1='"Form sent,Not Required,NA"', allow_blank=True)
dv_m_fb.add(f"T3:T{total_synced+50}")
ws_master.add_data_validation(dv_m_fb)

ws_master.freeze_panes = 'A3'
if hasattr(ws_master, 'views') and ws_master.views and ws_master.views.sheetView:
    ws_master.views.sheetView[0].showGridLines = True

master_min_widths = [16, 10, 10, 15, 10, 14, 12, 22, 30, 24, 14, 15, 14, 14, 18, 22, 20, 25, 25, 18, 15, 32, 18]
for col_idx in range(1, len(master_headers) + 1):
    col_letter = get_column_letter(col_idx)
    max_len = 0
    for r in range(2, current_out_row):
        val = ws_master.cell(r, col_idx).value
        if val is None: continue
        v_str = str(val).strip()
        if v_str.startswith('='): v_len = 15
        elif isinstance(val, (datetime.date, datetime.datetime)): v_len = 12
        elif isinstance(val, datetime.time): v_len = 10
        else: v_len = len(v_str)
        if v_len > max_len: max_len = v_len
    base_min = master_min_widths[col_idx - 1] if col_idx - 1 < len(master_min_widths) else 12
    ws_master.column_dimensions[col_letter].width = min(max(max_len + 3, base_min), 48)


# =============================================================================
# 3. BUILD COMPLETE, COLORFUL & AUTOMATED DASHBOARD WITH DT-3 & DT-4 FEEDBACK VISIBILITY
# =============================================================================
print("Building and refreshing Dashboard...")
if "Dashboard" in wb.sheetnames:
    del wb["Dashboard"]

ws_dash = wb.create_sheet("Dashboard", 0)

def style_dash_cell(ws, r, c, val=None, font=None, fill=None, border=None, align=None, num_format=None, hyperlink_loc=None):
    cell = ws.cell(r, c)
    if val is not None: cell.value = val
    if font: cell.font = font
    if fill: cell.fill = fill
    if border: cell.border = border
    if align: cell.alignment = align
    if num_format: cell.number_format = num_format
    if hyperlink_loc:
        col_letter = get_column_letter(c)
        cell.hyperlink = Hyperlink(ref=f"{col_letter}{r}", location=hyperlink_loc)
    return cell

def style_dash_range(ws, min_r, min_c, max_r, max_c, font=None, fill=None, border=None, align=None, num_format=None):
    for r in range(min_r, max_r + 1):
        for c in range(min_c, max_c + 1):
            cell = ws.cell(r, c)
            if font: cell.font = font
            if fill: cell.fill = fill
            if border: cell.border = border
            if align: cell.alignment = align
            if num_format: cell.number_format = num_format

# Banner
ws_dash.row_dimensions[1].height = 32
ws_dash.row_dimensions[2].height = 20
ws_dash.merge_cells('A1:S1')
ws_dash.merge_cells('A2:S2')

style_dash_cell(ws_dash, 1, 1, "🏢 CBRE | YEARLY EVENT OPERATIONS DASHBOARD", font=FONT_MAIN_TITLE, fill=FILL_BANNER_DARK, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 1, 1, 1, 19, fill=FILL_BANNER_DARK)

style_dash_cell(ws_dash, 2, 1, "Downtown-3 & Downtown-4 Facilities • 2026 Event Planning, Capacity Analytics & Operations Management", font=FONT_SUB_TITLE, fill=NAVY_DARK_FILL, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 2, 1, 2, 19, fill=NAVY_DARK_FILL)

ws_dash.row_dimensions[3].height = 8
ws_dash.row_dimensions[4].height = 20
ws_dash.row_dimensions[5].height = 26
ws_dash.merge_cells('A4:S4')
style_dash_cell(ws_dash, 4, 1, "OPERATIONS CALENDAR BOOKING POLICY & STATUS GUIDE:", font=FONT_BOLD_NAVY, align=ALIGN_LEFT_WRAP)

ws_dash.merge_cells('B5:E5')
style_dash_cell(ws_dash, 5, 2, "🟢 GREEN: Calendar Booking Done / Completed", font=FONT_TABLE_HEADER, fill=GREEN_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 5, 2, 5, 5, fill=GREEN_FILL, border=BLACK_BORDER)

ws_dash.merge_cells('G5:K5')
style_dash_cell(ws_dash, 5, 7, "🟡 YELLOW: Booking Pending (Upcoming / Action Required)", font=YELLOW_FONT, fill=YELLOW_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 5, 7, 5, 11, fill=YELLOW_FILL, border=BLACK_BORDER)

ws_dash.merge_cells('M5:P5')
style_dash_cell(ws_dash, 5, 13, "🔴 RED: Event Canceled by Organizer", font=RED_FONT, fill=RED_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 5, 13, 5, 16, fill=RED_FILL, border=BLACK_BORDER)

ws_dash.merge_cells('R5:S5')
style_dash_cell(ws_dash, 5, 18, "📑 Open Master Tracker ↗", font=FONT_TABLE_HEADER, fill=FILL_TEAL_BTN, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, hyperlink_loc="'Master_Event_Tracker_2026'!A1")
style_dash_range(ws_dash, 5, 18, 5, 19, fill=FILL_TEAL_BTN, border=BLACK_BORDER)

ws_dash.row_dimensions[6].height = 8

# KPI Cards (Cols B to S)
ws_dash.row_dimensions[7].height = 18
ws_dash.row_dimensions[8].height = 32
ws_dash.row_dimensions[9].height = 16

cards_cfg = [
    ('B', 'C', 'TOTAL EVENTS (2026)', '=B25+I28', 'DT-3 & DT-4 combined', FILL_CARD_GRAY, FONT_CARD_TITLE, FONT_KPI_NAVY, FONT_CARD_SUB),
    ('D', 'E', 'EVENTS CONDUCTED (DONE)', '=C25+J28', '🟢 Done / Confirmed', GREEN_FILL, FONT_TABLE_HEADER, FONT_KPI_WHITE, FONT_TABLE_HEADER),
    ('F', 'G', '🌟 BIG EVENTS (100+)', '=P25', '100+ Pax (Feedback Applicable)', BIG_EVENT_FILL, Font(name=FONT_FAMILY, size=9, bold=True, color='1B365D'), FONT_KPI_BLUE, Font(name=FONT_FAMILY, size=8, bold=True, color='1B365D')),
    ('H', 'I', '👥 SMALL EVENTS (<100)', '=Q25', '< 100 Pax (Routine Events)', FILL_CARD_GRAY, FONT_CARD_TITLE, FONT_KPI_NAVY, FONT_CARD_SUB),
    ('J', 'K', 'TOTAL ATTENDEES', '=F25+M28', '👥 Expected Pax', FILL_PURPLE_CARD, FONT_TABLE_HEADER, FONT_KPI_WHITE, FONT_TABLE_HEADER),
    ('L', 'M', 'DOWNTOWN-3 EVENTS', '=B25', '🏢 DT-3 Facility', FILL_CARD_GRAY, FONT_CARD_TITLE, FONT_KPI_NAVY, FONT_CARD_SUB),
    ('N', 'O', 'DOWNTOWN-4 EVENTS', '=I28', '🏢 DT-4 Facility', FILL_CARD_GRAY, FONT_CARD_TITLE, FONT_KPI_NAVY, FONT_CARD_SUB),
    ('P', 'Q', '📨 TOTAL FEEDBACK SENT', '=R25+S25', 'DT-3: 12 | DT-4: 37', FEEDBACK_SENT_FILL, Font(name=FONT_FAMILY, size=9, bold=True, color='1F4E79'), FONT_KPI_BLUE, Font(name=FONT_FAMILY, size=8, bold=True, color='1F4E79')),
    ('R', 'S', '⚪ FEEDBACK NOT REQUIRED', '=COUNTIF(Master_Event_Tracker_2026!$T$3:$T$5000, "Not Required")', 'Routine Events / NA', FILL_CARD_GRAY, FONT_CARD_TITLE, FONT_KPI_NAVY, FONT_CARD_SUB),
]

for c1, c2, title, formula, sub, fill, f_title, f_kpi, f_sub in cards_cfg:
    c1_idx = column_index_from_string(c1)
    c2_idx = column_index_from_string(c2)
    ws_dash.merge_cells(f'{c1}7:{c2}7')
    ws_dash.merge_cells(f'{c1}8:{c2}8')
    ws_dash.merge_cells(f'{c1}9:{c2}9')
    style_dash_range(ws_dash, 7, c1_idx, 9, c2_idx, fill=fill, border=BLACK_BORDER)
    style_dash_cell(ws_dash, 7, c1_idx, title, font=f_title, align=ALIGN_CENTER_WRAP)
    style_dash_cell(ws_dash, 8, c1_idx, formula, font=f_kpi, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, 9, c1_idx, sub, font=f_sub, align=ALIGN_CENTER_WRAP)

ws_dash.row_dimensions[10].height = 10

# Monthly Tables (Rows 11-28)
ws_dash.row_dimensions[11].height = 26
ws_dash.row_dimensions[12].height = 24

ws_dash.merge_cells('A11:F11')
style_dash_cell(ws_dash, 11, 1, "🏢 DOWNTOWN-3 (DT-3) MONTHLY EVENT TRACKER & NAVIGATION", font=FONT_SECTION_HEADER, fill=FILL_ROYAL_BLUE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 11, 1, 11, 6, fill=FILL_ROYAL_BLUE, border=BLACK_BORDER)

ws_dash.merge_cells('H11:M11')
style_dash_cell(ws_dash, 11, 8, "🏢 DOWNTOWN-4 (DT-4) MONTHLY EVENT TRACKER & NAVIGATION", font=FONT_SECTION_HEADER, fill=NAVY_DARK_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 11, 8, 11, 13, fill=NAVY_DARK_FILL, border=BLACK_BORDER)

ws_dash.merge_cells('O11:S11')
style_dash_cell(ws_dash, 11, 15, "📋 MONTHWISE EVENT SCALE & DT-3 / DT-4 FEEDBACK TRACKER", font=FONT_SECTION_HEADER, fill=FILL_ROYAL_BLUE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 11, 15, 11, 19, fill=FILL_ROYAL_BLUE, border=BLACK_BORDER)

dt3_headers = ['Month (Click to Open)', 'Total Events', '🟢 Confirmed', '🟡 Pending', '🔴 Canceled', 'Total Pax']
dt4_headers = ['Month (Click to Open)', 'Total Events', '🟢 Confirmed', '🟡 Pending', '🔴 Canceled', 'Total Pax']
fb_headers = ['Month', '🌟 Big (100+)', '👥 Small (<100)', '📨 DT-3 FB Sent', '📨 DT-4 FB Sent']

for idx, h in enumerate(dt3_headers, 1):
    style_dash_cell(ws_dash, 12, idx, h, font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)

for idx, h in enumerate(dt4_headers, 8):
    style_dash_cell(ws_dash, 12, idx, h, font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)

for idx, h in enumerate(fb_headers, 15):
    style_dash_cell(ws_dash, 12, idx, h, font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)

months_26 = ['Jan26', 'Feb26', 'Mar26', 'Apr26', 'May26', 'Jun26', 'Jul26', 'Aug26', 'Sep26', 'Oct26', 'Nov26', 'Dec26']
month_labels = ['Jan 2026 ↗', 'Feb 2026 ↗', 'Mar 2026 ↗', 'Apr 2026 ↗', 'May 2026 ↗', 'Jun 2026 ↗', 'Jul 2026 ↗', 'Aug 2026 ↗', 'Sep 2026 ↗', 'Oct 2026 ↗', 'Nov 2026 ↗', 'Dec 2026 ↗']
month_short = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

for idx, (m_tag, m_lbl, m_s) in enumerate(zip(months_26, month_labels, month_short)):
    r = 13 + idx
    ws_dash.row_dimensions[r].height = 20
    s_dt3 = f"DT-3 {m_tag}"
    style_dash_cell(ws_dash, r, 1, m_lbl, font=FONT_LINK, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, hyperlink_loc=f"'{s_dt3}'!A1")
    style_dash_cell(ws_dash, r, 2, f'=COUNTIF(\'{s_dt3}\'!$O$3:$O$500, "<>")', font=BOLD_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 3, f'=COUNTIF(\'{s_dt3}\'!$O$3:$O$500, "Confirmed (Done)")', font=GREEN_FONT, fill=GREEN_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 4, f'=COUNTIF(\'{s_dt3}\'!$O$3:$O$500, "Pending Calendar Booking")', font=YELLOW_FONT, fill=YELLOW_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 5, f'=COUNTIF(\'{s_dt3}\'!$O$3:$O$500, "Canceled")', font=RED_FONT, fill=RED_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 6, f'=SUM(\'{s_dt3}\'!$F$3:$F$500)', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

    s_dt4 = f"DT-4 {m_tag}"
    style_dash_cell(ws_dash, r, 8, m_lbl, font=FONT_LINK, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, hyperlink_loc=f"'{s_dt4}'!A1")
    style_dash_cell(ws_dash, r, 9, f'=COUNTIF(\'{s_dt4}\'!$P$3:$P$500, "<>")', font=BOLD_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 10, f'=COUNTIF(\'{s_dt4}\'!$P$3:$P$500, "Confirmed (Done)")', font=GREEN_FONT, fill=GREEN_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 11, f'=COUNTIF(\'{s_dt4}\'!$P$3:$P$500, "Pending Calendar Booking")', font=YELLOW_FONT, fill=YELLOW_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 12, f'=COUNTIF(\'{s_dt4}\'!$P$3:$P$500, "Canceled")', font=RED_FONT, fill=RED_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 13, f'=SUM(\'{s_dt4}\'!$G$3:$G$500)', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

    style_dash_cell(ws_dash, r, 15, m_lbl.replace(" ↗", ""), font=BOLD_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
    style_dash_cell(ws_dash, r, 16, f'=COUNTIFS(Master_Event_Tracker_2026!$C$3:$C$5000, "{m_s}", Master_Event_Tracker_2026!$L$3:$L$5000, "Big Event")', font=BIG_EVENT_FONT, fill=BIG_EVENT_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 17, f'=COUNTIFS(Master_Event_Tracker_2026!$C$3:$C$5000, "{m_s}", Master_Event_Tracker_2026!$L$3:$L$5000, "Small Event")', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    # DT-3 Feedback Sent for this month
    style_dash_cell(ws_dash, r, 18, f'=COUNTIF(\'{s_dt3}\'!$P$3:$P$500, "Form sent")', font=FEEDBACK_SENT_FONT, fill=FEEDBACK_SENT_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    # DT-4 Feedback Sent for this month
    style_dash_cell(ws_dash, r, 19, f'=COUNTIF(\'{s_dt4}\'!$Q$3:$Q$500, "Form sent")', font=FEEDBACK_SENT_FONT, fill=FEEDBACK_SENT_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

# DT-3 Total (Row 25)
ws_dash.row_dimensions[25].height = 22
style_dash_cell(ws_dash, 25, 1, "DT-3 Full Year Total", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_cell(ws_dash, 25, 2, "=SUM(B13:B24)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 25, 3, "=SUM(C13:C24)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 25, 4, "=SUM(D13:D24)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 25, 5, "=SUM(E13:E24)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 25, 6, "=SUM(F13:F24)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

# Monthwise Scale & Feedback Full Year Total (Row 25 in Cols O to S)
style_dash_cell(ws_dash, 25, 15, "Full Year Total", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_cell(ws_dash, 25, 16, "=SUM(P13:P24)", font=BIG_EVENT_FONT, fill=BIG_EVENT_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 25, 17, "=SUM(Q13:Q24)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 25, 18, "=SUM(R13:R24)", font=FEEDBACK_SENT_FONT, fill=FEEDBACK_SENT_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 25, 19, "=SUM(S13:S24)", font=FEEDBACK_SENT_FONT, fill=FEEDBACK_SENT_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

# DT-4 2027 Rows
months_27 = [('Jan27', "Jan '27 ↗"), ('Feb27', "Feb '27 ↗"), ('Mar27', "Mar '27 ↗")]
for idx, (m_tag, m_lbl) in enumerate(months_27):
    r = 25 + idx
    s_dt4_27 = f"DT-4 {m_tag}"
    style_dash_cell(ws_dash, r, 8, m_lbl, font=FONT_LINK, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, hyperlink_loc=f"'{s_dt4_27}'!A1")
    style_dash_cell(ws_dash, r, 9, f'=COUNTIF(\'{s_dt4_27}\'!$P$3:$P$500, "<>")', font=BOLD_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 10, f'=COUNTIF(\'{s_dt4_27}\'!$P$3:$P$500, "Confirmed (Done)")', font=GREEN_FONT, fill=GREEN_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 11, f'=COUNTIF(\'{s_dt4_27}\'!$P$3:$P$500, "Pending Calendar Booking")', font=YELLOW_FONT, fill=YELLOW_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 12, f'=COUNTIF(\'{s_dt4_27}\'!$P$3:$P$500, "Canceled")', font=RED_FONT, fill=RED_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 13, f'=SUM(\'{s_dt4_27}\'!$G$3:$G$500)', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

# DT-4 Subtotal (Row 28)
ws_dash.row_dimensions[28].height = 22
style_dash_cell(ws_dash, 28, 8, "DT-4 2026 Subtotal", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_cell(ws_dash, 28, 9, "=SUM(I13:I24)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 28, 10, "=SUM(J13:J24)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 28, 11, "=SUM(K13:K24)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 28, 12, "=SUM(L13:L24)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 28, 13, "=SUM(M13:M27)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

# Summary Mini Boxes (Rows 27-30)
style_dash_cell(ws_dash, 27, 15, "Event Scale Metric", font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_cell(ws_dash, 27, 16, "Summary", font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_cell(ws_dash, 28, 15, "🏆 Events Conducted", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
style_dash_cell(ws_dash, 28, 16, "=E45", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 29, 15, "🌟 Big Events (100+)", font=Font(name=FONT_FAMILY, size=10, bold=True, color='1F4E79'), fill=BIG_EVENT_FILL, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
style_dash_cell(ws_dash, 29, 16, "=P25", font=Font(name=FONT_FAMILY, size=10, bold=True, color='1F4E79'), fill=BIG_EVENT_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 30, 15, "👥 Small Events (<100)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
style_dash_cell(ws_dash, 30, 16, "=Q25", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

style_dash_cell(ws_dash, 27, 18, "Feedback Breakdown", font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_cell(ws_dash, 27, 19, "Count", font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_cell(ws_dash, 28, 18, "🏢 DT-3 Feedback Sent", font=FEEDBACK_SENT_FONT, fill=FEEDBACK_SENT_FILL, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
style_dash_cell(ws_dash, 28, 19, "=R25", font=FEEDBACK_SENT_FONT, fill=FEEDBACK_SENT_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 29, 18, "🏢 DT-4 Feedback Sent", font=FEEDBACK_SENT_FONT, fill=FEEDBACK_SENT_FILL, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
style_dash_cell(ws_dash, 29, 19, "=S25", font=FEEDBACK_SENT_FONT, fill=FEEDBACK_SENT_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 30, 18, "📨 Total Feedback Sent", font=Font(name=FONT_FAMILY, size=10, bold=True, color='1B365D'), fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
style_dash_cell(ws_dash, 30, 19, "=R25+S25", font=Font(name=FONT_FAMILY, size=10, bold=True, color='1B365D'), fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

# Combined Monthly Summary (Rows 31-45, Cols A to I)
ws_dash.row_dimensions[31].height = 26
ws_dash.row_dimensions[32].height = 24

ws_dash.merge_cells('A31:I31')
style_dash_cell(ws_dash, 31, 1, "📊 2026 COMBINED MONTHLY EVENTS SUMMARY", font=FONT_SECTION_HEADER, fill=NAVY_DARK_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 31, 1, 31, 9, fill=NAVY_DARK_FILL, border=BLACK_BORDER)

comb_headers = ['Month', 'DT-3 Events', 'DT-4 Events', 'Total Events', '🟢 Confirmed', '🟡 Pending', '🔴 Canceled', 'Confirmed %', 'Total Pax']
for idx, h in enumerate(comb_headers, 1):
    style_dash_cell(ws_dash, 32, idx, h, font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)

for idx, m_lbl in enumerate(month_short):
    r = 33 + idx
    ref_r = 13 + idx
    ws_dash.row_dimensions[r].height = 20
    style_dash_cell(ws_dash, r, 1, m_lbl, font=BOLD_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
    style_dash_cell(ws_dash, r, 2, f'=B{ref_r}', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 3, f'=I{ref_r}', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 4, f'=B{r}+C{r}', font=BOLD_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 5, f'=C{ref_r}+J{ref_r}', font=GREEN_FONT, fill=GREEN_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 6, f'=D{ref_r}+K{ref_r}', font=YELLOW_FONT, fill=YELLOW_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 7, f'=E{ref_r}+L{ref_r}', font=RED_FONT, fill=RED_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 8, f'=IFERROR(E{r}/(E{r}+F{r}), 0)', font=BOLD_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='0.0%')
    style_dash_cell(ws_dash, r, 9, f'=F{ref_r}+M{ref_r}', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

# Combined Total (Row 45)
ws_dash.row_dimensions[45].height = 22
style_dash_cell(ws_dash, 45, 1, "Full Year 2026 Total", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_cell(ws_dash, 45, 2, "=SUM(B33:B44)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 45, 3, "=SUM(C33:C44)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 45, 4, "=SUM(D33:D44)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 45, 5, "=SUM(E33:E44)", font=GREEN_FONT, fill=GREEN_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 45, 6, "=SUM(F33:F44)", font=YELLOW_FONT, fill=YELLOW_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 45, 7, "=SUM(G33:G44)", font=RED_FONT, fill=RED_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 45, 8, "=IFERROR(E45/(E45+F45), 0)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='0.0%')
style_dash_cell(ws_dash, 45, 9, "=SUM(I33:I44)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

ws_dash.row_dimensions[46].height = 10

# Interactive Month Deep Dive (Cols K to S)
ws_dash.merge_cells('K31:S31')
style_dash_cell(ws_dash, 31, 11, "🔍 INTERACTIVE MONTH SELECTION & DEEP DIVE", font=FONT_SECTION_HEADER, fill=FILL_ROYAL_BLUE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 31, 11, 31, 19, fill=FILL_ROYAL_BLUE, border=BLACK_BORDER)

ws_dash.merge_cells('K32:L32')
style_dash_cell(ws_dash, 32, 11, "📅 SELECT MONTH:", font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 32, 11, 32, 12, fill=FILL_NAVY_SLATE, border=BLACK_BORDER)

ws_dash.merge_cells('M32:O32')
style_dash_cell(ws_dash, 32, 13, "Jan", font=Font(name=FONT_FAMILY, size=12, bold=True, color='1B365D'), fill=FILL_DROPDOWN, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 32, 13, 32, 15, fill=FILL_DROPDOWN, border=BLACK_BORDER)

ws_dash.merge_cells('P32:S32')
style_dash_cell(ws_dash, 32, 16, "🟢 Live Auto-Calculated Drilldown", font=Font(name=FONT_FAMILY, size=10, bold=True, color='276A3C'), fill=FILL_SOFT_GREEN, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 32, 16, 32, 19, fill=FILL_SOFT_GREEN, border=BLACK_BORDER)

dv_month = DataValidation(type="list", formula1='"Full Year 2026,Jan,Feb,Mar,Apr,May,Jun,Jul,Aug,Sep,Oct,Nov,Dec"', allow_blank=False)
dv_month.add("M32")
ws_dash.add_data_validation(dv_month)

ws_dash.row_dimensions[33].height = 14
ws_dash.row_dimensions[34].height = 24
ws_dash.row_dimensions[35].height = 14
ws_dash.row_dimensions[36].height = 24
ws_dash.row_dimensions[37].height = 22

ws_dash.merge_cells('K33:L33')
style_dash_cell(ws_dash, 33, 11, "TOTAL EVENTS", font=FONT_CARD_TITLE, fill=FILL_CARD_GRAY, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 33, 11, 33, 12, fill=FILL_CARD_GRAY, border=BLACK_BORDER)

ws_dash.merge_cells('K34:L34')
style_dash_cell(ws_dash, 34, 11, '=IF($M$32="Full Year 2026", D45, IFERROR(INDEX($D$33:$D$44, MATCH($M$32, $A$33:$A$44, 0)), 0))', font=Font(name=FONT_FAMILY, size=14, bold=True, color='1B365D'), fill=FILL_CARD_GRAY, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_range(ws_dash, 34, 11, 34, 12, fill=FILL_CARD_GRAY, border=BLACK_BORDER)

ws_dash.merge_cells('M33:O33')
style_dash_cell(ws_dash, 33, 13, "CONFIRMED (DONE)", font=FONT_TABLE_HEADER, fill=GREEN_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 33, 13, 33, 15, fill=GREEN_FILL, border=BLACK_BORDER)

ws_dash.merge_cells('M34:O34')
style_dash_cell(ws_dash, 34, 13, '=IF($M$32="Full Year 2026", E45, IFERROR(INDEX($E$33:$E$44, MATCH($M$32, $A$33:$A$44, 0)), 0))', font=Font(name=FONT_FAMILY, size=14, bold=True, color='FFFFFF'), fill=GREEN_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_range(ws_dash, 34, 13, 34, 15, fill=GREEN_FILL, border=BLACK_BORDER)

ws_dash.merge_cells('P33:R33')
style_dash_cell(ws_dash, 33, 16, "PENDING BOOKING", font=YELLOW_FONT, fill=YELLOW_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 33, 16, 33, 18, fill=YELLOW_FILL, border=BLACK_BORDER)

ws_dash.merge_cells('P34:R34')
style_dash_cell(ws_dash, 34, 16, '=IF($M$32="Full Year 2026", F45, IFERROR(INDEX($F$33:$F$44, MATCH($M$32, $A$33:$A$44, 0)), 0))', font=Font(name=FONT_FAMILY, size=14, bold=True, color='000000'), fill=YELLOW_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_range(ws_dash, 34, 16, 34, 18, fill=YELLOW_FILL, border=BLACK_BORDER)

style_dash_cell(ws_dash, 33, 19, "CANCELED", font=FONT_TABLE_HEADER, fill=RED_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_cell(ws_dash, 34, 19, '=IF($M$32="Full Year 2026", G45, IFERROR(INDEX($G$33:$G$44, MATCH($M$32, $A$33:$A$44, 0)), 0))', font=Font(name=FONT_FAMILY, size=14, bold=True, color='FFFFFF'), fill=RED_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

ws_dash.merge_cells('K35:L35')
style_dash_cell(ws_dash, 35, 11, "CONFIRMED % RATE", font=FONT_CARD_TITLE, fill=FILL_CARD_GRAY, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 35, 11, 35, 12, fill=FILL_CARD_GRAY, border=BLACK_BORDER)

ws_dash.merge_cells('K36:L36')
style_dash_cell(ws_dash, 36, 11, '=IF($M$32="Full Year 2026", H45, IFERROR(INDEX($H$33:$H$44, MATCH($M$32, $A$33:$A$44, 0)), 0))', font=Font(name=FONT_FAMILY, size=13, bold=True, color='1B365D'), fill=FILL_CARD_GRAY, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='0.0%')
style_dash_range(ws_dash, 36, 11, 36, 12, fill=FILL_CARD_GRAY, border=BLACK_BORDER)

ws_dash.merge_cells('M35:O35')
style_dash_cell(ws_dash, 35, 13, "TOTAL ATTENDEES (PAX)", font=FONT_TABLE_HEADER, fill=FILL_PURPLE_CARD, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 35, 13, 35, 15, fill=FILL_PURPLE_CARD, border=BLACK_BORDER)

ws_dash.merge_cells('M36:O36')
style_dash_cell(ws_dash, 36, 13, '=IF($M$32="Full Year 2026", I45, IFERROR(INDEX($I$33:$I$44, MATCH($M$32, $A$33:$A$44, 0)), 0))', font=Font(name=FONT_FAMILY, size=14, bold=True, color='FFFFFF'), fill=FILL_PURPLE_CARD, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_range(ws_dash, 36, 13, 36, 15, fill=FILL_PURPLE_CARD, border=BLACK_BORDER)

ws_dash.merge_cells('P35:Q35')
style_dash_cell(ws_dash, 35, 16, "DT-3 EVENTS", font=FONT_CARD_TITLE, fill=FILL_CARD_GRAY, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 35, 16, 35, 17, fill=FILL_CARD_GRAY, border=BLACK_BORDER)

ws_dash.merge_cells('P36:Q36')
style_dash_cell(ws_dash, 36, 16, '=IF($M$32="Full Year 2026", B45, IFERROR(INDEX($B$33:$B$44, MATCH($M$32, $A$33:$A$44, 0)), 0))', font=Font(name=FONT_FAMILY, size=13, bold=True, color='1B365D'), fill=FILL_CARD_GRAY, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_range(ws_dash, 36, 16, 36, 17, fill=FILL_CARD_GRAY, border=BLACK_BORDER)

ws_dash.merge_cells('R35:S35')
style_dash_cell(ws_dash, 35, 18, "DT-4 EVENTS", font=FONT_CARD_TITLE, fill=FILL_CARD_GRAY, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 35, 18, 35, 19, fill=FILL_CARD_GRAY, border=BLACK_BORDER)

ws_dash.merge_cells('R36:S36')
style_dash_cell(ws_dash, 36, 18, '=IF($M$32="Full Year 2026", C45, IFERROR(INDEX($C$33:$C$44, MATCH($M$32, $A$33:$A$44, 0)), 0))', font=Font(name=FONT_FAMILY, size=13, bold=True, color='1B365D'), fill=FILL_CARD_GRAY, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_range(ws_dash, 36, 18, 36, 19, fill=FILL_CARD_GRAY, border=BLACK_BORDER)

# Row 37: Big Events, Small Events & Feedback Form Monthly Live Drilldown
ws_dash.merge_cells('K37:L37')
style_dash_cell(ws_dash, 37, 11, "🌟 BIG (100+):", font=Font(name=FONT_FAMILY, size=9, bold=True, color='1B365D'), fill=BIG_EVENT_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 37, 11, 37, 12, fill=BIG_EVENT_FILL, border=BLACK_BORDER)

style_dash_cell(ws_dash, 37, 13, '=IF($M$32="Full Year 2026", P25, IFERROR(INDEX($P$13:$P$24, MATCH($M$32, $A$33:$A$44, 0)), 0))', font=Font(name=FONT_FAMILY, size=13, bold=True, color='1B365D'), fill=BIG_EVENT_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

ws_dash.merge_cells('N37:O37')
style_dash_cell(ws_dash, 37, 14, "👥 SMALL (<100):", font=Font(name=FONT_FAMILY, size=9, bold=True, color='333333'), fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 37, 14, 37, 15, fill=FILL_TOTAL_ROW, border=BLACK_BORDER)

style_dash_cell(ws_dash, 37, 16, '=IF($M$32="Full Year 2026", Q25, IFERROR(INDEX($Q$13:$Q$24, MATCH($M$32, $A$33:$A$44, 0)), 0))', font=Font(name=FONT_FAMILY, size=13, bold=True, color='1B365D'), fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

ws_dash.merge_cells('Q37:R37')
style_dash_cell(ws_dash, 37, 17, "📨 DT3 / DT4 FB:", font=Font(name=FONT_FAMILY, size=8, bold=True, color='1F4E79'), fill=FEEDBACK_SENT_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 37, 17, 37, 18, fill=FEEDBACK_SENT_FILL, border=BLACK_BORDER)

style_dash_cell(ws_dash, 37, 19, '=IF($M$32="Full Year 2026", R25&"/"&S25, IFERROR(INDEX($R$13:$R$24, MATCH($M$32, $A$33:$A$44, 0))&"/"&INDEX($S$13:$S$24, MATCH($M$32, $A$33:$A$44, 0)), "0/0"))', font=Font(name=FONT_FAMILY, size=12, bold=True, color='1F4E79'), fill=FEEDBACK_SENT_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)

ws_dash.merge_cells('K38:S38')
style_dash_cell(ws_dash, 38, 11, "🚀 QUICK NAVIGATION TO MONTHLY TRACKERS", font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 38, 11, 38, 19, fill=FILL_NAVY_SLATE, border=BLACK_BORDER)

ws_dash.merge_cells('K39:S44')
help_text = "💡 HOW TO USE THIS DASHBOARD:\n1. Change the Month dropdown above (M32) to instantly update the Monthly Deep Dive KPIs.\n2. Click any Month in the DT-3 / DT-4 tables (Rows 13-27) to jump directly to that month's sheet.\n3. In any monthly sheet, click '🏠 ⮌ RETURN TO DASHBOARD' (A1) to return back here.\n4. Event Scale: Big Event (100+ pax) vs Small Event (<100 pax). Feedback forms recorded for previous entries are preserved; routine events are marked 'Not Required'."
style_dash_cell(ws_dash, 39, 11, help_text, font=Font(name=FONT_FAMILY, size=9, bold=False, color='1B365D'), fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)

# Facility Space & Venue Utilization (Cols A to D, Rows 47 to 58)
ws_dash.row_dimensions[47].height = 26
ws_dash.row_dimensions[48].height = 24

ws_dash.merge_cells('A47:D47')
style_dash_cell(ws_dash, 47, 1, "📍 FACILITY SPACE & VENUE UTILIZATION", font=FONT_SECTION_HEADER, fill=NAVY_DARK_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 47, 1, 47, 4, fill=NAVY_DARK_FILL, border=BLACK_BORDER)

venue_headers = ['Key Venue / Space', 'Building', 'Events', 'Total Pax']
for idx, h in enumerate(venue_headers, 1):
    style_dash_cell(ws_dash, 48, idx, h, font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)

venues = [
    ('Innovation Hub', 'Downtown-4', '*Innovation Hub*'),
    ('Cafeteria (L-1)', 'Downtown-4', '*Cafeteria*'),
    ('Hackable Area', 'Downtown-4', '*Hackable*'),
    ('Social Hub Areas', 'Downtown-4', '*Social Hub*'),
    ('Cafeteria (L-5)', 'Downtown-4', '*Cafeteria*'),
    ('Meeting Rooms (405, 506, VC)', 'Downtown-4', '*MR*'),
    ('Cafeteria Breakout', 'Downtown-3', '*Cafeteria*'),
    ('Breakout Area', 'Downtown-3', '*Breakout*'),
    ('Meeting Rooms (107, 305/306)', 'Downtown-3', '*MR*'),
]

for idx, (v_name, bldg, v_pattern) in enumerate(venues):
    r = 49 + idx
    ws_dash.row_dimensions[r].height = 20
    style_dash_cell(ws_dash, r, 1, v_name, font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
    style_dash_cell(ws_dash, r, 2, bldg, font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
    style_dash_cell(ws_dash, r, 3, f'=COUNTIFS(Master_Event_Tracker_2026!$H$3:$H$5000, "{v_pattern}", Master_Event_Tracker_2026!$F$3:$F$5000, "{bldg}")', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 4, f'=SUMIFS(Master_Event_Tracker_2026!$K$3:$K$5000, Master_Event_Tracker_2026!$H$3:$H$5000, "{v_pattern}", Master_Event_Tracker_2026!$F$3:$F$5000, "{bldg}")', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

ws_dash.row_dimensions[58].height = 22
ws_dash.merge_cells('A58:B58')
style_dash_cell(ws_dash, 58, 1, "Top Venues Subtotal", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 58, 1, 58, 2, fill=FILL_TOTAL_ROW, border=BLACK_BORDER)
style_dash_cell(ws_dash, 58, 3, "=SUM(C49:C57)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 58, 4, "=SUM(D49:D57)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

# Event Category & Purpose (Cols F to I, Rows 47 to 58)
ws_dash.merge_cells('F47:I47')
style_dash_cell(ws_dash, 47, 6, "🎯 EVENT CATEGORY & PURPOSE DISTRIBUTION", font=FONT_SECTION_HEADER, fill=NAVY_DARK_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 47, 6, 47, 9, fill=NAVY_DARK_FILL, border=BLACK_BORDER)

cat_headers = ['Event Category', 'Total Events', 'Confirmed', 'Total Pax']
for idx, h in enumerate(cat_headers, 6):
    style_dash_cell(ws_dash, 48, idx, h, font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)

categories = [
    'Townhall & All-Hands',
    'Training & Enablement',
    'Induction & Onboarding',
    'Assessment & Hiring',
    'Leadership & VIP Visit',
    'Social & Team Connect',
    'Tech & Innovation',
    'Dry Run & Rehearsal',
    'Team Meeting & Workshop'
]

for idx, cat_name in enumerate(categories):
    r = 49 + idx
    ws_dash.row_dimensions[r].height = 20
    style_dash_cell(ws_dash, r, 6, cat_name, font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
    style_dash_cell(ws_dash, r, 7, f'=COUNTIF(Master_Event_Tracker_2026!$J$3:$J$5000, "{cat_name}")', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 8, f'=COUNTIFS(Master_Event_Tracker_2026!$J$3:$J$5000, "{cat_name}", Master_Event_Tracker_2026!$S$3:$S$5000, "Confirmed (Done)")', font=GREEN_FONT, fill=GREEN_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 9, f'=SUMIF(Master_Event_Tracker_2026!$J$3:$J$5000, "{cat_name}", Master_Event_Tracker_2026!$K$3:$K$5000)', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

ws_dash.row_dimensions[58].height = 22
style_dash_cell(ws_dash, 58, 6, "Total Events Tracked", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_cell(ws_dash, 58, 7, "=SUM(G49:G57)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 58, 8, "=SUM(H49:H57)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 58, 9, "=SUM(I49:I57)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

# =============================================================================
# FEEDBACK FORM POLICY & APPLICABILITY RULES GUIDE (Cols K to S, Rows 47 to 58)
# =============================================================================
ws_dash.merge_cells('K47:S47')
style_dash_cell(ws_dash, 47, 11, "📜 FEEDBACK FORM POLICY & APPLICABILITY RULES GUIDE", font=FONT_SECTION_HEADER, fill=FILL_ROYAL_BLUE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 47, 11, 47, 19, fill=FILL_ROYAL_BLUE, border=BLACK_BORDER)

ws_dash.merge_cells('K48:L48')
style_dash_cell(ws_dash, 48, 11, "Event Category / Tier", font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 48, 11, 48, 12, fill=FILL_NAVY_SLATE, border=BLACK_BORDER)

ws_dash.merge_cells('M48:N48')
style_dash_cell(ws_dash, 48, 13, "Applicable Event Scope", font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 48, 13, 48, 14, fill=FILL_NAVY_SLATE, border=BLACK_BORDER)

style_dash_cell(ws_dash, 48, 15, "Pax Scale", font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)

ws_dash.merge_cells('P48:Q48')
style_dash_cell(ws_dash, 48, 16, "Feedback Form Rule", font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 48, 16, 48, 17, fill=FILL_NAVY_SLATE, border=BLACK_BORDER)

ws_dash.merge_cells('R48:S48')
style_dash_cell(ws_dash, 48, 18, "Operations Guidelines & Rationale", font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 48, 18, 48, 19, fill=FILL_NAVY_SLATE, border=BLACK_BORDER)

feedback_rules_data = [
    ("🌟 Tier 1: Major Townhalls", "All-Hands, Leadership Forums, Quarterly Townhalls", "100+ Pax", "📨 Form Sent (Mandatory)", "High stakeholder visibility & heavy AV/Cafeteria resource utilization."),
    ("🌟 Tier 1: Site Celebrations", "Independence Day, Women Empowerment, Cultural Meets", "100+ Pax", "📨 Form Sent (Mandatory)", "Measure employee engagement, catering satisfaction & event experience."),
    ("🌟 Tier 1: Strategic R&R", "Personal Banking R&R, PEB R&R, Annual Rewards", "70 - 150+", "📨 Form Sent (Mandatory)", "Recognize team achievements and track event organization feedback."),
    ("🎓 Tier 2: Tech Workshops", "Copilot Hacks, CSA Agentic AI, Gender Equity Forums", "25 - 80 Pax", "📨 Form Sent (Recommended)", "Evaluate curriculum depth, technical trainer effectiveness & lab setup."),
    ("🎓 Tier 2: Leadership Dev.", "People Leader Essentials, Think Like a Business, ALT", "20 - 50 Pax", "📨 Form Sent (Recommended)", "Track capability enhancement & leadership training effectiveness."),
    ("🎓 Tier 2: New Hire Induction", "New Joiners Induction, Grads Training, Immersion", "25 - 70 Pax", "📨 Form Sent (Recommended)", "Ensure seamless onboarding experience & facilities support feedback."),
    ("⚪ Tier 3: VIP / AU Visits", "AU Visitors & Senior Executive Visits (Sweta M, Allan Ha)", "1 - 5 Pax", "⚪ Not Required (Excluded)", "Escorted executive meetings; survey collection is NOT applicable."),
    ("⚪ Tier 3: Assessments", "Codility Assessments, Technical Tests, Hiring Drives", "20 - 50 Pax", "⚪ Not Required (Excluded)", "Recruitment technical assessments; feedback survey is NOT applicable."),
    ("⚪ Tier 3: Logistics & Routine", "Visitor Lunch setups, Bank Helpdesk, Daily Standups", "1 - 15 Pax", "⚪ Not Required (Excluded)", "Routine facilities operational tasks; feedback surveys are excluded."),
]

for idx, (cat_tier, scope_text, pax_str, rule_str, rationale_str) in enumerate(feedback_rules_data):
    r = 49 + idx
    ws_dash.row_dimensions[r].height = 20
    
    ws_dash.merge_cells(f'K{r}:L{r}')
    f_fill = BIG_EVENT_FILL if "Tier 1" in cat_tier else (FILL_CARD_GRAY if "Tier 2" in cat_tier else FILL_TOTAL_ROW)
    f_font = BIG_EVENT_FONT if "Tier 1" in cat_tier else (FONT_BOLD_NAVY if "Tier 2" in cat_tier else REGULAR_FONT)
    style_dash_cell(ws_dash, r, 11, cat_tier, font=f_font, fill=f_fill, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
    style_dash_range(ws_dash, r, 11, r, 12, fill=f_fill, border=BLACK_BORDER)

    ws_dash.merge_cells(f'M{r}:N{r}')
    style_dash_cell(ws_dash, r, 13, scope_text, font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
    style_dash_range(ws_dash, r, 13, r, 14, border=BLACK_BORDER)

    style_dash_cell(ws_dash, r, 15, pax_str, font=BOLD_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)

    ws_dash.merge_cells(f'P{r}:Q{r}')
    r_fill = FEEDBACK_SENT_FILL if "Mandatory" in rule_str or "Recommended" in rule_str else FEEDBACK_NOT_REQ_FILL
    r_font = FEEDBACK_SENT_FONT if "Mandatory" in rule_str or "Recommended" in rule_str else FEEDBACK_NOT_REQ_FONT
    style_dash_cell(ws_dash, r, 16, rule_str, font=r_font, fill=r_fill, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
    style_dash_range(ws_dash, r, 16, r, 17, fill=r_fill, border=BLACK_BORDER)

    ws_dash.merge_cells(f'R{r}:S{r}')
    style_dash_cell(ws_dash, r, 18, rationale_str, font=Font(name=FONT_FAMILY, size=8, color='333333'), border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
    style_dash_range(ws_dash, r, 18, r, 19, border=BLACK_BORDER)

ws_dash.row_dimensions[58].height = 22
ws_dash.merge_cells('K58:S58')
style_dash_cell(ws_dash, 58, 11, "💡 RULE SUMMARY: Feedback is required for Big Events (100+ Pax) & major workshops; routine syncs, AU visits & hiring tests are marked 'Not Required'.", font=Font(name=FONT_FAMILY, size=9, bold=True, color='1B365D'), fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
style_dash_range(ws_dash, 58, 11, 58, 19, fill=FILL_TOTAL_ROW, border=BLACK_BORDER)

ws_dash.row_dimensions[59].height = 10
ws_dash.row_dimensions[60].height = 6

# Priority Action Items (Cols A to I, Rows 61 to 72)
ws_dash.row_dimensions[61].height = 26
ws_dash.row_dimensions[62].height = 24

ws_dash.merge_cells('A61:I61')
style_dash_cell(ws_dash, 61, 1, "⚡ PRIORITY ACTION ITEMS: UPCOMING PENDING CALENDAR BOOKINGS (POST AUG 21)", font=FONT_SECTION_HEADER, fill=NAVY_DARK_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 61, 1, 61, 9, fill=NAVY_DARK_FILL, border=BLACK_BORDER)

action_headers = ['Event ID', 'Date', 'Day', 'Building', 'Location', 'Event Title', 'Pax', 'Booked By', 'Status']
for idx, h in enumerate(action_headers, 1):
    style_dash_cell(ws_dash, 62, idx, h, font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)

pending_future_sample = [
    ('EVT-2026-1117', datetime.date(2026, 8, 26), 'Wed', 'Downtown-4', 'L-5', 'Grads training', 50, 'Pragyasingh.kadyan@nab.com.au'),
    ('EVT-2026-1118', datetime.date(2026, 8, 26), 'Wed', 'Downtown-4', 'L-4', 'Grads training', 25, 'Ekta.sahrma@nab.com.au'),
    ('EVT-2026-1122', datetime.date(2026, 8, 26), 'Wed', 'Downtown-4', 'L-5', 'Potluck For Payments Ops', 80, 'Rupika.nagpal@nab.com.au'),
    ('EVT-2026-1126', datetime.date(2026, 8, 27), 'Thu', 'Downtown-4', 'L-5', 'Codility Assesment', 50, 'Rupika.nagpal@nab.com.au'),
    ('EVT-2026-1130', datetime.date(2026, 8, 28), 'Fri', 'Downtown-4', 'L-5', 'Codility Assesment', 50, 'Tarun.bajaj@nab.com.au'),
    ('EVT-2026-1135', datetime.date(2026, 8, 31), 'Mon', 'Downtown-4', 'L-5', 'Codility Assesment', 50, 'Tarun.bajaj@nab.com.au'),
    ('EVT-2026-1140', datetime.date(2026, 9, 1), 'Tue', 'Downtown-4', 'L-1', 'Townhall Session', 100, 'Operations Team'),
    ('EVT-2026-1145', datetime.date(2026, 9, 3), 'Thu', 'Downtown-4', 'L-5', 'Tech Enablement Workshop', 60, 'Tech Lead'),
    ('EVT-2026-1150', datetime.date(2026, 9, 7), 'Mon', 'Downtown-3', 'MR-305 & 306', 'New Joiners Induction', 40, 'Deshraj'),
    ('EVT-2026-1155', datetime.date(2026, 9, 10), 'Thu', 'Downtown-3', 'Cafeteria', 'Quarterly R&R Meet', 75, 'HR Team'),
]

for idx, (eid, dt, day, bldg, loc, title, pax, booked) in enumerate(pending_future_sample):
    r = 63 + idx
    ws_dash.row_dimensions[r].height = 20
    style_dash_cell(ws_dash, r, 1, eid, font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
    style_dash_cell(ws_dash, r, 2, dt, font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='dd-mmm-yyyy')
    style_dash_cell(ws_dash, r, 3, day, font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
    style_dash_cell(ws_dash, r, 4, bldg, font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
    style_dash_cell(ws_dash, r, 5, loc, font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
    style_dash_cell(ws_dash, r, 6, title, font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
    style_dash_cell(ws_dash, r, 7, pax, font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 8, booked, font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
    style_dash_cell(ws_dash, r, 9, "🟡 Pending Booking", font=YELLOW_FONT, fill=YELLOW_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)

# =============================================================================
# LIVE AUTOMATED FEEDBACK & EVENT SCALE PERFORMANCE (Cols K to S, Rows 61 to 72)
# =============================================================================
ws_dash.merge_cells('K61:S61')
style_dash_cell(ws_dash, 61, 11, "📊 LIVE AUTOMATED FEEDBACK & EVENT SCALE PERFORMANCE", font=FONT_SECTION_HEADER, fill=NAVY_DARK_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 61, 11, 61, 19, fill=NAVY_DARK_FILL, border=BLACK_BORDER)

ws_dash.merge_cells('K62:L62')
style_dash_cell(ws_dash, 62, 11, "Operations Breakdown", font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 62, 11, 62, 12, fill=FILL_NAVY_SLATE, border=BLACK_BORDER)

ws_dash.merge_cells('M62:N62')
style_dash_cell(ws_dash, 62, 13, "Total Events", font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 62, 13, 62, 14, fill=FILL_NAVY_SLATE, border=BLACK_BORDER)

style_dash_cell(ws_dash, 62, 15, "🟢 Done", font=FONT_TABLE_HEADER, fill=GREEN_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)

ws_dash.merge_cells('P62:Q62')
style_dash_cell(ws_dash, 62, 16, "📨 Form Sent", font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 62, 16, 62, 17, fill=FILL_NAVY_SLATE, border=BLACK_BORDER)

style_dash_cell(ws_dash, 62, 18, "⚪ Not Required", font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_cell(ws_dash, 62, 19, "Feedback %", font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)

auto_table_rows = [
    ("🌟 Big Events (100+ Pax)", '=COUNTIF(Master_Event_Tracker_2026!$L$3:$L$5000, "Big Event")', '=COUNTIFS(Master_Event_Tracker_2026!$L$3:$L$5000, "Big Event", Master_Event_Tracker_2026!$S$3:$S$5000, "Confirmed (Done)")', '=COUNTIFS(Master_Event_Tracker_2026!$L$3:$L$5000, "Big Event", Master_Event_Tracker_2026!$T$3:$T$5000, "Form sent")', '=COUNTIFS(Master_Event_Tracker_2026!$L$3:$L$5000, "Big Event", Master_Event_Tracker_2026!$T$3:$T$5000, "Not Required")', '=IFERROR(P63/O63, 0)'),
    ("👥 Small Events (<100 Pax)", '=COUNTIF(Master_Event_Tracker_2026!$L$3:$L$5000, "Small Event")', '=COUNTIFS(Master_Event_Tracker_2026!$L$3:$L$5000, "Small Event", Master_Event_Tracker_2026!$S$3:$S$5000, "Confirmed (Done)")', '=COUNTIFS(Master_Event_Tracker_2026!$L$3:$L$5000, "Small Event", Master_Event_Tracker_2026!$T$3:$T$5000, "Form sent")', '=COUNTIFS(Master_Event_Tracker_2026!$L$3:$L$5000, "Small Event", Master_Event_Tracker_2026!$T$3:$T$5000, "Not Required")', '=IFERROR(P64/O64, 0)'),
    ("🏢 Downtown-3 Operations", '=COUNTIF(Master_Event_Tracker_2026!$F$3:$F$5000, "Downtown-3")', '=COUNTIFS(Master_Event_Tracker_2026!$F$3:$F$5000, "Downtown-3", Master_Event_Tracker_2026!$S$3:$S$5000, "Confirmed (Done)")', '=COUNTIFS(Master_Event_Tracker_2026!$F$3:$F$5000, "Downtown-3", Master_Event_Tracker_2026!$T$3:$T$5000, "Form sent")', '=COUNTIFS(Master_Event_Tracker_2026!$F$3:$F$5000, "Downtown-3", Master_Event_Tracker_2026!$T$3:$T$5000, "Not Required")', '=IFERROR(P65/O65, 0)'),
    ("🏢 Downtown-4 Operations", '=COUNTIF(Master_Event_Tracker_2026!$F$3:$F$5000, "Downtown-4")', '=COUNTIFS(Master_Event_Tracker_2026!$F$3:$F$5000, "Downtown-4", Master_Event_Tracker_2026!$S$3:$S$5000, "Confirmed (Done)")', '=COUNTIFS(Master_Event_Tracker_2026!$F$3:$F$5000, "Downtown-4", Master_Event_Tracker_2026!$T$3:$T$5000, "Form sent")', '=COUNTIFS(Master_Event_Tracker_2026!$F$3:$F$5000, "Downtown-4", Master_Event_Tracker_2026!$T$3:$T$5000, "Not Required")', '=IFERROR(P66/O66, 0)'),
    ("🏆 Full Year Combined Total", '=COUNTIF(Master_Event_Tracker_2026!$L$3:$L$5000, "<>")', '=COUNTIF(Master_Event_Tracker_2026!$S$3:$S$5000, "Confirmed (Done)")', '=COUNTIF(Master_Event_Tracker_2026!$T$3:$T$5000, "Form sent")', '=COUNTIF(Master_Event_Tracker_2026!$T$3:$T$5000, "Not Required")', '=IFERROR(P67/O67, 0)'),
]

for idx, (lbl, f_tot, f_done, f_sent, f_not_req, f_pct) in enumerate(auto_table_rows):
    r = 63 + idx
    ws_dash.row_dimensions[r].height = 20
    is_tot = (idx == len(auto_table_rows) - 1)
    
    ws_dash.merge_cells(f'K{r}:L{r}')
    row_fill = FILL_TOTAL_ROW if is_tot else (BIG_EVENT_FILL if "Big Events" in lbl else NO_FILL)
    row_font = FONT_BOLD_NAVY if is_tot else (BIG_EVENT_FONT if "Big Events" in lbl else REGULAR_FONT)
    style_dash_cell(ws_dash, r, 11, lbl, font=row_font, fill=row_fill, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
    style_dash_range(ws_dash, r, 11, r, 12, fill=row_fill, border=BLACK_BORDER)

    ws_dash.merge_cells(f'M{r}:N{r}')
    style_dash_cell(ws_dash, r, 13, f_tot, font=BOLD_FONT, fill=row_fill, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_range(ws_dash, r, 13, r, 14, fill=row_fill, border=BLACK_BORDER)

    style_dash_cell(ws_dash, r, 15, f_done, font=GREEN_FONT, fill=GREEN_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

    ws_dash.merge_cells(f'P{r}:Q{r}')
    style_dash_cell(ws_dash, r, 16, f_sent, font=FEEDBACK_SENT_FONT, fill=FEEDBACK_SENT_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_range(ws_dash, r, 16, r, 17, fill=FEEDBACK_SENT_FILL, border=BLACK_BORDER)

    style_dash_cell(ws_dash, r, 18, f_not_req, font=REGULAR_FONT, fill=row_fill, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 19, f_pct, font=BOLD_FONT, fill=row_fill, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='0.0%')

for r in range(68, 73):
    ws_dash.row_dimensions[r].height = 20
    ws_dash.merge_cells(f'K{r}:S{r}')
    style_dash_cell(ws_dash, r, 11, "", font=REGULAR_FONT, fill=NO_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
    style_dash_range(ws_dash, r, 11, r, 19, fill=NO_FILL, border=BLACK_BORDER)

style_dash_cell(ws_dash, 68, 11, "📌 NOTE: All formulas on this dashboard update dynamically as new event records and feedback statuses are logged.", font=Font(name=FONT_FAMILY, size=8, italic=True, color='595959'), fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
style_dash_range(ws_dash, 68, 11, 68, 19, fill=FILL_TOTAL_ROW, border=BLACK_BORDER)

# Column Widths
dash_col_widths = {
    'A': 20.0, 'B': 14.0, 'C': 14.0, 'D': 14.0, 'E': 14.0, 'F': 22.0, 'G': 14.0, 'H': 20.0, 'I': 14.0,
    'J': 14.0, 'K': 15.0, 'L': 15.0, 'M': 14.0, 'N': 14.0, 'O': 18.0, 'P': 16.0, 'Q': 18.0, 'R': 16.0, 'S': 20.0
}
for col_l, w in dash_col_widths.items():
    ws_dash.column_dimensions[col_l].width = w

if hasattr(ws_dash, 'views') and ws_dash.views and ws_dash.views.sheetView:
    ws_dash.views.sheetView[0].showGridLines = True
elif hasattr(ws_dash, 'sheet_view'):
    ws_dash.sheet_view.showGridLines = True

# Save Workbook
saved = False
for attempt in range(5):
    try:
        wb.save(EXCEL_FILE)
        print(f"\n Master Event Tracker & Colorful Dashboard successfully synchronized and saved to: {EXCEL_FILE}")
        saved = True
        break
    except PermissionError:
        print(f" [RETRY {attempt+1}/5] '{EXCEL_FILE}' is currently open in Excel. Waiting 2 seconds...")
        time.sleep(2)
    except Exception as e:
        print(f" Error saving: {e}")

if not saved:
    alt_file = "Yearly Event Tracker- DT-3 and DT-4_Updated.xlsx"
    wb.save(alt_file)
    print(f"\n [INFO] Saved to '{alt_file}'")
