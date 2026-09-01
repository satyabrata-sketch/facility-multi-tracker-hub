"""
=============================================================================
ADD SPACE VENUE AND EVENT CATEGORY TABLES TO SIMPLIFIED DASHBOARD
=============================================================================
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.utils import get_column_letter, column_index_from_string
import datetime
import re
import os
import sys
import subprocess
import time

sys.stdout.reconfigure(encoding='utf-8')

SOURCE_FILE = "Yearly Event Tracker- DT-3 and DT-4 Updated.xlsx"
OUTPUT_FILE = "Yearly Event Tracker- DT-3 and DT-4 Simplified Master.xlsx"

if not os.path.exists(SOURCE_FILE):
    for alt in ["Yearly Event Tracker- DT-3 and DT-4 Latest(19th august).xlsx", "Yearly Event Tracker- DT-3 and DT-4.xlsx"]:
        if os.path.exists(alt):
            SOURCE_FILE = alt
            break

print(f"Loading source workbook: {SOURCE_FILE} ...")
temp_f = "temp_source_tracker_venues.xlsx"
try:
    # Use PowerShell to safely copy open Excel files
    subprocess.run(["powershell", "-Command", f"Copy-Item -Path '{SOURCE_FILE}' -Destination '{temp_f}' -Force"], capture_output=True)
    wb = openpyxl.load_workbook(temp_f)
    try: os.remove(temp_f)
    except: pass
except Exception as e:
    print(f"Loading directly due to: {e}")
    wb = openpyxl.load_workbook(SOURCE_FILE)

# Styling Definitions
FONT_FAMILY = "Segoe UI"
REGULAR_FONT = Font(name=FONT_FAMILY, size=10, color='000000')
BOLD_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color='000000')
HEADER_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color='FFFFFF')
BANNER_TITLE_FONT = Font(name=FONT_FAMILY, size=11, bold=True, color='FFFFFF')
HOME_BTN_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color='FFFFFF', underline='single')

FONT_MAIN_TITLE = Font(name=FONT_FAMILY, size=15, bold=True, color='FFFFFF')
FONT_SUB_TITLE = Font(name=FONT_FAMILY, size=10, bold=False, color='DCE6F1')
FONT_SECTION_HEADER = Font(name=FONT_FAMILY, size=11, bold=True, color='FFFFFF')
FONT_TABLE_HEADER = Font(name=FONT_FAMILY, size=10, bold=True, color='FFFFFF')
FONT_CARD_TITLE = Font(name=FONT_FAMILY, size=9, bold=True, color='595959')
FONT_CARD_SUB = Font(name=FONT_FAMILY, size=8, bold=False, color='7F7F7F')
FONT_BOLD_NAVY = Font(name=FONT_FAMILY, size=10, bold=True, color='1B365D')
FONT_LINK = Font(name=FONT_FAMILY, size=10, bold=True, color='0055AA', underline='single')
FONT_LINK_HIST = Font(name=FONT_FAMILY, size=9, bold=True, color='004488', underline='single')
FONT_NAV_BTN = Font(name=FONT_FAMILY, size=11, bold=True, color='FFFFFF', underline='single')

FONT_KPI_NAVY = Font(name=FONT_FAMILY, size=18, bold=True, color='1B365D')
FONT_KPI_WHITE = Font(name=FONT_FAMILY, size=18, bold=True, color='FFFFFF')
FONT_KPI_BLACK = Font(name=FONT_FAMILY, size=18, bold=True, color='000000')
FONT_KPI_BLUE = Font(name=FONT_FAMILY, size=18, bold=True, color='1F4E79')

NAVY_DARK_FILL = PatternFill(start_color='1B365D', end_color='1B365D', fill_type='solid')
NAVY_LIGHT_FILL = PatternFill(start_color='2C4D75', end_color='2C4D75', fill_type='solid')
FILL_NAVY_SLATE = PatternFill(start_color='2C4D75', end_color='2C4D75', fill_type='solid')
FILL_ROYAL_BLUE = PatternFill(start_color='204A87', end_color='204A87', fill_type='solid')
FILL_BANNER_DARK = PatternFill(start_color='0F2537', end_color='0F2537', fill_type='solid')
FILL_CARD_GRAY = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
FILL_TOTAL_ROW = PatternFill(start_color='E9EEF4', end_color='E9EEF4', fill_type='solid')
FILL_TEAL_BTN = PatternFill(start_color='008080', end_color='008080', fill_type='solid')
FILL_PURPLE_CARD = PatternFill(start_color='4A69BD', end_color='4A69BD', fill_type='solid')
FILL_DROPDOWN = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
FILL_SOFT_GREEN = PatternFill(start_color='E2F0D9', end_color='E2F0D9', fill_type='solid')
FILL_SOFT_ORANGE = PatternFill(start_color='FDEBD0', end_color='FDEBD0', fill_type='solid')
FILL_SOFT_BLUE = PatternFill(start_color='EBF5FB', end_color='EBF5FB', fill_type='solid')
FILL_SOFT_YELLOW = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')

# Status Colors
GREEN_FILL = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
GREEN_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color='FFFFFF')

YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
YELLOW_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color='000000')

RED_FILL = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
RED_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color='FFFFFF')

# Type & Feedback Colors
FEEDBACK_SENT_FILL = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
FEEDBACK_SENT_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color='1F4E79')
REQUEST_FILL = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
REQUEST_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color='856404')
EVENT_FILL = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
EVENT_FONT = Font(name=FONT_FAMILY, size=10, bold=False, color='000000')

NO_FILL = PatternFill(fill_type=None)

# Borders & Alignment
BLACK_THIN = Side(style='thin', color='000000')
BLACK_BORDER = Border(left=BLACK_THIN, right=BLACK_THIN, top=BLACK_THIN, bottom=BLACK_THIN)
ALIGN_CENTER_WRAP = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)

MONTH_ORDER = {'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6, 'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12}
MONTH_MAP = {'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6, 'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12, 'ap': 4}

def parse_date(val, default_year=2026):
    if val is None:
        return None
    if isinstance(val, (datetime.date, datetime.datetime)):
        return val.date() if isinstance(val, datetime.datetime) else val
    s = str(val).strip()
    if not s or s.lower() in ('na', 'none', '-', 'holiday', 'nl', 'no'):
        return None
    m = re.search(r'(\d{4})-(\d{1,2})-(\d{1,2})', s)
    if m:
        try:
            return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except: pass
    m = re.search(r'(\d{1,2})[/.-](\d{1,2})[/.-](\d{4})', s)
    if m:
        try:
            return datetime.date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except: pass
    m = re.search(r'(\d{1,2})(?:st|nd|rd|th)?[\s\'\-_]*([a-zA-Z]+)[\s\'\-_]*(\d{2,4})?', s)
    if m:
        day = int(m.group(1))
        mon_str = m.group(2).lower()[:3]
        mon = MONTH_MAP.get(mon_str)
        yr_str = m.group(3)
        if yr_str:
            yr = int(yr_str)
            if yr < 100: yr += 2000
        else:
            yr = default_year
        if mon:
            try:
                return datetime.date(yr, mon, day)
            except: pass
    return None

def format_date_val(val, default_year=2026):
    parsed = parse_date(val, default_year)
    return parsed if parsed else str(val or '').strip()

def parse_time_str(val):
    if val is None:
        return None
    if isinstance(val, datetime.time):
        return val.strftime('%I:%M %p')
    if isinstance(val, datetime.datetime):
        return val.strftime('%I:%M %p')
    if isinstance(val, (int, float)):
        if 0 <= val < 1:
            total_seconds = int(round(val * 86400))
            hours = (total_seconds // 3600) % 24
            minutes = (total_seconds % 3600) // 60
            ampm = 'PM' if hours >= 12 else 'AM'
            h12 = hours % 12
            if h12 == 0: h12 = 12
            return f"{h12:02d}:{minutes:02d} {ampm}"
        val = str(val)
    
    s = str(val).strip()
    if not s or s.lower() in ('na', 'none', '-', 'no', 'nl', 'yes'):
        return None
    if s.lower() in ('holiday', 'tbc', 'full day', 'all day'):
        return s.title()
        
    s_clean = s.replace(';', ':').replace('o', '0').replace('O', '0').replace('.', ':')
    s_clean = re.sub(r':([apAP][mM])', r' \1', s_clean)
    s_clean = re.sub(r'\s*:\s*', ':', s_clean)
    
    m = re.search(r'(\d{1,2}):(\d{2})(?::\d{2})?\s*([apAP][mM])?', s_clean)
    if m:
        hr = int(m.group(1))
        mn = int(m.group(2))
        ampm = m.group(3)
        if ampm:
            ampm = ampm.upper()
            if hr > 12: hr -= 12
            elif hr == 0: hr = 12
            return f"{hr:02d}:{mn:02d} {ampm}"
        else:
            if hr >= 12:
                ampm = 'PM'
                if hr > 12: hr -= 12
            else:
                ampm = 'AM'
                if hr == 0: hr = 12
            return f"{hr:02d}:{mn:02d} {ampm}"
            
    m2 = re.search(r'(\d{1,2})\s*([apAP][a-zA-Z]*)', s_clean)
    if m2:
        hr = int(m2.group(1))
        ampm_str = m2.group(2).upper()
        ampm = 'PM' if 'P' in ampm_str else 'AM'
        if hr > 12: hr -= 12
        elif hr == 0: hr = 12
        return f"{hr:02d}:00 {ampm}"
        
    m3 = re.search(r'(\d{2})(\d{2})\s*([apAP][mM])', s_clean)
    if m3:
        hr = int(m3.group(1))
        mn = int(m3.group(2))
        ampm = m3.group(3).upper()
        if hr > 12: hr -= 12
        elif hr == 0: hr = 12
        return f"{hr:02d}:{mn:02d} {ampm}"
        
    return None

def parse_day_str(val, date_obj=None):
    if date_obj and isinstance(date_obj, (datetime.date, datetime.datetime)):
        return date_obj.strftime('%a')
    if not val:
        return ''
    s = str(val).strip().lower()
    day_map = {
        'mon': 'Mon', 'monday': 'Mon',
        'tue': 'Tue', 'tues': 'Tue', 'tuesday': 'Tue',
        'wed': 'Wed', 'wednesday': 'Wed',
        'thu': 'Thu', 'thur': 'Thu', 'thurs': 'Thu', 'thursday': 'Thu',
        'fri': 'Fri', 'friday': 'Fri',
        'sat': 'Sat', 'saturday': 'Sat',
        'sun': 'Sun', 'sunday': 'Sun'
    }
    return day_map.get(s, str(val).strip())

def parse_capacity(val):
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip()
    if not s or s.lower() in ('na', 'none', '-', 'tbc', 'nl', 'no'):
        return 0
    m = re.search(r'(\d+)', s)
    if m:
        return int(m.group(1))
    return 0

def is_request_entry(title, loc, rem, fb, log, pax):
    t_lower = str(title or '').lower().strip()
    loc_lower = str(loc or '').lower().strip()
    rem_lower = str(rem or '').lower().strip()
    fb_lower = str(fb or '').lower().strip()
    log_lower = str(log or '').lower().strip()
    combined = f"{t_lower} | {loc_lower} | {rem_lower} | {fb_lower} | {log_lower}"
    
    if pax >= 10:
        if any(w in combined for w in ['townhall', 'all-hands', 'all hands', 'induction', 'training', 'workshop', 'session', 'hiring drive', 'celebration', 'annual day', 'ceremony', 'connect', 'assessment', 'felicitation', 'potluck', 'pot luck']):
            return False

    pure_req_titles = [
        'water bottle', 'water bottles', 'tissue box', 'tissue paper', 'tissues', 
        'sanitizer', 'diet coke', 'markers', 'white board marker', 'extension cord',
        'power extension', 'chair setup only', 'chairs only', 'notepad chairs only',
        'dustbin', 'cleaning', 'housekeeping', 'stationery', 'arrangement support for visitor lunch',
        'pantry support', 'arrnange water', 'arrange water', 'supplies', 'cutlery'
    ]
    for pr in pure_req_titles:
        if pr in t_lower:
            return True
            
    if any(k in t_lower for k in ['au visitor', 'visitor', 'visitors', 'visit']) and pax <= 3:
        if any(w in rem_lower for w in ['tissue', 'water bottle', 'coke', 'sanitizer', 'check in mr', 'check in by', 'candies', 'mentos', 'sticky noted', 'offline singage', 'offline note', 'offline message', 'replenish']):
            return True
        if not rem_lower and not any(w in t_lower for w in ['training', 'townhall', 'meeting', 'workshop', 'session']):
            return True

    if not t_lower and any(w in rem_lower for w in ['water bottles', 'tissue box', 'cutlery', 'sanitizer']):
        if pax <= 5:
            return True

    return False

def standardize_status(status_val):
    if not status_val:
        return 'Pending Calendar Booking'
    s = str(status_val).strip().lower()
    if 'confirm' in s or 'done' in s or 'booked' in s or 'completed' in s:
        return 'Confirmed (Done)'
    elif 'cancel' in s or 'reject' in s:
        return 'Canceled'
    else:
        return 'Pending Calendar Booking'

def standardize_feedback(feedback_val, status_val=None):
    if not feedback_val:
        return 'NA'
    s = str(feedback_val).strip().lower()
    if 'sent' in s or 'yes' in s or 'done' in s:
        return 'Form sent'
    else:
        return 'NA'

def categorize_event(title, loc, remarks, req_type='Event'):
    if req_type == 'Request':
        return 'Facility & Hospitality Request'
    s = f"{title} {remarks} {loc}".lower()
    if any(w in s for w in ['townhall', 'town hall', 'all-hands', 'all hands', 'global call', 'q1 r&r', 'felicitation']):
        return 'Townhall & All-Hands'
    elif any(w in s for w in ['induction', 'onboarding', 'new joiner', 'orientation', 'welcome']):
        return 'Induction & Onboarding'
    elif any(w in s for w in ['training', 'workshop', 'learning', 'upskilling', 'bootcamp', 'program', 'ops academy', 'grads']):
        return 'Training & Enablement'
    elif any(w in s for w in ['assessment', 'hiring', 'interview', 'codility', 'recruitment', 'sourcing']):
        return 'Assessment & Hiring'
    elif any(w in s for w in ['vip', 'visitor', 'leadership', 'executive', 'client visit', 'director', 'sweta mehra', 'peter']):
        return 'Leadership & VIP Visit'
    elif any(w in s for w in ['potluck', 'pot luck', 'fun', 'lunch', 'celebration', 'festival', 'diwali', 'republic day', 'annual day', 'hi tea', 'hi-tea', 'connect']):
        return 'Social & Team Connect'
    elif any(w in s for w in ['gen ai', 'tech', 'ai session', 'hackathon', 'innovation', 'rendezvous']):
        return 'Tech & Innovation'
    elif any(w in s for w in ['dry run', 'rehearsal', 'audio visual check', 'av check', 'prep']):
        return 'Dry Run & Rehearsal'
    else:
        return 'Team Meeting & Workshop'

def style_and_merge(ws, start_col, end_col, val, font, fill, hyperlink_loc=None):
    for c in range(start_col, end_col + 1):
        cell = ws.cell(1, c)
        cell.font = font
        cell.fill = fill
        cell.border = BLACK_BORDER
        cell.alignment = ALIGN_CENTER_WRAP
        cell.value = None
    c_start = ws.cell(1, start_col)
    c_start.value = val
    if hyperlink_loc:
        c_start.hyperlink = Hyperlink(ref=f"{get_column_letter(start_col)}1", location=hyperlink_loc)
    if end_col > start_col:
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

print("Formatting monthly and historical sheets...")
monthly_sheets = [s for s in wb.sheetnames if s not in ('Dashboard', 'Master_Event_Tracker_2026')]

for sname in monthly_sheets:
    ws = wb[sname]
    is_dt3 = "DT-3" in sname
    is_dt4 = "DT-4" in sname
    
    raw_rows = []
    max_r = ws.max_row
    max_c = ws.max_column
    
    if max_r >= 2:
        for r in range(3, max_r + 1):
            row_data = {c: ws.cell(r, c).value for c in range(1, max_c + 1)}
            raw_rows.append(row_data)

    for m in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(m))
        
    ws.delete_rows(1, ws.max_row + 10)
    
    if is_dt3:
        headers = [
            "Event ID", "Event Date", "Day", "Location", "Type", "Event Scale",
            "Event Type / Title", "Capacity", "Event Start Time", "Event End Time",
            "F& B", "Logistic Support", "Booked By", "Requester Email", "Booking Date",
            "Status", "Feedback Form", "Remarks", "Booking Taken By"
        ]
        m_tag = sname.replace("DT-3", "").strip().upper()
        style_and_merge(ws, 1, 3, "🏠 ⮌ RETURN TO DASHBOARD", HOME_BTN_FONT, NAVY_DARK_FILL, "'Dashboard'!A1")
        style_and_merge(ws, 4, 14, f"DOWNTOWN-3 EVENT OPERATIONS TRACKER • {m_tag}", BANNER_TITLE_FONT, NAVY_LIGHT_FILL)
        style_and_merge(ws, 15, 19, "🟢 Green: Confirmed | 🟡 Yellow: Pending | 🔴 Red: Canceled", HEADER_FONT, NAVY_DARK_FILL)
    elif is_dt4:
        headers = [
            "Event ID", "Event Date", "Day", "Location", "Floor", "Type", "Event Scale",
            "Event Type / Title", "Capacity", "Event Start Time", "Event End Time",
            "F& B", "Logistic Support", "Booked By", "Requester Email", "Booking Date",
            "Status", "Feedback Form", "Remarks", "Booking Taken By"
        ]
        m_tag = sname.replace("DT-4", "").strip().upper()
        style_and_merge(ws, 1, 3, "🏠 ⮌ RETURN TO DASHBOARD", HOME_BTN_FONT, NAVY_DARK_FILL, "'Dashboard'!A1")
        style_and_merge(ws, 4, 15, f"DOWNTOWN-4 EVENT OPERATIONS TRACKER • {m_tag}", BANNER_TITLE_FONT, NAVY_LIGHT_FILL)
        style_and_merge(ws, 16, 20, "🟢 Green: Confirmed | 🟡 Yellow: Pending | 🔴 Red: Canceled", HEADER_FONT, NAVY_DARK_FILL)
    else:
        headers = [
            "Event ID", "Event Date", "Day", "Location", "Type", "Event Scale",
            "Event Type / Title", "Capacity", "Event Start Time", "Event End Time",
            "F& B", "Logistic Support", "Booked By", "Requester Email", "Booking Date",
            "Status", "Feedback Form", "Remarks", "Booking Taken By"
        ]
        style_and_merge(ws, 1, 3, "🏠 ⮌ RETURN TO DASHBOARD", HOME_BTN_FONT, NAVY_DARK_FILL, "'Dashboard'!A1")
        style_and_merge(ws, 4, 14, f"EVENT OPERATIONS TRACKER • {sname.upper()}", BANNER_TITLE_FONT, NAVY_LIGHT_FILL)
        style_and_merge(ws, 15, 19, "🟢 Green: Confirmed | 🟡 Yellow: Pending | 🔴 Red: Canceled", HEADER_FONT, NAVY_DARK_FILL)

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 28
    
    for col_idx, h_text in enumerate(headers, 1):
        cell = ws.cell(2, col_idx, h_text)
        cell.font = HEADER_FONT
        cell.fill = NAVY_DARK_FILL
        cell.alignment = ALIGN_CENTER_WRAP
        cell.border = BLACK_BORDER

    curr_r = 3
    for r_idx, r_dict in enumerate(raw_rows):
        if is_dt4:
            eid = r_dict.get(1)
            dt = r_dict.get(2)
            day = r_dict.get(3)
            loc = str(r_dict.get(4) or '').strip()
            floor = str(r_dict.get(5) or '').strip()
            if len(r_dict) >= 19 and r_dict.get(6) in ("Event", "Request"):
                type_val = r_dict.get(6)
                title = str(r_dict.get(8) or '').strip()
                cap = r_dict.get(9)
                st = r_dict.get(10)
                et = r_dict.get(11)
                fnb = str(r_dict.get(12) or '').strip()
                log = str(r_dict.get(13) or '').strip()
                booked = str(r_dict.get(14) or '').strip()
                email = str(r_dict.get(15) or '').strip()
                bdate = r_dict.get(16)
                status = r_dict.get(17)
                fb = r_dict.get(18)
                rem = str(r_dict.get(19) or '').strip()
                taken = str(r_dict.get(20) or '').strip()
            else:
                title = str(r_dict.get(6) or '').strip()
                cap = r_dict.get(7)
                st = r_dict.get(8)
                et = r_dict.get(9)
                fnb = str(r_dict.get(10) or '').strip()
                log = str(r_dict.get(11) or '').strip()
                booked = str(r_dict.get(12) or '').strip()
                email = str(r_dict.get(13) or '').strip()
                bdate = r_dict.get(14)
                status = r_dict.get(15)
                fb = r_dict.get(16)
                rem = str(r_dict.get(17) or '').strip()
                taken = str(r_dict.get(18) or '').strip()
                type_val = "Request" if is_request_entry(title, loc, rem, fnb, log, parse_capacity(cap)) else "Event"
        elif is_dt3:
            eid = r_dict.get(1)
            dt = r_dict.get(2)
            day = r_dict.get(3)
            loc = str(r_dict.get(4) or '').strip()
            floor = 'L-1' if '1' in loc else ('L-5' if '5' in loc else ('L-4' if '4' in loc else ('L-6' if '6' in loc else 'General')))
            if len(r_dict) >= 18 and r_dict.get(5) in ("Event", "Request"):
                type_val = r_dict.get(5)
                title = str(r_dict.get(7) or '').strip()
                cap = r_dict.get(8)
                st = r_dict.get(9)
                et = r_dict.get(10)
                fnb = str(r_dict.get(11) or '').strip()
                log = str(r_dict.get(12) or '').strip()
                booked = str(r_dict.get(13) or '').strip()
                email = str(r_dict.get(14) or '').strip()
                bdate = r_dict.get(15)
                status = r_dict.get(16)
                fb = r_dict.get(17)
                rem = str(r_dict.get(18) or '').strip()
                taken = str(r_dict.get(19) or '').strip()
            else:
                title = str(r_dict.get(5) or '').strip()
                cap = r_dict.get(6)
                st = r_dict.get(7)
                et = r_dict.get(8)
                fnb = str(r_dict.get(9) or '').strip()
                log = str(r_dict.get(10) or '').strip()
                booked = str(r_dict.get(11) or '').strip()
                email = str(r_dict.get(12) or '').strip()
                bdate = r_dict.get(13)
                status = r_dict.get(14)
                fb = r_dict.get(15)
                rem = str(r_dict.get(16) or '').strip()
                taken = str(r_dict.get(17) or '').strip()
                type_val = "Request" if is_request_entry(title, loc, rem, fnb, log, parse_capacity(cap)) else "Event"
        else:
            eid = f"EVT-HIST-{curr_r-2:04d}"
            dt = r_dict.get(1)
            day = r_dict.get(2)
            loc = str(r_dict.get(3) or '').strip()
            title = str(r_dict.get(4) or '').strip()
            cap = r_dict.get(5)
            st = r_dict.get(6)
            et = r_dict.get(7)
            fnb = str(r_dict.get(8) or '').strip()
            log = str(r_dict.get(9) or '').strip()
            booked = str(r_dict.get(10) or '').strip()
            email = str(r_dict.get(11) or '').strip()
            bdate = r_dict.get(12)
            status = r_dict.get(13)
            fb = 'NA'
            rem = str(r_dict.get(14) or '').strip()
            taken = str(r_dict.get(15) or '').strip()
            type_val = "Request" if is_request_entry(title, loc, rem, fnb, log, parse_capacity(cap)) else "Event"

        pax = parse_capacity(cap)
        has_content = any(v is not None and str(v).strip() != '' for v in (dt, day, loc, title, booked, rem, status)) or (pax > 0)
        
        if not has_content and not dt:
            continue

        if is_dt4:
            scale_formula = f'=IF(F{curr_r}="","",IF(F{curr_r}="Request","Request",IF(I{curr_r}>=30,"Large Event",IF(I{curr_r}>0,"Small Event",IF(OR(ISNUMBER(SEARCH("Townhall",H{curr_r})),ISNUMBER(SEARCH("Cafeteria",D{curr_r})),ISNUMBER(SEARCH("Hiring",H{curr_r})),ISNUMBER(SEARCH("Innovation",D{curr_r})),ISNUMBER(SEARCH("All-Hands",H{curr_r})),ISNUMBER(SEARCH("All Hands",H{curr_r}))),"Large Event","Pending PAX")))))'
            row_vals = [
                eid, dt, day, loc, floor, type_val, scale_formula, title, pax, st, et,
                fnb, log, booked, email, bdate, status, fb, rem, taken
            ]
        else:
            scale_formula = f'=IF(E{curr_r}="","",IF(E{curr_r}="Request","Request",IF(H{curr_r}>=30,"Large Event",IF(H{curr_r}>0,"Small Event",IF(OR(ISNUMBER(SEARCH("Townhall",G{curr_r})),ISNUMBER(SEARCH("Cafeteria",D{curr_r})),ISNUMBER(SEARCH("Hiring",G{curr_r})),ISNUMBER(SEARCH("Innovation",D{curr_r})),ISNUMBER(SEARCH("All-Hands",G{curr_r})),ISNUMBER(SEARCH("All Hands",G{curr_r}))),"Large Event","Pending PAX")))))'
            row_vals = [
                eid, dt, day, loc, type_val, scale_formula, title, pax, st, et,
                fnb, log, booked, email, bdate, status, fb, rem, taken
            ]

        ws.row_dimensions[curr_r].height = 24
        for c_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(curr_r, c_idx, val)
            cell.font = REGULAR_FONT
            cell.border = BLACK_BORDER
            cell.alignment = ALIGN_CENTER_WRAP
            
            if is_dt4:
                cap_c, stat_c, fb_c, type_c = 9, 17, 18, 6
            else:
                cap_c, stat_c, fb_c, type_c = 8, 16, 17, 5
                
            if c_idx == cap_c:
                cell.number_format = '#,##0'
            elif c_idx == type_c:
                if val == "Request":
                    cell.fill = REQUEST_FILL
                    cell.font = REQUEST_FONT
                else:
                    cell.fill = EVENT_FILL
                    cell.font = EVENT_FONT
            elif c_idx == stat_c:
                s_str = str(val or '').lower()
                if 'confirm' in s_str or 'done' in s_str:
                    cell.value = "Confirmed (Done)"
                    cell.fill = GREEN_FILL
                    cell.font = GREEN_FONT
                elif 'cancel' in s_str:
                    cell.value = "Canceled"
                    cell.fill = RED_FILL
                    cell.font = RED_FONT
                elif has_content:
                    cell.value = "Pending Calendar Booking"
                    cell.fill = YELLOW_FILL
                    cell.font = YELLOW_FONT
            elif c_idx == fb_c:
                f_str = str(val or '').lower()
                if 'sent' in f_str:
                    cell.value = "Form sent"
                    cell.fill = FEEDBACK_SENT_FILL
                    cell.font = FEEDBACK_SENT_FONT
                elif has_content:
                    cell.value = "NA"
                    cell.fill = NO_FILL
                    cell.font = REGULAR_FONT
            elif c_idx in (2, 15 if not is_dt4 else 16):
                if isinstance(val, (datetime.date, datetime.datetime)):
                    cell.number_format = 'dd-mmm-yyyy'

        curr_r += 1

    if is_dt4:
        type_col_letter, stat_col_letter, fb_col_letter = 'F', 'Q', 'R'
    else:
        type_col_letter, stat_col_letter, fb_col_letter = 'E', 'P', 'Q'
        
    dv_type = DataValidation(type="list", formula1='"Event,Request"', allow_blank=True)
    dv_type.add(f"{type_col_letter}3:{type_col_letter}{curr_r+30}")
    ws.add_data_validation(dv_type)
    
    dv_status = DataValidation(type="list", formula1='"Confirmed (Done),Pending Calendar Booking,Canceled"', allow_blank=True)
    dv_status.add(f"{stat_col_letter}3:{stat_col_letter}{curr_r+30}")
    ws.add_data_validation(dv_status)
    
    dv_fb = DataValidation(type="list", formula1='"Form sent,NA"', allow_blank=True)
    dv_fb.add(f"{fb_col_letter}3:{fb_col_letter}{curr_r+30}")
    ws.add_data_validation(dv_fb)

    ws.freeze_panes = 'A3'
    if hasattr(ws, 'views') and ws.views and ws.views.sheetView:
        ws.views.sheetView[0].showGridLines = True

    for c_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(c_idx)
        max_l = max([len(str(ws.cell(r, c_idx).value or '')) for r in range(2, min(curr_r, 40))] + [len(headers[c_idx-1])])
        ws.column_dimensions[col_letter].width = min(max(max_l + 3, 12), 40)

print("Monthly sheets successfully prepared.")

# =============================================================================
# 2. BUILD MASTER EVENT TRACKER 2026
# =============================================================================
print("Building Master Event Tracker 2026...")
if "Master_Event_Tracker_2026" in wb.sheetnames:
    ws_master = wb["Master_Event_Tracker_2026"]
else:
    ws_master = wb.create_sheet("Master_Event_Tracker_2026")

for m in list(ws_master.merged_cells.ranges):
    ws_master.unmerge_cells(str(m))
ws_master.delete_rows(1, ws_master.max_row + 10)

ws_master.row_dimensions[1].height = 28
style_and_merge(ws_master, 1, 3, "🏠 ⮌ BACK TO DASHBOARD", HOME_BTN_FONT, NAVY_DARK_FILL, "'Dashboard'!A1")
style_and_merge(ws_master, 4, 18, "MASTER EVENT OPERATIONS DATABASE • FULL YEAR 2026 (DOWNTOWN-3 & DOWNTOWN-4)", BANNER_TITLE_FONT, NAVY_LIGHT_FILL)
style_and_merge(ws_master, 19, 24, "🟢 Green: Confirmed | 🟡 Yellow: Pending | 🔴 Red: Canceled", HEADER_FONT, NAVY_DARK_FILL)

master_headers = [
    "Event ID", "Year", "Month", "Event Date", "Day", "Building", "Floor", 
    "Location", "Type", "Event Scale", "Event Title / Purpose", "Category", "Capacity / Pax", 
    "Start Time", "End Time", "F&B Support", "Logistic Support", 
    "Booked By", "Requester Email", "Calendar Booking Status", 
    "Feedback Form", "Booking Date", "Remarks", "Booking Taken By"
]

ws_master.row_dimensions[2].height = 28
for col_idx, h_text in enumerate(master_headers, 1):
    c = ws_master.cell(2, col_idx, h_text)
    c.fill = NAVY_DARK_FILL
    c.font = HEADER_FONT
    c.alignment = ALIGN_CENTER_WRAP
    c.border = BLACK_BORDER

sheets_2026 = [s for s in wb.sheetnames if ('DT-3' in s or 'DT-4' in s) and ('26' in s or '27' in s)]

def sort_key(s):
    yr = 1 if '27' in s else 0
    b = 0 if 'DT-3' in s else 1
    m_num = 1
    for k, v in MONTH_ORDER.items():
        if k in s:
            m_num = v
            break
    return (yr, m_num, b)

sheets_2026.sort(key=sort_key)

m_curr_r = 3
total_master_records = 0

for sname in sheets_2026:
    ws = wb[sname]
    bldg = "Downtown-3" if "DT-3" in sname else "Downtown-4"
    is_dt3 = (bldg == "Downtown-3")
    yr_val = 2027 if "27" in sname else 2026
    m_name = "Jan"
    for k in MONTH_ORDER.keys():
        if k in sname:
            m_name = k
            break
            
    for r in range(3, ws.max_row + 1):
        if is_dt3:
            d_val = ws.cell(r, 2).value
            day_val = ws.cell(r, 3).value
            loc_val = str(ws.cell(r, 4).value or '').strip()
            type_val = str(ws.cell(r, 5).value or 'Event').strip()
            title_val = str(ws.cell(r, 7).value or '').strip()
            cap_val = ws.cell(r, 8).value
            start_time = ws.cell(r, 9).value
            end_time = ws.cell(r, 10).value
            fnb_val = str(ws.cell(r, 11).value or '').strip()
            logistics_val = str(ws.cell(r, 12).value or '').strip()
            booked_by_val = str(ws.cell(r, 13).value or '').strip()
            email_val = str(ws.cell(r, 14).value or '').strip()
            booking_date_val = ws.cell(r, 15).value
            status_val = ws.cell(r, 16).value
            feedback_val = ws.cell(r, 17).value
            remarks_val = str(ws.cell(r, 18).value or '').strip()
            taken_by_val = str(ws.cell(r, 19).value or '').strip()
            floor_val = 'L-1' if '1' in loc_val else ('L-5' if '5' in loc_val else ('L-4' if '4' in loc_val else ('L-6' if '6' in loc_val else 'General')))
        else:
            d_val = ws.cell(r, 2).value
            day_val = ws.cell(r, 3).value
            loc_val = str(ws.cell(r, 4).value or '').strip()
            floor_val = str(ws.cell(r, 5).value or '').strip()
            type_val = str(ws.cell(r, 6).value or 'Event').strip()
            title_val = str(ws.cell(r, 8).value or '').strip()
            cap_val = ws.cell(r, 9).value
            start_time = ws.cell(r, 10).value
            end_time = ws.cell(r, 11).value
            fnb_val = str(ws.cell(r, 12).value or '').strip()
            logistics_val = str(ws.cell(r, 13).value or '').strip()
            booked_by_val = str(ws.cell(r, 14).value or '').strip()
            email_val = str(ws.cell(r, 15).value or '').strip()
            booking_date_val = ws.cell(r, 16).value
            status_val = ws.cell(r, 17).value
            feedback_val = ws.cell(r, 18).value
            remarks_val = str(ws.cell(r, 19).value or '').strip()
            taken_by_val = str(ws.cell(r, 20).value or '').strip()

        pax = parse_capacity(cap_val)
        has_event = any(v is not None and str(v).strip() != '' for v in (d_val, day_val, loc_val, title_val, booked_by_val, remarks_val, status_val)) or (pax > 0)
        
        if not has_event:
            continue
            
        total_master_records += 1
        
        evt_formula = f'=IF(OR(D{m_curr_r}<>"",H{m_curr_r}<>"",K{m_curr_r}<>"",R{m_curr_r}<>"",T{m_curr_r}<>""), "EVT-" & TEXT(B{m_curr_r}, "0000") & "-" & TEXT(COUNTIF(OFFSET($A$1, 1, 0, ROW()-2, 1), "EVT*") + 1, "0000"), "")'
        scale_formula = f'=IF(I{m_curr_r}="","",IF(I{m_curr_r}="Request","Request",IF(M{m_curr_r}>=30,"Large Event",IF(M{m_curr_r}>0,"Small Event",IF(OR(ISNUMBER(SEARCH("Townhall",K{m_curr_r})),ISNUMBER(SEARCH("Cafeteria",H{m_curr_r})),ISNUMBER(SEARCH("Hiring",K{m_curr_r})),ISNUMBER(SEARCH("Innovation",H{m_curr_r})),ISNUMBER(SEARCH("All-Hands",K{m_curr_r})),ISNUMBER(SEARCH("All Hands",K{m_curr_r}))),"Large Event","Pending PAX")))))'
        
        date_obj = parse_date(d_val, yr_val)
        formatted_date = date_obj if date_obj else format_date_val(d_val, yr_val)
        formatted_day = parse_day_str(day_val, date_obj)
        formatted_start = parse_time_str(start_time)
        formatted_end = parse_time_str(end_time)
        booking_date_obj = parse_date(booking_date_val, yr_val)
        formatted_booking_date = booking_date_obj if booking_date_obj else format_date_val(booking_date_val, yr_val)
        
        final_status = standardize_status(status_val)
        final_feedback = standardize_feedback(feedback_val, final_status)
        cat = categorize_event(title_val, loc_val, remarks_val, type_val)
        
        row_vals = [
            evt_formula, yr_val, m_name, formatted_date, formatted_day, bldg, floor_val,
            loc_val, type_val, scale_formula, title_val, cat, pax, formatted_start, formatted_end,
            fnb_val, logistics_val, booked_by_val, email_val, final_status,
            final_feedback, formatted_booking_date, remarks_val, taken_by_val
        ]
        
        ws_master.row_dimensions[m_curr_r].height = 24
        for mc_idx, m_val in enumerate(row_vals, 1):
            mc = ws_master.cell(m_curr_r, mc_idx, m_val)
            mc.font = REGULAR_FONT
            mc.border = BLACK_BORDER
            mc.alignment = ALIGN_CENTER_WRAP
            
            if mc_idx == 9:
                if type_val == "Request":
                    mc.fill = REQUEST_FILL
                    mc.font = REQUEST_FONT
                else:
                    mc.fill = EVENT_FILL
                    mc.font = EVENT_FONT
            elif mc_idx == 13:
                mc.number_format = '#,##0'
            elif mc_idx in (14, 15):
                mc.number_format = 'hh:mm AM/PM'
            elif mc_idx in (4, 22):
                if isinstance(m_val, (datetime.date, datetime.datetime)):
                    mc.number_format = 'dd-mmm-yyyy'
            elif mc_idx == 20:
                if "Confirmed" in final_status:
                    mc.fill = GREEN_FILL
                    mc.font = GREEN_FONT
                elif "Canceled" in final_status:
                    mc.fill = RED_FILL
                    mc.font = RED_FONT
                else:
                    mc.fill = YELLOW_FILL
                    mc.font = YELLOW_FONT
            elif mc_idx == 21:
                if final_feedback == "Form sent":
                    mc.fill = FEEDBACK_SENT_FILL
                    mc.font = FEEDBACK_SENT_FONT
                else:
                    mc.fill = NO_FILL
                    mc.font = REGULAR_FONT

        m_curr_r += 1

dv_m_type = DataValidation(type="list", formula1='"Event,Request"', allow_blank=True)
dv_m_type.add(f"I3:I{m_curr_r+50}")
ws_master.add_data_validation(dv_m_type)

dv_m_status = DataValidation(type="list", formula1='"Confirmed (Done),Pending Calendar Booking,Canceled"', allow_blank=True)
dv_m_status.add(f"T3:T{m_curr_r+50}")
ws_master.add_data_validation(dv_m_status)

dv_m_fb = DataValidation(type="list", formula1='"Form sent,NA"', allow_blank=True)
dv_m_fb.add(f"U3:U{m_curr_r+50}")
ws_master.add_data_validation(dv_m_fb)

ws_master.freeze_panes = 'A3'
if hasattr(ws_master, 'views') and ws_master.views and ws_master.views.sheetView:
    ws_master.views.sheetView[0].showGridLines = True

master_widths = [16, 8, 8, 14, 8, 14, 10, 20, 12, 14, 30, 24, 14, 12, 12, 16, 20, 20, 25, 24, 14, 14, 30, 18]
for c_idx, w in enumerate(master_widths, 1):
    ws_master.column_dimensions[get_column_letter(c_idx)].width = w

print("Master Tracker complete.")

# =============================================================================
# 3. BUILD SHORT, SIMPLIFIED, EXECUTIVE DASHBOARD (NUMBERS ONLY & NAV PAGE)
# =============================================================================
print("Constructing Simplified Executive Dashboard with Venue & Category Tables...")
if "Dashboard" in wb.sheetnames:
    del wb["Dashboard"]

ws_dash = wb.create_sheet("Dashboard", 0)

def style_dash_cell(ws, r, c, val=None, font=None, fill=None, border=None, align=None, num_format=None, hyperlink_loc=None):
    cell = ws.cell(r, c)
    if val is not None: cell.value = val
    if font: cell.font = font
    if fill: cell.fill = fill
    if border: cell.border = border
    if align: cell.alignment = align
    if num_format: cell.number_format = num_format
    if hyperlink_loc:
        col_letter = get_column_letter(c)
        cell.hyperlink = Hyperlink(ref=f"{col_letter}{r}", location=hyperlink_loc)
    return cell

def style_dash_range(ws, min_r, min_c, max_r, max_c, font=None, fill=None, border=None, align=None, num_format=None):
    for r in range(min_r, max_r + 1):
        for c in range(min_c, max_c + 1):
            cell = ws.cell(r, c)
            if font: cell.font = font
            if fill: cell.fill = fill
            if border: cell.border = border
            if align: cell.alignment = align
            if num_format: cell.number_format = num_format

# Row 1 & 2: Header Banners
ws_dash.row_dimensions[1].height = 32
ws_dash.row_dimensions[2].height = 20
ws_dash.merge_cells('A1:W1')
ws_dash.merge_cells('A2:W2')

style_dash_cell(ws_dash, 1, 1, "🏢 CBRE | EVENT & REQUEST OPERATIONS EXECUTIVE DASHBOARD", font=FONT_MAIN_TITLE, fill=FILL_BANNER_DARK, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 1, 1, 1, 23, fill=FILL_BANNER_DARK)

style_dash_cell(ws_dash, 2, 1, "Downtown-3 & Downtown-4 Facilities • 2026 Simplified Operations, Capacity Analytics & Auto-Navigation", font=FONT_SUB_TITLE, fill=NAVY_DARK_FILL, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 2, 1, 2, 23, fill=NAVY_DARK_FILL)

ws_dash.row_dimensions[3].height = 6

# Row 4 & 5: Legend & Jump Buttons
ws_dash.row_dimensions[4].height = 18
ws_dash.row_dimensions[5].height = 26
ws_dash.merge_cells('A4:W4')
style_dash_cell(ws_dash, 4, 1, "STATUS & CLASSIFICATION LEGEND GUIDE (CLICK MASTER TO JUMP):", font=FONT_BOLD_NAVY, align=ALIGN_LEFT_WRAP)

ws_dash.merge_cells('A5:C5')
style_dash_cell(ws_dash, 5, 1, "📑 OPEN MASTER DATABASE ↗", font=FONT_TABLE_HEADER, fill=FILL_TEAL_BTN, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, hyperlink_loc="'Master_Event_Tracker_2026'!A1")
style_dash_range(ws_dash, 5, 1, 5, 3, fill=FILL_TEAL_BTN, border=BLACK_BORDER)

ws_dash.merge_cells('D5:F5')
style_dash_cell(ws_dash, 5, 4, "🟢 GREEN: Confirmed Booking", font=FONT_TABLE_HEADER, fill=GREEN_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 5, 4, 5, 6, fill=GREEN_FILL, border=BLACK_BORDER)

ws_dash.merge_cells('G5:I5')
style_dash_cell(ws_dash, 5, 7, "🟡 YELLOW: Booking Pending", font=YELLOW_FONT, fill=YELLOW_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 5, 7, 5, 9, fill=YELLOW_FILL, border=BLACK_BORDER)

ws_dash.merge_cells('J5:L5')
style_dash_cell(ws_dash, 5, 10, "🔴 RED: Canceled Booking", font=RED_FONT, fill=RED_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 5, 10, 5, 12, fill=RED_FILL, border=BLACK_BORDER)

ws_dash.merge_cells('M5:O5')
style_dash_cell(ws_dash, 5, 13, "🎪 EVENTS (Large ≥30 / Small <30)", font=FONT_TABLE_HEADER, fill=FILL_PURPLE_CARD, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 5, 13, 5, 15, fill=FILL_PURPLE_CARD, border=BLACK_BORDER)

ws_dash.merge_cells('P5:R5')
style_dash_cell(ws_dash, 5, 16, "📋 REQUESTS (Water / Tissues / Supplies)", font=Font(name=FONT_FAMILY, size=10, bold=True, color='856404'), fill=FILL_SOFT_ORANGE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 5, 16, 5, 18, fill=FILL_SOFT_ORANGE, border=BLACK_BORDER)

ws_dash.merge_cells('S5:U5')
style_dash_cell(ws_dash, 5, 19, "⏳ PENDING PAX (PAX Unassigned / TBD)", font=Font(name=FONT_FAMILY, size=10, bold=True, color='7D6608'), fill=FILL_SOFT_YELLOW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 5, 19, 5, 21, fill=FILL_SOFT_YELLOW, border=BLACK_BORDER)

ws_dash.merge_cells('V5:W5')
style_dash_cell(ws_dash, 5, 22, "📨 FEEDBACK (Sent vs NA)", font=FEEDBACK_SENT_FONT, fill=FEEDBACK_SENT_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 5, 22, 5, 23, fill=FEEDBACK_SENT_FILL, border=BLACK_BORDER)

ws_dash.row_dimensions[6].height = 8

# Rows 7-9: Executive KPI Cards (Numbers Only)
ws_dash.row_dimensions[7].height = 18
ws_dash.row_dimensions[8].height = 32
ws_dash.row_dimensions[9].height = 16

cards = [
    ('A', 'B', 'TOTAL RECORDS', '=B25+N28', 'DT-3 & DT-4 Combined', FILL_CARD_GRAY, FONT_CARD_TITLE, FONT_KPI_NAVY, FONT_CARD_SUB),
    ('C', 'D', '🎪 TOTAL EVENTS', '=C25+O28', 'Meetings & Sessions', FILL_CARD_GRAY, FONT_CARD_TITLE, FONT_KPI_BLUE, FONT_CARD_SUB),
    ('E', 'F', '📋 TOTAL REQUESTS', '=D25+P28', 'Water / Tissues / Supplies', FILL_SOFT_ORANGE, Font(name=FONT_FAMILY, size=9, bold=True, color='856404'), Font(name=FONT_FAMILY, size=18, bold=True, color='856404'), Font(name=FONT_FAMILY, size=8, bold=True, color='856404')),
    ('G', 'H', '🏢 LARGE EVENTS (≥30)', '=E25+Q28', 'High Capacity (≥30 Pax)', FILL_SOFT_GREEN, Font(name=FONT_FAMILY, size=9, bold=True, color='155724'), Font(name=FONT_FAMILY, size=18, bold=True, color='155724'), Font(name=FONT_FAMILY, size=8, bold=True, color='155724')),
    ('I', 'J', '👥 SMALL EVENTS (<30)', '=F25+R28', 'Standard (<30 Pax)', FILL_CARD_GRAY, FONT_CARD_TITLE, FONT_KPI_NAVY, FONT_CARD_SUB),
    ('K', 'L', '⏳ PENDING PAX', '=G25+S28', 'Unassigned / TBD PAX', FILL_SOFT_YELLOW, Font(name=FONT_FAMILY, size=9, bold=True, color='7D6608'), Font(name=FONT_FAMILY, size=18, bold=True, color='7D6608'), Font(name=FONT_FAMILY, size=8, bold=True, color='7D6608')),
    ('M', 'N', '🟢 CONFIRMED', '=H25+T28', 'Calendar Booking Done', GREEN_FILL, FONT_TABLE_HEADER, FONT_KPI_WHITE, FONT_TABLE_HEADER),
    ('O', 'P', '🟡 PENDING', '=I25+U28', 'Action Required', YELLOW_FILL, YELLOW_FONT, FONT_KPI_BLACK, YELLOW_FONT),
    ('Q', 'R', '🔴 CANCELED', '=J25+V28', 'Canceled by Requester', RED_FILL, FONT_TABLE_HEADER, FONT_KPI_WHITE, FONT_TABLE_HEADER),
    ('S', 'T', '👥 TOTAL PAX', '=K25+W28', 'Total Expected Attendees', FILL_PURPLE_CARD, FONT_TABLE_HEADER, FONT_KPI_WHITE, FONT_TABLE_HEADER),
    ('U', 'V', '📨 FEEDBACK SENT', '=COUNTIF(Master_Event_Tracker_2026!$U$3:$U$5000, "Form sent")', 'Forms Sent (2026)', FEEDBACK_SENT_FILL, Font(name=FONT_FAMILY, size=9, bold=True, color='1F4E79'), FONT_KPI_BLUE, Font(name=FONT_FAMILY, size=8, bold=True, color='1F4E79')),
    ('W', 'W', '⚪ FB NA', '=COUNTIF(Master_Event_Tracker_2026!$U$3:$U$5000, "NA")', 'NA Form', FILL_CARD_GRAY, FONT_CARD_TITLE, FONT_KPI_NAVY, FONT_CARD_SUB)
]

for item in cards:
    c1, c2, title, formula, sub, fill, f_title, f_kpi, f_sub = item
    c1_idx = column_index_from_string(c1)
    c2_idx = column_index_from_string(c2)
    if c2_idx > c1_idx:
        ws_dash.merge_cells(f'{c1}7:{c2}7')
        ws_dash.merge_cells(f'{c1}8:{c2}8')
        ws_dash.merge_cells(f'{c1}9:{c2}9')
    style_dash_range(ws_dash, 7, c1_idx, 9, c2_idx, fill=fill, border=BLACK_BORDER)
    style_dash_cell(ws_dash, 7, c1_idx, title, font=f_title, align=ALIGN_CENTER_WRAP)
    style_dash_cell(ws_dash, 8, c1_idx, formula, font=f_kpi, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, 9, c1_idx, sub, font=f_sub, align=ALIGN_CENTER_WRAP)

ws_dash.row_dimensions[10].height = 10

# Rows 11-28: Two Side-by-Side Facility Monthly Breakdown Tables (DT-3 on Left, DT-4 on Right)
ws_dash.row_dimensions[11].height = 26
ws_dash.row_dimensions[12].height = 24

ws_dash.merge_cells('A11:K11')
style_dash_cell(ws_dash, 11, 1, "🏢 DOWNTOWN-3 (DT-3) 2026 OPERATIONS • MONTHLY BREAKDOWN & AUTO-NAVIGATE", font=FONT_SECTION_HEADER, fill=FILL_ROYAL_BLUE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 11, 1, 11, 11, fill=FILL_ROYAL_BLUE, border=BLACK_BORDER)

ws_dash.merge_cells('M11:W11')
style_dash_cell(ws_dash, 11, 13, "🏢 DOWNTOWN-4 (DT-4) OPERATIONS • MONTHLY BREAKDOWN & AUTO-NAVIGATE", font=FONT_SECTION_HEADER, fill=NAVY_DARK_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 11, 13, 11, 23, fill=NAVY_DARK_FILL, border=BLACK_BORDER)

dt_headers = ['Month ↗', 'Total', '🎪 Events', '📋 Requests', 'Large (≥30)', 'Small (<30)', '⏳ Pend.PAX', '🟢 Conf.', '🟡 Pend.', '🔴 Canc.', 'Total Pax']

for idx, h in enumerate(dt_headers, 1):
    style_dash_cell(ws_dash, 12, idx, h, font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)

for idx, h in enumerate(dt_headers, 13):
    style_dash_cell(ws_dash, 12, idx, h, font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)

months_26 = ['Jan26', 'Feb26', 'Mar26', 'Apr26', 'May26', 'Jun26', 'Jul26', 'Aug26', 'Sep26', 'Oct26', 'Nov26', 'Dec26']
month_labels = ['Jan 2026 ↗', 'Feb 2026 ↗', 'Mar 2026 ↗', 'Apr 2026 ↗', 'May 2026 ↗', 'Jun 2026 ↗', 'Jul 2026 ↗', 'Aug 2026 ↗', 'Sep 2026 ↗', 'Oct 2026 ↗', 'Nov 2026 ↗', 'Dec 2026 ↗']

for idx, (m_tag, m_lbl) in enumerate(zip(months_26, month_labels)):
    r = 13 + idx
    ws_dash.row_dimensions[r].height = 20
    s_dt3 = f"DT-3 {m_tag}"
    s_dt4 = f"DT-4 {m_tag}"
    
    # DT-3 Columns (1 to 11):
    style_dash_cell(ws_dash, r, 1, m_lbl, font=FONT_LINK, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, hyperlink_loc=f"'{s_dt3}'!A1")
    style_dash_cell(ws_dash, r, 2, f'=COUNTIF(\'{s_dt3}\'!$P$3:$P$500, "<>")', font=BOLD_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 3, f'=COUNTIF(\'{s_dt3}\'!$E$3:$E$500, "Event")', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 4, f'=COUNTIF(\'{s_dt3}\'!$E$3:$E$500, "Request")', font=Font(name=FONT_FAMILY, size=9, bold=True, color='856404'), fill=FILL_SOFT_ORANGE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 5, f'=COUNTIF(\'{s_dt3}\'!$F$3:$F$500, "Large Event")', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 6, f'=COUNTIF(\'{s_dt3}\'!$F$3:$F$500, "Small Event")', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 7, f'=COUNTIF(\'{s_dt3}\'!$F$3:$F$500, "Pending PAX")', font=Font(name=FONT_FAMILY, size=9, bold=True, color='7D6608'), fill=FILL_SOFT_YELLOW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 8, f'=COUNTIF(\'{s_dt3}\'!$P$3:$P$500, "Confirmed (Done)")', font=GREEN_FONT, fill=GREEN_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 9, f'=COUNTIF(\'{s_dt3}\'!$P$3:$P$500, "Pending Calendar Booking")', font=YELLOW_FONT, fill=YELLOW_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 10, f'=COUNTIF(\'{s_dt3}\'!$P$3:$P$500, "Canceled")', font=RED_FONT, fill=RED_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 11, f'=SUM(\'{s_dt3}\'!$H$3:$H$500)', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

    # DT-4 Columns (13 to 23):
    style_dash_cell(ws_dash, r, 13, m_lbl, font=FONT_LINK, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, hyperlink_loc=f"'{s_dt4}'!A1")
    style_dash_cell(ws_dash, r, 14, f'=COUNTIF(\'{s_dt4}\'!$Q$3:$Q$500, "<>")', font=BOLD_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 15, f'=COUNTIF(\'{s_dt4}\'!$F$3:$F$500, "Event")', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 16, f'=COUNTIF(\'{s_dt4}\'!$F$3:$F$500, "Request")', font=Font(name=FONT_FAMILY, size=9, bold=True, color='856404'), fill=FILL_SOFT_ORANGE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 17, f'=COUNTIF(\'{s_dt4}\'!$G$3:$G$500, "Large Event")', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 18, f'=COUNTIF(\'{s_dt4}\'!$G$3:$G$500, "Small Event")', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 19, f'=COUNTIF(\'{s_dt4}\'!$G$3:$G$500, "Pending PAX")', font=Font(name=FONT_FAMILY, size=9, bold=True, color='7D6608'), fill=FILL_SOFT_YELLOW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 20, f'=COUNTIF(\'{s_dt4}\'!$Q$3:$Q$500, "Confirmed (Done)")', font=GREEN_FONT, fill=GREEN_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 21, f'=COUNTIF(\'{s_dt4}\'!$Q$3:$Q$500, "Pending Calendar Booking")', font=YELLOW_FONT, fill=YELLOW_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 22, f'=COUNTIF(\'{s_dt4}\'!$Q$3:$Q$500, "Canceled")', font=RED_FONT, fill=RED_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 23, f'=SUM(\'{s_dt4}\'!$I$3:$I$500)', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

# DT-3 Full Year Total (Row 25 in Cols A to K)
ws_dash.row_dimensions[25].height = 22
style_dash_cell(ws_dash, 25, 1, "DT-3 Total (2026)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
for c in range(2, 12):
    c_let = get_column_letter(c)
    style_dash_cell(ws_dash, 25, c, f"=SUM({c_let}13:{c_let}24)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

# DT-4 2027 Projections (Rows 25 to 27 in Cols 13 to 23)
months_27 = [('Jan27', "Jan '27 ↗"), ('Feb27', "Feb '27 ↗"), ('Mar27', "Mar '27 ↗")]
for idx, (m_tag, m_lbl) in enumerate(months_27):
    r = 25 + idx
    ws_dash.row_dimensions[r].height = 20
    s_dt4_27 = f"DT-4 {m_tag}"
    style_dash_cell(ws_dash, r, 13, m_lbl, font=FONT_LINK, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, hyperlink_loc=f"'{s_dt4_27}'!A1")
    style_dash_cell(ws_dash, r, 14, f'=COUNTIF(\'{s_dt4_27}\'!$Q$3:$Q$500, "<>")', font=BOLD_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 15, f'=COUNTIF(\'{s_dt4_27}\'!$F$3:$F$500, "Event")', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 16, f'=COUNTIF(\'{s_dt4_27}\'!$F$3:$F$500, "Request")', font=Font(name=FONT_FAMILY, size=9, bold=True, color='856404'), fill=FILL_SOFT_ORANGE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 17, f'=COUNTIF(\'{s_dt4_27}\'!$G$3:$G$500, "Large Event")', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 18, f'=COUNTIF(\'{s_dt4_27}\'!$G$3:$G$500, "Small Event")', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 19, f'=COUNTIF(\'{s_dt4_27}\'!$G$3:$G$500, "Pending PAX")', font=Font(name=FONT_FAMILY, size=9, bold=True, color='7D6608'), fill=FILL_SOFT_YELLOW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 20, f'=COUNTIF(\'{s_dt4_27}\'!$Q$3:$Q$500, "Confirmed (Done)")', font=GREEN_FONT, fill=GREEN_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 21, f'=COUNTIF(\'{s_dt4_27}\'!$Q$3:$Q$500, "Pending Calendar Booking")', font=YELLOW_FONT, fill=YELLOW_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 22, f'=COUNTIF(\'{s_dt4_27}\'!$Q$3:$Q$500, "Canceled")', font=RED_FONT, fill=RED_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 23, f'=SUM(\'{s_dt4_27}\'!$I$3:$I$500)', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

# DT-4 2026 Subtotal (Row 28 in Cols 13 to 23)
ws_dash.row_dimensions[28].height = 22
style_dash_cell(ws_dash, 28, 13, "DT-4 Subtotal (2026)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
for c in range(14, 24):
    c_let = get_column_letter(c)
    style_dash_cell(ws_dash, 28, c, f"=SUM({c_let}13:{c_let}24)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

# Rows 26-28 Left Side: Facility Summary Box (DT-3 vs DT-4 vs Total)
ws_dash.merge_cells('A26:D26')
style_dash_cell(ws_dash, 26, 1, "🏢 FACILITY 2026 BREAKDOWN", font=FONT_TABLE_HEADER, fill=FILL_ROYAL_BLUE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 26, 1, 26, 4, fill=FILL_ROYAL_BLUE, border=BLACK_BORDER)

ws_dash.merge_cells('E26:G26')
style_dash_cell(ws_dash, 26, 5, "🎪 EVENTS", font=FONT_TABLE_HEADER, fill=FILL_ROYAL_BLUE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 26, 5, 26, 7, fill=FILL_ROYAL_BLUE, border=BLACK_BORDER)

ws_dash.merge_cells('H26:K26')
style_dash_cell(ws_dash, 26, 8, "📋 REQUESTS", font=FONT_TABLE_HEADER, fill=FILL_ROYAL_BLUE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 26, 8, 26, 11, fill=FILL_ROYAL_BLUE, border=BLACK_BORDER)

# Row 27: DT-3 vs DT-4 breakdown
ws_dash.merge_cells('A27:D27')
style_dash_cell(ws_dash, 27, 1, "Downtown-3 (DT-3)", font=FONT_BOLD_NAVY, fill=FILL_CARD_GRAY, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
style_dash_range(ws_dash, 27, 1, 27, 4, fill=FILL_CARD_GRAY, border=BLACK_BORDER)

ws_dash.merge_cells('E27:G27')
style_dash_cell(ws_dash, 27, 5, "=C25", font=FONT_KPI_BLUE, fill=FILL_CARD_GRAY, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_range(ws_dash, 27, 5, 27, 7, fill=FILL_CARD_GRAY, border=BLACK_BORDER)

ws_dash.merge_cells('H27:K27')
style_dash_cell(ws_dash, 27, 8, "=D25", font=Font(name=FONT_FAMILY, size=11, bold=True, color='856404'), fill=FILL_SOFT_ORANGE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_range(ws_dash, 27, 8, 27, 11, fill=FILL_SOFT_ORANGE, border=BLACK_BORDER)

# Row 28: DT-4 breakdown
ws_dash.merge_cells('A28:D28')
style_dash_cell(ws_dash, 28, 1, "Downtown-4 (DT-4)", font=FONT_BOLD_NAVY, fill=FILL_CARD_GRAY, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
style_dash_range(ws_dash, 28, 1, 28, 4, fill=FILL_CARD_GRAY, border=BLACK_BORDER)

ws_dash.merge_cells('E28:G28')
style_dash_cell(ws_dash, 28, 5, "=O28", font=FONT_KPI_BLUE, fill=FILL_CARD_GRAY, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_range(ws_dash, 28, 5, 28, 7, fill=FILL_CARD_GRAY, border=BLACK_BORDER)

ws_dash.merge_cells('H28:K28')
style_dash_cell(ws_dash, 28, 8, "=P28", font=Font(name=FONT_FAMILY, size=11, bold=True, color='856404'), fill=FILL_SOFT_ORANGE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_range(ws_dash, 28, 8, 28, 11, fill=FILL_SOFT_ORANGE, border=BLACK_BORDER)

ws_dash.row_dimensions[29].height = 10

# Rows 30-44: Consolidated 2026 Combined Monthly Summary (Cols A to L) + Navigation Directory Page (Cols N to W)
ws_dash.row_dimensions[30].height = 26
ws_dash.row_dimensions[31].height = 24

ws_dash.merge_cells('A30:L30')
style_dash_cell(ws_dash, 30, 1, "📊 2026 CONSOLIDATED OPERATIONS SUMMARY (DT-3 + DT-4 COMBINED)", font=FONT_SECTION_HEADER, fill=NAVY_DARK_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 30, 1, 30, 12, fill=NAVY_DARK_FILL, border=BLACK_BORDER)

comb_headers = ['Month', 'Total Records', '🎪 Events', '📋 Requests', 'Large (≥30)', 'Small (<30)', '⏳ Pend.PAX', '🟢 Confirmed', '🟡 Pending', '🔴 Canceled', 'Confirm %', 'Total Pax']
for idx, h in enumerate(comb_headers, 1):
    style_dash_cell(ws_dash, 31, idx, h, font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)

months_short = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
for idx, m_short in enumerate(months_short):
    r = 32 + idx
    ref_r = 13 + idx
    ws_dash.row_dimensions[r].height = 20
    style_dash_cell(ws_dash, r, 1, m_short, font=BOLD_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
    style_dash_cell(ws_dash, r, 2, f'=$C{r}+$D{r}', font=BOLD_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 3, f'=C{ref_r}+O{ref_r}', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 4, f'=D{ref_r}+P{ref_r}', font=Font(name=FONT_FAMILY, size=9, bold=True, color='856404'), fill=FILL_SOFT_ORANGE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 5, f'=E{ref_r}+Q{ref_r}', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 6, f'=F{ref_r}+R{ref_r}', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 7, f'=G{ref_r}+S{ref_r}', font=Font(name=FONT_FAMILY, size=9, bold=True, color='7D6608'), fill=FILL_SOFT_YELLOW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 8, f'=H{ref_r}+T{ref_r}', font=GREEN_FONT, fill=GREEN_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 9, f'=I{ref_r}+U{ref_r}', font=YELLOW_FONT, fill=YELLOW_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 10, f'=J{ref_r}+V{ref_r}', font=RED_FONT, fill=RED_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    style_dash_cell(ws_dash, r, 11, f'=IFERROR($H{r}/($B{r}-$J{r}), 0)', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='0.0%')
    style_dash_cell(ws_dash, r, 12, f'=K{ref_r}+W{ref_r}', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

# Full Year 2026 Combined Total (Row 44)
ws_dash.row_dimensions[44].height = 22
style_dash_cell(ws_dash, 44, 1, "Full Year 2026 Total", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
for c in range(2, 11):
    c_let = get_column_letter(c)
    style_dash_cell(ws_dash, 44, c, f"=SUM({c_let}32:{c_let}43)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
style_dash_cell(ws_dash, 44, 11, "=IFERROR(H44/(B44-J44), 0)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='0.0%')
style_dash_cell(ws_dash, 44, 12, "=SUM(L32:L43)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

# Right Side: MASTER NAVIGATION DIRECTORY (Cols N to W, Rows 30-44)
ws_dash.merge_cells('N30:W30')
style_dash_cell(ws_dash, 30, 14, "🚀 MASTER NAVIGATION PAGE & DIRECTORY (CLICK ANY TILE TO OPEN)", font=FONT_SECTION_HEADER, fill=FILL_ROYAL_BLUE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 30, 14, 30, 23, fill=FILL_ROYAL_BLUE, border=BLACK_BORDER)

# Row 31: Master Database Big Button
ws_dash.merge_cells('N31:W31')
ws_dash.row_dimensions[31].height = 24
style_dash_cell(ws_dash, 31, 14, "📑 CLICK HERE TO OPEN MASTER EVENT & REQUEST DATABASE 2026 ↗", font=FONT_NAV_BTN, fill=FILL_TEAL_BTN, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, hyperlink_loc="'Master_Event_Tracker_2026'!A1")
style_dash_range(ws_dash, 31, 14, 31, 23, fill=FILL_TEAL_BTN, border=BLACK_BORDER)

# Row 32: Header DT-3 2026 Navigation
ws_dash.merge_cells('N32:W32')
ws_dash.row_dimensions[32].height = 20
style_dash_cell(ws_dash, 32, 14, "🏢 DOWNTOWN-3 (DT-3) MONTHLY SHEETS (2026):", font=FONT_TABLE_HEADER, fill=NAVY_DARK_FILL, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
style_dash_range(ws_dash, 32, 14, 32, 23, fill=NAVY_DARK_FILL, border=BLACK_BORDER)

dt3_nav_1 = [('Jan26', 'N33', 'O33'), ('Feb26', 'P33', 'Q33'), ('Mar26', 'R33', 'S33'), ('Apr26', 'T33', 'T33'), ('May26', 'U33', 'U33'), ('Jun26', 'V33', 'W33')]
dt3_nav_2 = [('Jul26', 'N34', 'O34'), ('Aug26', 'P34', 'Q34'), ('Sep26', 'R34', 'S34'), ('Oct26', 'T34', 'T34'), ('Nov26', 'U34', 'U34'), ('Dec26', 'V34', 'W34')]

for m_t, c_start, c_end in dt3_nav_1:
    ws_dash.merge_cells(f"{c_start}:{c_end}")
    c_idx = column_index_from_string(c_start[:1])
    r_idx = int(c_start[1:])
    style_dash_cell(ws_dash, r_idx, c_idx, f"DT-3 {m_t[:3]} ↗", font=FONT_LINK, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, hyperlink_loc=f"'DT-3 {m_t}'!A1")
    style_dash_range(ws_dash, r_idx, c_idx, r_idx, column_index_from_string(c_end[:1]), fill=FILL_TOTAL_ROW, border=BLACK_BORDER)

for m_t, c_start, c_end in dt3_nav_2:
    ws_dash.merge_cells(f"{c_start}:{c_end}")
    c_idx = column_index_from_string(c_start[:1])
    r_idx = int(c_start[1:])
    style_dash_cell(ws_dash, r_idx, c_idx, f"DT-3 {m_t[:3]} ↗", font=FONT_LINK, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, hyperlink_loc=f"'DT-3 {m_t}'!A1")
    style_dash_range(ws_dash, r_idx, c_idx, r_idx, column_index_from_string(c_end[:1]), fill=FILL_TOTAL_ROW, border=BLACK_BORDER)

# Row 35: Header DT-4 2026 Navigation
ws_dash.merge_cells('N35:W35')
ws_dash.row_dimensions[35].height = 20
style_dash_cell(ws_dash, 35, 14, "🏢 DOWNTOWN-4 (DT-4) MONTHLY SHEETS (2026 & 2027):", font=FONT_TABLE_HEADER, fill=NAVY_DARK_FILL, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
style_dash_range(ws_dash, 35, 14, 35, 23, fill=NAVY_DARK_FILL, border=BLACK_BORDER)

dt4_nav_1 = [('Jan26', 'N36', 'O36'), ('Feb26', 'P36', 'Q36'), ('Mar26', 'R36', 'S36'), ('Apr26', 'T36', 'T36'), ('May26', 'U36', 'U36'), ('Jun26', 'V36', 'W36')]
dt4_nav_2 = [('Jul26', 'N37', 'O37'), ('Aug26', 'P37', 'Q37'), ('Sep26', 'R37', 'S37'), ('Oct26', 'T37', 'T37'), ('Nov26', 'U37', 'U37'), ('Dec26', 'V37', 'W37')]
dt4_nav_3 = [('Jan27', 'N38', 'P38'), ('Feb27', 'Q38', 'S38'), ('Mar27', 'T38', 'W38')]

for m_t, c_start, c_end in dt4_nav_1:
    ws_dash.merge_cells(f"{c_start}:{c_end}")
    c_idx = column_index_from_string(c_start[:1])
    r_idx = int(c_start[1:])
    style_dash_cell(ws_dash, r_idx, c_idx, f"DT-4 {m_t[:3]} ↗", font=FONT_LINK, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, hyperlink_loc=f"'DT-4 {m_t}'!A1")
    style_dash_range(ws_dash, r_idx, c_idx, r_idx, column_index_from_string(c_end[:1]), fill=FILL_TOTAL_ROW, border=BLACK_BORDER)

for m_t, c_start, c_end in dt4_nav_2:
    ws_dash.merge_cells(f"{c_start}:{c_end}")
    c_idx = column_index_from_string(c_start[:1])
    r_idx = int(c_start[1:])
    style_dash_cell(ws_dash, r_idx, c_idx, f"DT-4 {m_t[:3]} ↗", font=FONT_LINK, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, hyperlink_loc=f"'DT-4 {m_t}'!A1")
    style_dash_range(ws_dash, r_idx, c_idx, r_idx, column_index_from_string(c_end[:1]), fill=FILL_TOTAL_ROW, border=BLACK_BORDER)

ws_dash.row_dimensions[38].height = 20
for m_t, c_start, c_end in dt4_nav_3:
    ws_dash.merge_cells(f"{c_start}:{c_end}")
    c_idx = column_index_from_string(c_start[:1])
    r_idx = int(c_start[1:])
    style_dash_cell(ws_dash, r_idx, c_idx, f"DT-4 {m_t} (Projection) ↗", font=FONT_LINK, fill=FILL_SOFT_GREEN, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, hyperlink_loc=f"'DT-4 {m_t}'!A1")
    style_dash_range(ws_dash, r_idx, c_idx, r_idx, column_index_from_string(c_end[:1]), fill=FILL_SOFT_GREEN, border=BLACK_BORDER)

# Row 39: Header Historical Year Navigation
ws_dash.merge_cells('N39:W39')
ws_dash.row_dimensions[39].height = 20
style_dash_cell(ws_dash, 39, 14, "📁 HISTORICAL ARCHIVE SHEETS (2025 & 2024 DIRECT ACCESS):", font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
style_dash_range(ws_dash, 39, 14, 39, 23, fill=FILL_NAVY_SLATE, border=BLACK_BORDER)

# Row 40: 2025 Key Archive Tiles
hist_25_tiles = [('Jan25', 'Jan25 ↗', 'N40', 'O40'), ('Feb25', 'Feb25 ↗', 'P40', 'Q40'), ('Jun25', 'Jun25 ↗', 'R40', 'S40'), ('Sep\'25- Downtown-3', 'Sep25(DT3) ↗', 'T40', 'T40'), ('Downtown-4 Sep\'25', 'Sep25(DT4) ↗', 'U40', 'U40'), ('Dec25', 'Dec25 ↗', 'V40', 'W40')]
ws_dash.row_dimensions[40].height = 20
for s_name, lbl, c_start, c_end in hist_25_tiles:
    ws_dash.merge_cells(f"{c_start}:{c_end}")
    c_idx = column_index_from_string(c_start[:1])
    r_idx = int(c_start[1:])
    s_escaped = s_name.replace("'", "''")
    style_dash_cell(ws_dash, r_idx, c_idx, lbl, font=FONT_LINK_HIST, fill=FILL_SOFT_BLUE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, hyperlink_loc=f"'{s_escaped}'!A1")
    style_dash_range(ws_dash, r_idx, c_idx, r_idx, column_index_from_string(c_end[:1]), fill=FILL_SOFT_BLUE, border=BLACK_BORDER)

# Row 41: 2024 Key Archive Tiles
hist_24_tiles = [('Feb\'24', 'Feb24 ↗', 'N41', 'O41'), ('May\'24', 'May24 ↗', 'P41', 'Q41'), ('Jul\'24', 'Jul24 ↗', 'R41', 'S41'), ('Sep24', 'Sep24 ↗', 'T41', 'T41'), ('Nov24', 'Nov24 ↗', 'U41', 'U41'), ('Dec24', 'Dec24 ↗', 'V41', 'W41')]
ws_dash.row_dimensions[41].height = 20
for s_name, lbl, c_start, c_end in hist_24_tiles:
    ws_dash.merge_cells(f"{c_start}:{c_end}")
    c_idx = column_index_from_string(c_start[:1])
    r_idx = int(c_start[1:])
    s_escaped = s_name.replace("'", "''")
    style_dash_cell(ws_dash, r_idx, c_idx, lbl, font=FONT_LINK_HIST, fill=FILL_SOFT_BLUE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, hyperlink_loc=f"'{s_escaped}'!A1")
    style_dash_range(ws_dash, r_idx, c_idx, r_idx, column_index_from_string(c_end[:1]), fill=FILL_SOFT_BLUE, border=BLACK_BORDER)

# Row 42-44: Quick User Guide
ws_dash.merge_cells('N42:W44')
help_box = "💡 QUICK USER GUIDE:\n• Click any blue underlined link (↗) to instantly jump to that month or the Master Tracker.\n• In any monthly sheet or Master Tracker, click '🏠 ⮌ RETURN TO DASHBOARD' in cell A1 to return here.\n• Water bottles, tissues & supplies are classified under 'Request'; Meetings & Sessions are 'Event'.\n• Bookings with PAX=0 are smart-classified (e.g. Townhall/Cafeteria -> Large Event, else Pending PAX)."
style_dash_cell(ws_dash, 42, 14, help_box, font=Font(name=FONT_FAMILY, size=9, bold=False, color='1B365D'), fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
style_dash_range(ws_dash, 42, 14, 44, 23, fill=FILL_TOTAL_ROW, border=BLACK_BORDER)

ws_dash.row_dimensions[45].height = 12

# =============================================================================
# 4. ADD SPACE VENUE & EVENT CATEGORY TABLES (ROWS 46 TO 59)
# =============================================================================
# Section Headers (Row 46)
ws_dash.row_dimensions[46].height = 26
ws_dash.merge_cells('A46:K46')
style_dash_cell(ws_dash, 46, 1, "🏛️ SPACE & VENUE UTILIZATION SUMMARY (DT-3 & DT-4 2026)", font=FONT_SECTION_HEADER, fill=FILL_ROYAL_BLUE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 46, 1, 46, 11, fill=FILL_ROYAL_BLUE, border=BLACK_BORDER)

ws_dash.merge_cells('M46:W46')
style_dash_cell(ws_dash, 46, 13, "🎯 EVENT CATEGORY & PURPOSE BREAKDOWN (DT-3 & DT-4 2026)", font=FONT_SECTION_HEADER, fill=NAVY_DARK_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
style_dash_range(ws_dash, 46, 13, 46, 23, fill=NAVY_DARK_FILL, border=BLACK_BORDER)

# Table Headers (Row 47)
ws_dash.row_dimensions[47].height = 24

venue_headers = ['Space / Venue', 'Total', '🎪 Events', '📋 Requests', 'Large (≥30)', 'Small (<30)', '⏳ Pend.PAX', '🟢 Conf.', '🟡 Pend.', '🔴 Canc.', 'Total Pax']
for idx, h in enumerate(venue_headers, 1):
    style_dash_cell(ws_dash, 47, idx, h, font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)

cat_headers = ['Event Category', 'Total', '🎪 Events', '📋 Requests', 'Large (≥30)', 'Small (<30)', '⏳ Pend.PAX', '🟢 Conf.', '🟡 Pend.', '🔴 Canc.', 'Total Pax']
for idx, h in enumerate(cat_headers, 13):
    style_dash_cell(ws_dash, 47, idx, h, font=FONT_TABLE_HEADER, fill=FILL_NAVY_SLATE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)

# Venue Definition List: (Label, CriteriaPattern)
venue_rows = [
    ('Breakout Areas & Zones', '*Breakout*'),
    ('Innovation Hub', '*Innovation*'),
    ('Social Hub Areas', '*Social Hub*'),
    ('Hackable Space / Area', '*Hackable*'),
    ('Cafeteria & Dining', '*Cafeteria*'),
    ('Meeting Rooms (305/306)', '*305*'),
    ('Level 1 Spaces (L-1)', '*L-1*'),
    ('Level 4 Spaces (L-4)', '*L-4*'),
    ('Level 5 Spaces (L-5)', '*L-5*'),
    ('Level 6 Spaces (L-6)', '*L-6*'),
    ('Other / General Venues', 'Other')
]

for idx, (v_lbl, v_crit) in enumerate(venue_rows):
    r = 48 + idx
    ws_dash.row_dimensions[r].height = 20
    style_dash_cell(ws_dash, r, 1, v_lbl, font=BOLD_FONT, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
    
    if v_crit == 'Other':
        style_dash_cell(ws_dash, r, 2, f'=B44-SUM(B48:B{r-1})', font=BOLD_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 3, f'=C44-SUM(C48:C{r-1})', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 4, f'=D44-SUM(D48:D{r-1})', font=Font(name=FONT_FAMILY, size=9, bold=True, color='856404'), fill=FILL_SOFT_ORANGE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 5, f'=E44-SUM(E48:E{r-1})', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 6, f'=F44-SUM(F48:F{r-1})', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 7, f'=G44-SUM(G48:G{r-1})', font=Font(name=FONT_FAMILY, size=9, bold=True, color='7D6608'), fill=FILL_SOFT_YELLOW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 8, f'=H44-SUM(H48:H{r-1})', font=GREEN_FONT, fill=GREEN_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 9, f'=I44-SUM(I48:I{r-1})', font=YELLOW_FONT, fill=YELLOW_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 10, f'=J44-SUM(J48:J{r-1})', font=RED_FONT, fill=RED_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 11, f'=L44-SUM(K48:K{r-1})', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    else:
        style_dash_cell(ws_dash, r, 2, f'=COUNTIF(Master_Event_Tracker_2026!$H$3:$H$5000, "{v_crit}")', font=BOLD_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 3, f'=COUNTIFS(Master_Event_Tracker_2026!$H$3:$H$5000, "{v_crit}", Master_Event_Tracker_2026!$I$3:$I$5000, "Event")', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 4, f'=COUNTIFS(Master_Event_Tracker_2026!$H$3:$H$5000, "{v_crit}", Master_Event_Tracker_2026!$I$3:$I$5000, "Request")', font=Font(name=FONT_FAMILY, size=9, bold=True, color='856404'), fill=FILL_SOFT_ORANGE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 5, f'=COUNTIFS(Master_Event_Tracker_2026!$H$3:$H$5000, "{v_crit}", Master_Event_Tracker_2026!$J$3:$J$5000, "Large Event")', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 6, f'=COUNTIFS(Master_Event_Tracker_2026!$H$3:$H$5000, "{v_crit}", Master_Event_Tracker_2026!$J$3:$J$5000, "Small Event")', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 7, f'=COUNTIFS(Master_Event_Tracker_2026!$H$3:$H$5000, "{v_crit}", Master_Event_Tracker_2026!$J$3:$J$5000, "Pending PAX")', font=Font(name=FONT_FAMILY, size=9, bold=True, color='7D6608'), fill=FILL_SOFT_YELLOW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 8, f'=COUNTIFS(Master_Event_Tracker_2026!$H$3:$H$5000, "{v_crit}", Master_Event_Tracker_2026!$T$3:$T$5000, "Confirmed (Done)")', font=GREEN_FONT, fill=GREEN_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 9, f'=COUNTIFS(Master_Event_Tracker_2026!$H$3:$H$5000, "{v_crit}", Master_Event_Tracker_2026!$T$3:$T$5000, "Pending Calendar Booking")', font=YELLOW_FONT, fill=YELLOW_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 10, f'=COUNTIFS(Master_Event_Tracker_2026!$H$3:$H$5000, "{v_crit}", Master_Event_Tracker_2026!$T$3:$T$5000, "Canceled")', font=RED_FONT, fill=RED_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 11, f'=SUMIFS(Master_Event_Tracker_2026!$M$3:$M$5000, Master_Event_Tracker_2026!$H$3:$H$5000, "{v_crit}")', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

# Venue Subtotal Row (Row 59 in Cols A to K)
ws_dash.row_dimensions[59].height = 22
style_dash_cell(ws_dash, 59, 1, "Total Venue Utilization", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
for c in range(2, 12):
    c_let = get_column_letter(c)
    style_dash_cell(ws_dash, 59, c, f"=SUM({c_let}48:{c_let}58)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

# Category Definition List: Label
category_rows = [
    'Team Meeting & Workshop',
    'Training & Enablement',
    'Townhall & All-Hands',
    'Assessment & Hiring',
    'Social & Team Connect',
    'Induction & Onboarding',
    'Tech & Innovation',
    'Leadership & VIP Visit',
    'Facility & Hospitality Request',
    'Dry Run & Rehearsal',
    'Other / Uncategorized'
]

for idx, c_lbl in enumerate(category_rows):
    r = 48 + idx
    style_dash_cell(ws_dash, r, 13, c_lbl, font=BOLD_FONT, border=BLACK_BORDER, align=ALIGN_LEFT_WRAP)
    
    if c_lbl == 'Other / Uncategorized':
        style_dash_cell(ws_dash, r, 14, f'=B44-SUM(N48:N{r-1})', font=BOLD_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 15, f'=C44-SUM(O48:O{r-1})', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 16, f'=D44-SUM(P48:P{r-1})', font=Font(name=FONT_FAMILY, size=9, bold=True, color='856404'), fill=FILL_SOFT_ORANGE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 17, f'=E44-SUM(Q48:Q{r-1})', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 18, f'=F44-SUM(R48:R{r-1})', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 19, f'=G44-SUM(S48:S{r-1})', font=Font(name=FONT_FAMILY, size=9, bold=True, color='7D6608'), fill=FILL_SOFT_YELLOW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 20, f'=H44-SUM(T48:T{r-1})', font=GREEN_FONT, fill=GREEN_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 21, f'=I44-SUM(U48:U{r-1})', font=YELLOW_FONT, fill=YELLOW_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 22, f'=J44-SUM(V48:V{r-1})', font=RED_FONT, fill=RED_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 23, f'=L44-SUM(W48:W{r-1})', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
    else:
        style_dash_cell(ws_dash, r, 14, f'=COUNTIF(Master_Event_Tracker_2026!$L$3:$L$5000, "{c_lbl}")', font=BOLD_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 15, f'=COUNTIFS(Master_Event_Tracker_2026!$L$3:$L$5000, "{c_lbl}", Master_Event_Tracker_2026!$I$3:$I$5000, "Event")', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 16, f'=COUNTIFS(Master_Event_Tracker_2026!$L$3:$L$5000, "{c_lbl}", Master_Event_Tracker_2026!$I$3:$I$5000, "Request")', font=Font(name=FONT_FAMILY, size=9, bold=True, color='856404'), fill=FILL_SOFT_ORANGE, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 17, f'=COUNTIFS(Master_Event_Tracker_2026!$L$3:$L$5000, "{c_lbl}", Master_Event_Tracker_2026!$J$3:$J$5000, "Large Event")', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 18, f'=COUNTIFS(Master_Event_Tracker_2026!$L$3:$L$5000, "{c_lbl}", Master_Event_Tracker_2026!$J$3:$J$5000, "Small Event")', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 19, f'=COUNTIFS(Master_Event_Tracker_2026!$L$3:$L$5000, "{c_lbl}", Master_Event_Tracker_2026!$J$3:$J$5000, "Pending PAX")', font=Font(name=FONT_FAMILY, size=9, bold=True, color='7D6608'), fill=FILL_SOFT_YELLOW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 20, f'=COUNTIFS(Master_Event_Tracker_2026!$L$3:$L$5000, "{c_lbl}", Master_Event_Tracker_2026!$T$3:$T$5000, "Confirmed (Done)")', font=GREEN_FONT, fill=GREEN_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 21, f'=COUNTIFS(Master_Event_Tracker_2026!$L$3:$L$5000, "{c_lbl}", Master_Event_Tracker_2026!$T$3:$T$5000, "Pending Calendar Booking")', font=YELLOW_FONT, fill=YELLOW_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 22, f'=COUNTIFS(Master_Event_Tracker_2026!$L$3:$L$5000, "{c_lbl}", Master_Event_Tracker_2026!$T$3:$T$5000, "Canceled")', font=RED_FONT, fill=RED_FILL, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')
        style_dash_cell(ws_dash, r, 23, f'=SUMIFS(Master_Event_Tracker_2026!$M$3:$M$5000, Master_Event_Tracker_2026!$L$3:$L$5000, "{c_lbl}")', font=REGULAR_FONT, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

# Category Subtotal Row (Row 59 in Cols M to W)
style_dash_cell(ws_dash, 59, 13, "Total All Categories", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP)
for c in range(14, 24):
    c_let = get_column_letter(c)
    style_dash_cell(ws_dash, 59, c, f"=SUM({c_let}48:{c_let}58)", font=FONT_BOLD_NAVY, fill=FILL_TOTAL_ROW, border=BLACK_BORDER, align=ALIGN_CENTER_WRAP, num_format='#,##0')

# Column Widths for Dashboard
dash_col_widths = {
    'A': 22.0, 'B': 12.0, 'C': 12.0, 'D': 13.0, 'E': 13.0, 'F': 13.0, 'G': 13.0, 'H': 12.0, 'I': 12.0, 'J': 12.0, 'K': 13.0,
    'L': 3.0,
    'M': 26.0, 'N': 12.0, 'O': 12.0, 'P': 13.0, 'Q': 13.0, 'R': 13.0, 'S': 13.0, 'T': 12.0, 'U': 12.0, 'V': 12.0, 'W': 13.0
}
for col_l, w in dash_col_widths.items():
    ws_dash.column_dimensions[col_l].width = w

if hasattr(ws_dash, 'views') and ws_dash.views and ws_dash.views.sheetView:
    ws_dash.views.sheetView[0].showGridLines = True
elif hasattr(ws_dash, 'sheet_view'):
    ws_dash.sheet_view.showGridLines = True

print(f"Saving updated duplicate workbook to: {OUTPUT_FILE} ...")
saved = False
for attempt in range(5):
    try:
        wb.save(OUTPUT_FILE)
        print(f"✅ DUPLICATE WORKBOOK UPDATED SUCCESSFULLY: {OUTPUT_FILE}")
        saved = True
        break
    except PermissionError:
        print(f" [RETRY {attempt+1}/5] '{OUTPUT_FILE}' is open in Excel. Waiting 2 seconds...")
        time.sleep(2)
    except Exception as e:
        print(f" Error saving: {e}")
        break

if not saved:
    alt_out = "Yearly Event Tracker- DT-3 and DT-4 Simplified Master_New.xlsx"
    wb.save(alt_out)
    print(f" Saved to alternative file: {alt_out}")
